import io
import uuid
from datetime import date, datetime
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_, and_
from decimal import Decimal

from app.api import deps
from app.models.attendance import LeaveApplication, LeaveType, LeaveStatus, LeaveEncashment
from app.models.employee import Employee
from app.models.organization import Organization
from app.schemas.leave import LeavePayrollExportRequest

router = APIRouter()

@router.post("/payroll-export", dependencies=[Depends(deps.check_permission("55"))])
def export_leave_for_payroll(
    request: LeavePayrollExportRequest,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Export leave and encashment data for payroll processing as an Excel file.
    Summarizes approved leaves and encashments within the date range.
    """
    # 1. Fetch all active employees
    employees = db.query(Employee).filter(
        Employee.organization_id == current_org.id,
        Employee.is_deleted == False
    ).all()
    
    emp_ids = [e.id for e in employees]
    
    # 2. Fetch Approved Leave Applications overlapping with the period
    leaves = db.query(LeaveApplication).filter(
        LeaveApplication.organization_id == current_org.id,
        LeaveApplication.status == LeaveStatus.APPROVED,
        or_(
            and_(LeaveApplication.from_date <= request.to_date, LeaveApplication.to_date >= request.from_date)
        )
    ).options(joinedload(LeaveApplication.leave_type)).all()
    
    # Get all unique leave names for headers
    all_leave_names = sorted(list(set(leave.leave_type.leave_name for leave in leaves)))
    
    # 3. Fetch Approved Encashments within the period
    encashments = []
    if request.include_encashments:
        encashments = db.query(LeaveEncashment).filter(
            LeaveEncashment.organization_id == current_org.id,
            LeaveEncashment.status == "approved",
            LeaveEncashment.encashment_date >= request.from_date,
            LeaveEncashment.encashment_date <= request.to_date
        ).all()
        
    # 4. Group data by employee
    from collections import defaultdict
    emp_leaves = defaultdict(list)
    for leave in leaves:
        emp_leaves[leave.employee_id].append(leave)
        
    emp_encashments = defaultdict(list)
    for enc in encashments:
        emp_encashments[enc.employee_id].append(enc)
        
    # 5. Generate Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Leave Payroll Export"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center")
    
    # Headers
    headers = ["Employee Code", "Employee Name"] + all_leave_names + [
        "Total Paid Days", "Total Unpaid Days", "Encashment Days", "Encashment Amount"
    ]
    
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    # 6. Process each employee
    for emp in employees:
        e_leaves = emp_leaves.get(emp.id, [])
        e_encs = emp_encashments.get(emp.id, [])
        
        summaries = defaultdict(lambda: {"days": 0.0, "is_paid": True})
        total_paid = 0.0
        total_unpaid = 0.0
        
        for leave in e_leaves:
            overlap_start = max(leave.from_date, request.from_date)
            overlap_end = min(leave.to_date, request.to_date)
            overlap_days = (overlap_end - overlap_start).days + 1
            if overlap_days < 0: overlap_days = 0
            
            days = float(overlap_days)
            summaries[leave.leave_type.leave_name]["days"] += days
            summaries[leave.leave_type.leave_name]["is_paid"] = leave.leave_type.is_paid
            
            if leave.leave_type.is_paid:
                total_paid += days
            else:
                total_unpaid += days
                
        total_enc_days = sum(float(enc.encashment_days) for enc in e_encs)
        total_enc_amount = sum(float(enc.encashment_amount) for enc in e_encs)
        
        # Prepare row
        row = [emp.employee_code, f"{emp.first_name} {emp.last_name}"]
        for name in all_leave_names:
            row.append(summaries[name]["days"])
        
        row.extend([total_paid, total_unpaid, total_enc_days, total_enc_amount])
        ws.append(row)
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 3

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"leave_payroll_{request.from_date}_to_{request.to_date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

