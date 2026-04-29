import uuid
from typing import List, Optional, Union
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_, func
from datetime import datetime, date, timedelta
from decimal import Decimal

from app.api import deps
from app.models.attendance import (
    AttendanceLog, AttendanceRecord, AttendanceStatus, 
    CheckType, AttendanceSource, ShiftMaster, ShiftRoster,
    AttendanceRegularization, RegularizationStatus,
    OvertimeRequest, OvertimeStatus, BiometricDevice
)
from app.models.employee import Employee, Department, Location, EmploymentStatus
from app.models.organization import Organization
from app.schemas.attendance import (
    AttendanceCheckIn, AttendanceCheckInResponse,
    AttendanceCheckOut, AttendanceCheckOutResponse,
    AttendanceBreak, AttendanceBreakResponse, AttendanceBreakEndResponse,
    AttendanceCurrentStatusResponse, AttendanceRecordListResponse,
    AttendanceRecordResponse, ManualAttendanceCreate, AttendanceRecordUpdate,
    EmployeeAttendanceResponse, AttendanceSummaryResponse,
    AttendanceDashboardResponse, AttendanceLogListResponse,
    AttendanceLogProcessRequest, AttendanceLogProcessResponse,
    AttendanceSyncRequest, AttendanceSyncResponse,
    AttendanceRegularizationListResponse, AttendanceRegularizationResponse,
    AttendanceRegularizationCreate, AttendanceRegularizationApproval,
    AttendanceRegularizationRejection,
    OvertimeRequestListResponse, OvertimeRequestResponse,
    AttendanceImportResponse, AttendanceImportError,
    PayrollExportRequest, PayrollExportResponse, EmployeePayrollAttendance,
    BiometricDeviceListResponse, BiometricDeviceCreate, BiometricDeviceUpdate,
    BiometricDeviceResponse, BiometricDeviceStatusResponse
)
import csv
import io

from app.core.permissions import AttendancePolicyPermissions

router = APIRouter()

@router.post("/check-in", response_model=AttendanceCheckInResponse)
def check_in(
    *,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    check_in_in: AttendanceCheckIn
):
    """
    Clock-in attendance for an employee.
    """
    # Ownership Check: Employees can only check-in for themselves.
    if isinstance(current_user, Employee):
        if str(current_user.uuid) != str(check_in_in.employee_uuid):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only check-in for yourself."
            )
        current_org_id = current_user.organization_id
    else:
        # For Organization logins
        current_org_id = current_user.id

    # 1. Validate employee
    employee = db.query(Employee).filter(
        Employee.uuid == check_in_in.employee_uuid,
        Employee.organization_id == current_org_id
    ).first()
    
    if not employee:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"success": False, "message": "Employee not found", "data": None}
        )

    # Use provided timestamp or server UTC time
    punch_time = (check_in_in.timestamp or datetime.utcnow()).replace(tzinfo=None)
    punch_date = punch_time.date()

    # 2. Create Attendance Log (Raw Punch)
    attendance_log = AttendanceLog(
        organization_id=current_org_id,
        employee_id=employee.id,
        punch_time=punch_time,
        punch_date=punch_date,
        check_type=CheckType.CHECK_IN,
        source=check_in_in.source,
        device_id=check_in_in.device_id,
        latitude=check_in_in.latitude,
        longitude=check_in_in.longitude,
        location=check_in_in.location_name
    )
    db.add(attendance_log)

    # 3. Handle Daily Attendance Record
    attendance_record = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id == employee.id,
        AttendanceRecord.attendance_date == punch_date,
        AttendanceRecord.organization_id == current_org_id
    ).first()

    if not attendance_record:
        # Resolve Shift for this day
        # Priority: 1. Published Roster
        #           2. Future implementation: Default Organization Shift
        
        roster = db.query(ShiftRoster).filter(
            ShiftRoster.employee_id == employee.id,
            ShiftRoster.roster_date == punch_date,
            ShiftRoster.is_published == True,
            ShiftRoster.is_deleted == False
        ).options(joinedload(ShiftRoster.shift)).first()
        
        shift_id = roster.shift_id if roster else None
        shift_start = roster.shift.start_time if roster else None
        shift_end = roster.shift.end_time if roster else None
        
        attendance_record = AttendanceRecord(
            organization_id=current_org_id,
            employee_id=employee.id,
            attendance_date=punch_date,
            shift_id=shift_id,
            shift_start_time=shift_start,
            shift_end_time=shift_end,
            first_check_in=punch_time,
            status=AttendanceStatus.PRESENT,
            check_in_latitude=check_in_in.latitude,
            check_in_longitude=check_in_in.longitude,
            check_in_location=check_in_in.location_name,
            check_in_source=check_in_in.source,
            check_in_device_id=check_in_in.device_id
        )
        db.add(attendance_record)
    else:
        # Update existing record if first_check_in is missing (covers cases like Absent -> Present)
        if not attendance_record.first_check_in:
            attendance_record.first_check_in = punch_time
            attendance_record.check_in_latitude = check_in_in.latitude
            attendance_record.check_in_longitude = check_in_in.longitude
            attendance_record.check_in_location = check_in_in.location_name
            attendance_record.check_in_source = check_in_in.source
            attendance_record.check_in_device_id = check_in_in.device_id
            
            # Update status if it was absent/blank
            if attendance_record.status == AttendanceStatus.ABSENT:
                attendance_record.status = AttendanceStatus.PRESENT
            db.add(attendance_record)

    # 4. Synchronize with Shift Roster (mirror actual times)
    roster = db.query(ShiftRoster).filter(
        ShiftRoster.employee_id == employee.id,
        ShiftRoster.roster_date == punch_date,
        ShiftRoster.is_deleted == False
    ).first()
    if roster and not roster.actual_start_time:
        roster.actual_start_time = punch_time.time()
        db.add(roster)

    try:
        db.commit()
        db.refresh(attendance_log)
    except Exception as e:
        db.rollback()
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Error during check-in: {str(e)}", "data": None}
        )

    return AttendanceCheckInResponse(
        success=True,
        message=f"Checked in successfully at {punch_time.strftime('%H:%M:%S')}",
        data=attendance_log
    )

@router.post("/check-out", response_model=AttendanceCheckOutResponse)
def check_out(
    *,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    check_out_in: AttendanceCheckOut
):
    """
    Clock-out attendance for an employee.
    Calculates working hours.
    """
    # Ownership Check
    if isinstance(current_user, Employee):
        if str(current_user.uuid) != str(check_out_in.employee_uuid):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only check-out for yourself."
            )
        current_org_id = current_user.organization_id
    else:
        current_org_id = current_user.id

    # 1. Validate employee
    employee = db.query(Employee).filter(
        Employee.uuid == check_out_in.employee_uuid,
        Employee.organization_id == current_org_id
    ).first()
    
    if not employee:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"success": False, "message": "Employee not found", "data": None}
        )

    # Use provided timestamp or server UTC time
    punch_time = (check_out_in.timestamp or datetime.utcnow()).replace(tzinfo=None)
    punch_date = punch_time.date()

    # 2. Create Attendance Log (Raw Punch)
    attendance_log = AttendanceLog(
        organization_id=current_org_id,
        employee_id=employee.id,
        punch_time=punch_time,
        punch_date=punch_date,
        check_type=CheckType.CHECK_OUT,
        source=check_out_in.source,
        device_id=check_out_in.device_id,
        latitude=check_out_in.latitude,
        longitude=check_out_in.longitude,
        location=check_out_in.location_name
    )
    db.add(attendance_log)

    # 3. Handle Daily Attendance Record
    attendance_record = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id == employee.id,
        AttendanceRecord.attendance_date == punch_date,
        AttendanceRecord.organization_id == current_org_id
    ).first()

    # Night Shift Logic: If no record for today, check for an unclosed record from yesterday
    if not attendance_record or not attendance_record.first_check_in:
        yesterday_date = punch_date - timedelta(days=1)
        yesterday_record = db.query(AttendanceRecord).filter(
            AttendanceRecord.employee_id == employee.id,
            AttendanceRecord.attendance_date == yesterday_date,
            AttendanceRecord.organization_id == current_org_id,
            AttendanceRecord.last_check_out == None
        ).first()
        
        if yesterday_record:
            attendance_record = yesterday_record

    if not attendance_record:
        # Checking out without checking in? We create the record but it might be incomplete
        attendance_record = AttendanceRecord(
            organization_id=current_org_id,
            employee_id=employee.id,
            attendance_date=punch_date,
            last_check_out=punch_time,
            status=AttendanceStatus.PRESENT,
            check_out_latitude=check_out_in.latitude,
            check_out_longitude=check_out_in.longitude,
            check_out_location=check_out_in.location_name,
            check_out_source=check_out_in.source,
            check_out_device_id=check_out_in.device_id
        )
        db.add(attendance_record)
    else:
        # Normal update
        attendance_record.last_check_out = punch_time
        attendance_record.check_out_latitude = check_out_in.latitude
        attendance_record.check_out_longitude = check_out_in.longitude
        attendance_record.check_out_location = check_out_in.location_name
        attendance_record.check_out_source = check_out_in.source
        attendance_record.check_out_device_id = check_out_in.device_id
        
        # Calculate Working Hours
        if attendance_record.first_check_in:
            duration = punch_time - attendance_record.first_check_in
            total_seconds = duration.total_seconds()
            hours = total_seconds / 3600
            
            attendance_record.total_work_hours = round(hours, 2)
            # For now, net work hours = total work hours (breaks logic can be added later)
            attendance_record.net_work_hours = round(hours, 2)

    # 4. Synchronize with Shift Roster (mirror actual times)
    roster = db.query(ShiftRoster).filter(
        ShiftRoster.employee_id == employee.id,
        ShiftRoster.roster_date == attendance_record.attendance_date,
        ShiftRoster.is_deleted == False
    ).first()
    if roster:
        roster.actual_end_time = punch_time.time()
        db.add(roster)

    try:
        db.commit()
        db.refresh(attendance_log)
    except Exception as e:
        db.rollback()
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Error during check-out: {str(e)}", "data": None}
        )

    return AttendanceCheckOutResponse(
        success=True,
        message=f"Checked out successfully at {punch_time.strftime('%H:%M:%S')}",
        data={
            "punch": {
                "uuid": str(attendance_log.uuid),
                "punch_time": attendance_log.punch_time.isoformat(),
            },
            "summary": {
                "total_work_hours": float(attendance_record.total_work_hours or 0),
                "attendance_date": str(attendance_record.attendance_date)
            }
        }
    )

@router.post("/break-start", response_model=AttendanceBreakResponse)
def break_start(
    *,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    break_in: AttendanceBreak
):
    """
    Start a break for an employee.
    """
    # Ownership Check
    if isinstance(current_user, Employee):
        if str(current_user.uuid) != str(break_in.employee_uuid):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only perform this action for yourself."
            )
        current_org_id = current_user.organization_id
    else:
        current_org_id = current_user.id

    # 1. Validate employee
    employee = db.query(Employee).filter(
        Employee.uuid == break_in.employee_uuid,
        Employee.organization_id == current_org_id
    ).first()
    
    if not employee:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"success": False, "message": "Employee not found", "data": None}
        )

    # Use provided timestamp or server UTC time
    punch_time = (break_in.timestamp or datetime.utcnow()).replace(tzinfo=None)
    punch_date = punch_time.date()

    # 2. Find active attendance record (Today or Yesterday for night shifts)
    attendance_record = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id == employee.id,
        AttendanceRecord.attendance_date == punch_date,
        AttendanceRecord.organization_id == current_org_id
    ).first()

    if not attendance_record or not attendance_record.first_check_in:
        yesterday_date = punch_date - timedelta(days=1)
        yesterday_record = db.query(AttendanceRecord).filter(
            AttendanceRecord.employee_id == employee.id,
            AttendanceRecord.attendance_date == yesterday_date,
            AttendanceRecord.organization_id == current_org_id,
            AttendanceRecord.last_check_out == None
        ).first()
        
        if yesterday_record:
            attendance_record = yesterday_record

    if not attendance_record:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "message": "Cannot start break without an active check-in.", "data": None}
        )

    # 3. Create Attendance Log (Raw Punch)
    attendance_log = AttendanceLog(
        organization_id=current_org_id,
        employee_id=employee.id,
        punch_time=punch_time,
        punch_date=punch_date,
        check_type=CheckType.BREAK_START,
        source=break_in.source,
        device_id=break_in.device_id
    )
    db.add(attendance_log)

    try:
        db.commit()
        db.refresh(attendance_log)
    except Exception as e:
        db.rollback()
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Error starting break: {str(e)}", "data": None}
        )

    return AttendanceBreakResponse(
        success=True,
        message=f"Break started at {punch_time.strftime('%H:%M:%S')}",
        data=attendance_log
    )

@router.post("/break-end", response_model=AttendanceBreakEndResponse)
def break_end(
    *,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    break_in: AttendanceBreak
):
    """
    End a break for an employee.
    Calculates break duration and updates the attendance record.
    """
    # Ownership Check
    if isinstance(current_user, Employee):
        if str(current_user.uuid) != str(break_in.employee_uuid):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only perform this action for yourself."
            )
        current_org_id = current_user.organization_id
    else:
        current_org_id = current_user.id

    # 1. Validate employee
    employee = db.query(Employee).filter(
        Employee.uuid == break_in.employee_uuid,
        Employee.organization_id == current_org_id
    ).first()
    
    if not employee:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"success": False, "message": "Employee not found", "data": None}
        )

    # Use provided timestamp or server UTC time
    punch_time = (break_in.timestamp or datetime.utcnow()).replace(tzinfo=None)
    punch_date = punch_time.date()

    # 2. Find active attendance record
    attendance_record = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id == employee.id,
        AttendanceRecord.attendance_date == punch_date,
        AttendanceRecord.organization_id == current_org_id
    ).first()

    # Night Shift Logic
    if not attendance_record or not attendance_record.first_check_in:
        yesterday_date = punch_date - timedelta(days=1)
        yesterday_record = db.query(AttendanceRecord).filter(
            AttendanceRecord.employee_id == employee.id,
            AttendanceRecord.attendance_date == yesterday_date,
            AttendanceRecord.organization_id == current_org_id,
            AttendanceRecord.last_check_out == None
        ).first()
        
        if yesterday_record:
            attendance_record = yesterday_record

    if not attendance_record:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "message": "No active attendance record found.", "data": None}
        )

    # 3. Find the most recent BREAK_START punch to calculate duration
    last_break_start = db.query(AttendanceLog).filter(
        AttendanceLog.employee_id == employee.id,
        AttendanceLog.check_type == CheckType.BREAK_START,
        AttendanceLog.organization_id == current_org_id,
        AttendanceLog.punch_time < punch_time
    ).order_by(AttendanceLog.punch_time.desc()).first()

    if not last_break_start:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "message": "No 'Break Start' record found for this session.", "data": None}
        )

    # 4. Create Attendance Log (Raw Punch)
    attendance_log = AttendanceLog(
        organization_id=current_org_id,
        employee_id=employee.id,
        punch_time=punch_time,
        punch_date=punch_date,
        check_type=CheckType.BREAK_END,
        source=break_in.source,
        device_id=break_in.device_id
    )
    db.add(attendance_log)

    # 5. Calculate break duration and update record
    break_duration_seconds = (punch_time - last_break_start.punch_time).total_seconds()
    break_hours = round(break_duration_seconds / 3600, 2)
    
    # Update total break hours for the day
    current_break_hours = float(attendance_record.break_hours or 0)
    attendance_record.break_hours = current_break_hours + break_hours
    
    # Recalculate net work hours (Total - Breaks)
    total_hours = float(attendance_record.total_work_hours or 0)
    attendance_record.net_work_hours = round(total_hours - float(attendance_record.break_hours), 2)
    
    db.add(attendance_record)

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Error ending break: {str(e)}", "data": None}
        )

    return AttendanceBreakEndResponse(
        success=True,
        message=f"Break ended at {punch_time.strftime('%H:%M:%S')}",
        data={
            "punch": {
                "uuid": str(attendance_log.uuid),
                "punch_time": attendance_log.punch_time.isoformat(),
            },
            "summary": {
                "session_break_hours": break_hours,
                "total_daily_break_hours": float(attendance_record.break_hours),
                "net_work_hours": float(attendance_record.net_work_hours)
            }
        }
    )

@router.get("/current-status", response_model=AttendanceCurrentStatusResponse)
def get_current_status(
    *,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    employee_uuid: uuid.UUID = Query(..., description="Employee UUID")
):
    """
    Get the current attendance status of an employee.
    """
    # Ownership Check / Org Resolution
    if isinstance(current_user, Employee):
        if str(current_user.uuid) != str(employee_uuid):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only access your own attendance status."
            )
        current_org_id = current_user.organization_id
    else:
        current_org_id = current_user.id

    # 1. Resolve employee
    employee = db.query(Employee).filter(
        Employee.uuid == employee_uuid,
        Employee.organization_id == current_org_id
    ).first()
    
    if not employee:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"success": False, "message": "Employee not found", "data": None}
        )

    # 2. Get the last punch
    last_punch = db.query(AttendanceLog).filter(
        AttendanceLog.employee_id == employee.id,
        AttendanceLog.organization_id == current_org_id
    ).order_by(AttendanceLog.punch_time.desc()).first()

    # 3. Determine status
    is_checked_in = False
    is_on_break = False
    last_punch_type = None
    last_punch_time = None

    if last_punch:
        last_punch_type = last_punch.check_type
        last_punch_time = last_punch.punch_time
        
        if last_punch_type == CheckType.CHECK_IN:
            is_checked_in = True
        elif last_punch_type == CheckType.BREAK_START:
            is_checked_in = True
            is_on_break = True
        elif last_punch_type == CheckType.BREAK_END:
            is_checked_in = True
        elif last_punch_type == CheckType.CHECK_OUT:
            is_checked_in = False

    # 4. Get active record
    now_utc = datetime.utcnow()
    current_date_utc = now_utc.date()
    
    # Priority 1: Use the date from the last punch if currently checked in to ensure we find the active session
    search_date = last_punch_time.date() if (is_checked_in and last_punch_time) else current_date_utc
    
    attendance_record = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id == employee.id,
        AttendanceRecord.attendance_date == search_date,
        AttendanceRecord.organization_id == current_org_id
    ).options(joinedload(AttendanceRecord.shift)).first()

    # Priority 2: Night shift look-back (if search_date didn't return a record)
    if not attendance_record:
        yesterday_date = current_date_utc - timedelta(days=1)
        attendance_record = db.query(AttendanceRecord).filter(
            AttendanceRecord.employee_id == employee.id,
            AttendanceRecord.attendance_date == yesterday_date,
            AttendanceRecord.organization_id == current_org_id,
            AttendanceRecord.last_check_out == None
        ).options(joinedload(AttendanceRecord.shift)).first()
        
        if attendance_record:
            search_date = yesterday_date
        else:
            search_date = current_date_utc

    # 5. Build response with real-time calculations
    total_work_hours = float(attendance_record.total_work_hours or 0) if attendance_record else 0
    total_break_hours = float(attendance_record.break_hours or 0) if attendance_record else 0
    net_work_hours = float(attendance_record.net_work_hours or 0) if attendance_record else 0

    if attendance_record:
        # Real-time Total Work Hours (Gross)
        if is_checked_in and attendance_record.first_check_in:
            # If currently checked in (or on break), calculate elapsed time since first check-in
            elapsed_seconds = (now_utc - attendance_record.first_check_in).total_seconds()
            total_work_hours = round(max(0, elapsed_seconds / 3600), 2)
        
        # Real-time Break Hours
        if is_on_break and last_punch_time:
            # If currently on break, add current session to existing (completed) break hours
            session_break_seconds = (now_utc - last_punch_time).total_seconds()
            total_break_hours = round(total_break_hours + (session_break_seconds / 3600), 2)
            
        # Real-time Net Work Hours
        net_work_hours = round(max(0, total_work_hours - total_break_hours), 2)

    status_data = {
        "employee_uuid": employee.uuid,
        "is_checked_in": is_checked_in,
        "is_on_break": is_on_break,
        "last_punch_type": last_punch_type,
        "last_punch_time": last_punch_time,
        "attendance_date": search_date,
        "total_work_hours": total_work_hours,
        "total_break_hours": total_break_hours,
        "net_work_hours": net_work_hours,
        "current_shift": attendance_record.shift if attendance_record and attendance_record.shift else None
    }

    return AttendanceCurrentStatusResponse(
        success=True,
        message="Current status retrieved successfully",
        data=status_data
    )

@router.get("/records", response_model=AttendanceRecordListResponse)
def list_attendance_records(
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    _: bool = Depends(deps.check_permission("33")),
    page: Optional[int] = Query(None, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, description="Items per page"),
    employee_uuid: Optional[uuid.UUID] = Query(None, description="Filter by Employee UUID"),
    department_uuid: Optional[uuid.UUID] = Query(None, description="Filter by Department UUID"),
    from_date: Optional[date] = Query(None, description="Filter from attendance date"),
    to_date: Optional[date] = Query(None, description="Filter to attendance date"),
    status: Optional[AttendanceStatus] = Query(None, description="Filter by Attendance Status"),
    search: Optional[str] = Query(None, description="Search by employee name or code"),
    sort_by: Optional[str] = Query("attendance_date", description="Sort by: employee, attendance_date, net_work_hours, status"),
    order: Optional[str] = Query("desc", regex="^(asc|desc)$", description="Sort order: asc or desc")
):
    """
    List daily attendance records with filtering, advanced sorting, searching, and pagination.
    """
    current_org_id = current_user.id if isinstance(current_user, Organization) else current_user.organization_id

    query = db.query(AttendanceRecord).filter(
        AttendanceRecord.organization_id == current_org_id
    )
    
    # 1. Join Employee upfront if needed for sorting or filtering
    query = query.join(Employee, AttendanceRecord.employee_id == Employee.id)
    
    # 2. Filtering logic
    if department_uuid:
        # Explicit onclause fix for ambiguous join error
        query = query.join(Department, Employee.department_id == Department.id).filter(Department.uuid == department_uuid)
    
    if employee_uuid:
        query = query.filter(Employee.uuid == employee_uuid)
        
    if from_date:
        query = query.filter(AttendanceRecord.attendance_date >= from_date)
    if to_date:
        query = query.filter(AttendanceRecord.attendance_date <= to_date)
    if status:
        query = query.filter(AttendanceRecord.status == status)
        
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Employee.first_name.ilike(search_term),
                Employee.last_name.ilike(search_term),
                Employee.employee_code.ilike(search_term)
            )
        )
        
    # 3. Sorting logic
    sort_map = {
        "employee": [Employee.first_name, Employee.last_name],
        "attendance_date": [AttendanceRecord.attendance_date],
        "net_work_hours": [AttendanceRecord.net_work_hours],
        "status": [AttendanceRecord.status]
    }
    
    fields = sort_map.get(sort_by.lower(), [AttendanceRecord.attendance_date])
    
    for field in fields:
        if order.lower() == "desc":
            query = query.order_by(field.desc())
        else:
            query = query.order_by(field.asc())

    # 4. Optimization: Early loading
    query = query.options(
        joinedload(AttendanceRecord.employee),
        joinedload(AttendanceRecord.shift)
    )
    
    # 5. Pagination
    total_records = query.count()
    
    pagination_data = None
    if page is not None:
        total_pages = (total_records + limit - 1) // limit
        skip = (page - 1) * limit
        records = query.offset(skip).limit(limit).all()
        pagination_data = {
            "total_records": total_records,
            "current_page": page,
            "total_pages": total_pages,
            "page_size": limit
        }
    else:
        records = query.all()
        pagination_data = {
            "total_records": total_records,
            "current_page": 1,
            "total_pages": 1,
            "page_size": total_records if total_records > 0 else 0
        }
        
    return AttendanceRecordListResponse(
        success=True,
        message="Attendance records retrieved successfully",
        data=records,
        pagination=pagination_data
    )

@router.get("/records/{record_uuid}", response_model=AttendanceRecordResponse)
def get_attendance_record(
    record_uuid: uuid.UUID,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    _: bool = Depends(deps.check_permission("33"))
):
    """
    Get details of a specific attendance record.
    """
    current_org_id = current_user.id if isinstance(current_user, Organization) else current_user.organization_id

    record = db.query(AttendanceRecord).filter(
        AttendanceRecord.uuid == record_uuid,
        AttendanceRecord.organization_id == current_org_id
    ).options(
        joinedload(AttendanceRecord.employee),
        joinedload(AttendanceRecord.shift)
    ).first()
    
    if not record:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"success": False, "message": "Attendance record not found", "data": None}
        )
        
    return AttendanceRecordResponse(
        success=True,
        message="Attendance record retrieved successfully",
        data=record
    )

@router.post("/records", response_model=AttendanceRecordResponse)
def create_manual_attendance(
    *,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    manual_in: ManualAttendanceCreate
):
    """
    Manual attendance entry by HR/Admin or Self-Service by Employee.
    """
    current_org_id = current_user.id if isinstance(current_user, Organization) else current_user.organization_id

    # 1. RBAC Check
    is_org = isinstance(current_user, Organization)
    is_authorized_manager = not is_org and deps.has_permission(db, current_user, "34")
    is_self_service = not is_org and current_user.uuid == manual_in.employee_uuid

    if not is_org and not is_authorized_manager and not is_self_service:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not have permission to create this record (Requires code 34 for managing others)"
        )

    # 2. Validate employee
    employee = db.query(Employee).filter(
        Employee.uuid == manual_in.employee_uuid,
        Employee.organization_id == current_org_id
    ).first()
    
    if not employee:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"success": False, "message": "Employee not found", "data": None}
        )

    # 3. Check if record already exists
    existing_record = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id == employee.id,
        AttendanceRecord.attendance_date == manual_in.attendance_date,
        AttendanceRecord.organization_id == current_org_id
    ).first()
    
    if existing_record:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "message": "Attendance record already exists for this date. Use update instead.", "data": None}
        )

    # 4. Resolve Shift (lookup roster)
    roster = db.query(ShiftRoster).filter(
        ShiftRoster.employee_id == employee.id,
        ShiftRoster.roster_date == manual_in.attendance_date,
        ShiftRoster.is_published == True,
        ShiftRoster.is_deleted == False
    ).options(joinedload(ShiftRoster.shift)).first()
    
    shift_id = roster.shift_id if roster else None
    shift_start = roster.shift.start_time if roster else None
    shift_end = roster.shift.end_time if roster else None

    # 5. Calculate hours
    total_hours = 0
    if manual_in.check_out:
        duration = manual_in.check_out - manual_in.check_in
        total_hours = round(duration.total_seconds() / 3600, 2)

    # 6. Create Attendance Record
    attendance_record = AttendanceRecord(
        organization_id=current_org_id,
        employee_id=employee.id,
        attendance_date=manual_in.attendance_date,
        shift_id=shift_id,
        shift_start_time=shift_start,
        shift_end_time=shift_end,
        first_check_in=manual_in.check_in.replace(tzinfo=None),
        last_check_out=manual_in.check_out.replace(tzinfo=None) if manual_in.check_out else None,
        total_work_hours=total_hours,
        net_work_hours=total_hours,
        status=AttendanceStatus.PRESENT,
        is_manual_entry=True,
        manual_entry_reason=manual_in.reason,
        check_in_source=AttendanceSource.MANUAL,
        check_out_source=AttendanceSource.MANUAL if manual_in.check_out else None
    )
    db.add(attendance_record)

    # 7. Create Audit Logs
    check_in_log = AttendanceLog(
        organization_id=current_org_id,
        employee_id=employee.id,
        punch_time=manual_in.check_in.replace(tzinfo=None),
        punch_date=manual_in.attendance_date,
        check_type=CheckType.CHECK_IN,
        source=AttendanceSource.MANUAL,
        location="Manual Entry"
    )
    db.add(check_in_log)
    
    if manual_in.check_out:
        check_out_log = AttendanceLog(
            organization_id=current_org_id,
            employee_id=employee.id,
            punch_time=manual_in.check_out.replace(tzinfo=None),
            punch_date=manual_in.attendance_date,
            check_type=CheckType.CHECK_OUT,
            source=AttendanceSource.MANUAL,
            location="Manual Entry"
        )
        db.add(check_out_log)

    try:
        db.commit()
        db.refresh(attendance_record)
    except Exception as e:
        db.rollback()
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Error creating manual entry: {str(e)}", "data": None}
        )

    return AttendanceRecordResponse(
        success=True,
        message="Manual attendance record created successfully",
        data=attendance_record
    )

@router.put("/records/{record_uuid}", response_model=AttendanceRecordResponse)
def update_attendance_record(
    record_uuid: uuid.UUID,
    record_in: AttendanceRecordUpdate,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    _: bool = Depends(deps.check_permission("35"))
):
    """
    Update an attendance record (e.g., HR correcting times or approving regularization).
    """
    current_org_id = current_user.id if isinstance(current_user, Organization) else current_user.organization_id

    record = db.query(AttendanceRecord).filter(
        AttendanceRecord.uuid == record_uuid,
        AttendanceRecord.organization_id == current_org_id
    ).first()
    
    if not record:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"success": False, "message": "Attendance record not found", "data": None}
        )
        
    update_data = record_in.model_dump(exclude_unset=True)
    
    for field, value in update_data.items():
        if field in ['first_check_in', 'last_check_out'] and value:
            value = value.replace(tzinfo=None)
        setattr(record, field, value)
        
    # Trigger recalculation if times were changed
    if 'first_check_in' in update_data or 'last_check_out' in update_data:
        if record.first_check_in and record.last_check_out:
            duration = record.last_check_out - record.first_check_in
            total_hours = round(duration.total_seconds() / 3600, 2)
            record.total_work_hours = total_hours
            record.net_work_hours = round(total_hours - float(record.break_hours or 0), 2)

    db.add(record)
    
    try:
        db.commit()
        db.refresh(record)
        # Re-fetch with joined relations for the response
        record = db.query(AttendanceRecord).filter(AttendanceRecord.id == record.id).options(
            joinedload(AttendanceRecord.employee),
            joinedload(AttendanceRecord.shift)
        ).first()
    except Exception as e:
        db.rollback()
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Error updating record: {str(e)}", "data": None}
        )
        
    return AttendanceRecordResponse(
        success=True,
        message="Attendance record updated successfully",
        data=record
    )

@router.post("/bulk-import", response_model=AttendanceImportResponse)
def bulk_import_attendance(
    *,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    file: UploadFile = File(...),
    _: bool = Depends(deps.check_permission(AttendancePolicyPermissions.READ))
):
    """
    Bulk import attendance records from an Excel (.xlsx) file.
    Expected columns: employee_code, date, check_in, check_out, status, remarks
    - 'date' should be in YYYY-MM-DD format.
    - 'check_in' and 'check_out' should be in HH:MM format (time) or full ISO datetime.
    """
    rows = []
    file_extension = file.filename.split('.')[-1].lower()
    
    try:
        if file_extension in ['xlsx', 'xls']:
            df = pd.read_excel(file.file)
            # Convert NaN to None for consistent processing
            rows = df.where(pd.notnull(df), None).to_dict(orient='records')
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, 
                detail="Invalid file format. Please upload an Excel (.xlsx or .xls) file."
            )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, 
            detail=f"Error reading Excel file: {str(e)}"
        )
    
    success_count = 0
    errors = []
    
    for row_idx, row in enumerate(rows, start=2):
        try:
            employee_code = row.get('employee_code', '').strip()
            email = row.get('email', '').strip()
            attendance_date_str = row.get('date', '').strip()
            check_in_str = row.get('check_in', '').strip()
            check_out_str = row.get('check_out', '').strip()
            status_str = row.get('status', '').strip().lower()
            remarks = row.get('remarks', '').strip()
            
            if not (employee_code or email) or not attendance_date_str:
                errors.append(AttendanceImportError(
                    row=row_idx, 
                    error="Missing required fields: employee_code/email or date"
                ))
                continue
                
            # Resolve employee
            if employee_code:
                employee = db.query(Employee).filter(
                    Employee.employee_code == employee_code,
                    Employee.organization_id == current_org.id,
                    Employee.is_deleted == False
                ).first()
            else:
                employee = db.query(Employee).filter(
                    Employee.work_email == email,
                    Employee.organization_id == current_org.id,
                    Employee.is_deleted == False
                ).first()
            
            if not employee:
                errors.append(AttendanceImportError(
                    row=row_idx, 
                    error=f"Employee not found with code/email: {employee_code or email}"
                ))
                continue
            
            try:
                attendance_date = datetime.strptime(attendance_date_str, '%Y-%m-%d').date()
            except ValueError:
                errors.append(AttendanceImportError(
                    row=row_idx, 
                    error=f"Invalid date format: {attendance_date_str}. Expected YYYY-MM-DD"
                ))
                continue
            
            # Helper to parse time/datetime
            def parse_stamp(stamp_str, base_date):
                if not stamp_str: return None
                try:
                    # HH:MM format
                    t = datetime.strptime(stamp_str, '%H:%M').time()
                    return datetime.combine(base_date, t)
                except ValueError:
                    try:
                        # HH:MM:SS format
                        t = datetime.strptime(stamp_str, '%H:%M:%S').time()
                        return datetime.combine(base_date, t)
                    except ValueError:
                        try:
                            # Full ISO format
                            return datetime.fromisoformat(stamp_str.replace('Z', '+00:00')).replace(tzinfo=None)
                        except ValueError:
                            return None

            check_in = parse_stamp(check_in_str, attendance_date)
            check_out = parse_stamp(check_out_str, attendance_date)

            # Resolve Shift (lookup roster)
            roster = db.query(ShiftRoster).filter(
                ShiftRoster.employee_id == employee.id,
                ShiftRoster.roster_date == attendance_date,
                ShiftRoster.is_published == True,
                ShiftRoster.is_deleted == False
            ).options(joinedload(ShiftRoster.shift)).first()
            
            shift_id = roster.shift_id if roster else None
            shift_start = roster.shift.start_time if roster else None
            shift_end = roster.shift.end_time if roster else None

            # Map status
            attendance_status = AttendanceStatus.PRESENT
            if status_str:
                try:
                    attendance_status = AttendanceStatus(status_str)
                except ValueError:
                    pass

            total_hours = 0
            if check_in and check_out:
                duration = check_out - check_in
                total_hours = round(duration.total_seconds() / 3600, 2)

            # Create or Update record
            record = db.query(AttendanceRecord).filter(
                AttendanceRecord.employee_id == employee.id,
                AttendanceRecord.attendance_date == attendance_date,
                AttendanceRecord.organization_id == current_org.id
            ).first()

            if not record:
                record = AttendanceRecord(
                    organization_id=current_org.id,
                    employee_id=employee.id,
                    attendance_date=attendance_date,
                    shift_id=shift_id,
                    shift_start_time=shift_start,
                    shift_end_time=shift_end,
                    first_check_in=check_in,
                    last_check_out=check_out,
                    total_work_hours=total_hours,
                    net_work_hours=total_hours,
                    status=attendance_status,
                    remarks=remarks,
                    is_manual_entry=True,
                    manual_entry_reason="Bulk Import"
                )
                db.add(record)
            else:
                record.first_check_in = check_in
                record.last_check_out = check_out
                record.total_work_hours = total_hours
                record.net_work_hours = total_hours
                record.status = attendance_status
                record.remarks = remarks
                record.is_manual_entry = True
                db.add(record)

            success_count += 1
            
        except Exception as e:
            errors.append(AttendanceImportError(row=row_idx, error=f"Unexpected error: {str(e)}"))

    if success_count > 0:
        try:
            db.commit()
        except Exception as e:
            db.rollback()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, 
                detail=f"Database error during commit: {str(e)}"
            )
    
    return AttendanceImportResponse(
        success=True,
        message=f"Import completed. {success_count} records processed.",
        success_count=success_count,
        error_count=len(errors),
        errors=errors
    )

@router.get("/import-template")
def get_attendance_import_template(
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    _: bool = Depends(deps.check_permission(AttendancePolicyPermissions.READ))
):
    """
    Generate an Excel template for bulk attendance import.
    """
    wb = Workbook()
    
    # --- LOOKUPS SHEET (Hidden) ---
    ws_lookup = wb.create_sheet("Lookups")
    ws_lookup.sheet_state = 'hidden'
    
    # 1. Attendance Statuses
    ws_lookup.cell(row=1, column=1, value="Attendance Statuses")
    valid_statuses = [s.value for s in AttendanceStatus]
    for idx, s in enumerate(valid_statuses, start=2):
        ws_lookup.cell(row=idx, column=1, value=s)
        
    # --- MAIN SHEET ---
    ws = wb.active
    ws.title = "Attendance Import"
    
    headers = [
        "employee_code*", "email", "date*", "check_in*", "check_out*", "status*", "remarks"
    ]
    ws.append(headers)
    
    # Styling headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = 20

    # Data Validation for Status
    status_dv = DataValidation(type="list", formula1=f"Lookups!$A$2:$A${len(valid_statuses)+1}")
    ws.add_data_validation(status_dv)
    status_dv.add("F2:F1000")

    # Sample Data
    ws.append(["EMP001", "employee@example.com", datetime.now().strftime('%Y-%m-%d'), "09:00", "18:00", "present", "Bulk import sample"])

    # Help text
    ws['I1'] = "Instructions:"
    ws['I1'].font = Font(bold=True)
    ws['I2'] = "* Mandatory fields"
    ws['I3'] = "Dates must be YYYY-MM-DD"
    ws['I4'] = "Times must be HH:MM (24-hour format)"

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=attendance_import_template.xlsx"}
    )
@router.get("/employees/{employee_uuid}", response_model=EmployeeAttendanceResponse)
def get_employee_attendance(
    employee_uuid: uuid.UUID,
    from_date: date = Query(..., description="Start date"),
    to_date: date = Query(..., description="End date"),
    page: Optional[int] = Query(None, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, description="Items per page"),
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Get attendance records and summary for a specific employee over a period.
    """
    # 1. Ownership Check / Org Resolution
    if isinstance(current_user, Employee):
        if str(current_user.uuid) != str(employee_uuid):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only access your own attendance report."
            )
        current_org_id = current_user.organization_id
    else:
        current_org_id = current_user.id

    # 2. Resolve employee
    employee = db.query(Employee).filter(
        Employee.uuid == employee_uuid,
        Employee.organization_id == current_org_id
    ).first()
    
    if not employee:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"success": False, "message": "Employee not found", "summary": None, "data": []}
        )

    # 3. Base Query
    base_query = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id == employee.id,
        AttendanceRecord.organization_id == current_org_id,
        AttendanceRecord.attendance_date >= from_date,
        AttendanceRecord.attendance_date <= to_date
    )

    # 4. Calculate Summary (Always for the full period requested)
    all_records_for_period = base_query.all()
    
    summary_data = {
        "total_days": (to_date - from_date).days + 1,
        "present_days": 0,
        "absent_days": 0,
        "half_days": 0,
        "leave_days": 0,
        "holiday_days": 0,
        "late_days": 0,
        "early_departure_days": 0,
        "total_work_hours": 0.0,
        "total_break_hours": 0.0,
        "total_net_work_hours": 0.0
    }

    for record in all_records_for_period:
        if record.status == AttendanceStatus.PRESENT:
            summary_data["present_days"] += 1
        elif record.status == AttendanceStatus.ABSENT:
            summary_data["absent_days"] += 1
        elif record.status == AttendanceStatus.HALF_DAY:
            summary_data["half_days"] += 1
        elif record.status == AttendanceStatus.ON_LEAVE:
            summary_data["leave_days"] += 1
        elif record.status == AttendanceStatus.HOLIDAY:
            summary_data["holiday_days"] += 1
        
        if record.is_late:
            summary_data["late_days"] += 1
        if record.is_early_departure:
            summary_data["early_departure_days"] += 1
        
        summary_data["total_work_hours"] += float(record.total_work_hours or 0)
        summary_data["total_break_hours"] += float(record.break_hours or 0)
        summary_data["total_net_work_hours"] += float(record.net_work_hours or 0)

    # Round totals
    summary_data["total_work_hours"] = round(summary_data["total_work_hours"], 2)
    summary_data["total_break_hours"] = round(summary_data["total_break_hours"], 2)
    summary_data["total_net_work_hours"] = round(summary_data["total_net_work_hours"], 2)

    # 4. Apply Pagination to Records
    total_records = len(all_records_for_period)
    records_query = base_query.options(
        joinedload(AttendanceRecord.shift)
    ).order_by(AttendanceRecord.attendance_date.asc())
    
    pagination_data = None
    if page is not None:
        total_pages = (total_records + limit - 1) // limit
        skip = (page - 1) * limit
        records = records_query.offset(skip).limit(limit).all()
        pagination_data = {
            "total_records": total_records,
            "current_page": page,
            "total_pages": total_pages,
            "page_size": limit
        }
    else:
        records = all_records_for_period # Already loaded
        pagination_data = {
            "total_records": total_records,
            "current_page": 1,
            "total_pages": 1,
            "page_size": total_records if total_records > 0 else 0
        }

    return EmployeeAttendanceResponse(
        success=True,
        message=f"Attendance report for {from_date} to {to_date} retrieved successfully",
        summary=summary_data,
        pagination=pagination_data,
        data=records
    )

@router.get("/summary", response_model=AttendanceSummaryResponse)
def get_attendance_summary(
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    _: bool = Depends(deps.check_permission("33")),
    employee_uuid: Optional[uuid.UUID] = Query(None, description="Filter by Employee UUID"),
    department_uuid: Optional[uuid.UUID] = Query(None, description="Filter by Department UUID"),
    from_date: date = Query(..., description="Start date"),
    to_date: date = Query(..., description="End date")
):
    """
    Get aggregate attendance statistics for a period, optionally filtered by employee or department.
    """
    query = db.query(AttendanceRecord).filter(
        AttendanceRecord.organization_id == current_org.id,
        AttendanceRecord.attendance_date >= from_date,
        AttendanceRecord.attendance_date <= to_date
    )
    
    if department_uuid:
        query = query.join(Employee, AttendanceRecord.employee_id == Employee.id).join(Department).filter(Department.uuid == department_uuid)
    
    if employee_uuid:
        if not department_uuid:
            query = query.join(Employee, AttendanceRecord.employee_id == Employee.id)
        query = query.filter(Employee.uuid == employee_uuid)
        
    records = query.all()
    
    summary_data = {
        "total_days": (to_date - from_date).days + 1,
        "present_days": 0,
        "absent_days": 0,
        "half_days": 0,
        "leave_days": 0,
        "holiday_days": 0,
        "late_days": 0,
        "early_departure_days": 0,
        "total_work_hours": 0.0,
        "total_break_hours": 0.0,
        "total_net_work_hours": 0.0
    }

    for record in records:
        if record.status == AttendanceStatus.PRESENT:
            summary_data["present_days"] += 1
        elif record.status == AttendanceStatus.ABSENT:
            summary_data["absent_days"] += 1
        elif record.status == AttendanceStatus.HALF_DAY:
            summary_data["half_days"] += 1
        elif record.status == AttendanceStatus.ON_LEAVE:
            summary_data["leave_days"] += 1
        elif record.status == AttendanceStatus.HOLIDAY:
            summary_data["holiday_days"] += 1
            
        if record.is_late:
            summary_data["late_days"] += 1
        if record.is_early_departure:
            summary_data["early_departure_days"] += 1
            
        summary_data["total_work_hours"] += float(record.total_work_hours or 0)
        summary_data["total_break_hours"] += float(record.break_hours or 0)
        summary_data["total_net_work_hours"] += float(record.net_work_hours or 0)

    # Round totals
    for key in ["total_work_hours", "total_break_hours", "total_net_work_hours"]:
        summary_data[key] = round(summary_data[key], 2)

    return AttendanceSummaryResponse(
        success=True,
        message="Attendance summary retrieved successfully",
        data=summary_data
    )

@router.get("/dashboard", response_model=AttendanceDashboardResponse)
def get_attendance_dashboard(
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    _: bool = Depends(deps.check_permission("33")),
    date_query: date = Query(default=date.today(), alias="date"),
    department_uuid: Optional[uuid.UUID] = Query(None, description="Filter by Department UUID"),
    location_uuid: Optional[uuid.UUID] = Query(None, description="Filter by Location UUID")
):
    """
    Real-time attendance dashboard showing status of all employees for a specific date.
    Includes current status (On Break, Checked In) if the date is today.
    """
    # 1. Base Employee Query
    emp_query = db.query(Employee).filter(
        Employee.organization_id == current_org.id,
        Employee.employment_status == EmploymentStatus.ACTIVE,
        Employee.is_deleted == False
    )
    
    if department_uuid:
        emp_query = emp_query.join(Department, Employee.department_id == Department.id).filter(Department.uuid == department_uuid)
    if location_uuid:
        emp_query = emp_query.join(Location, Employee.location_id == Location.id).filter(Location.uuid == location_uuid)
        
    employees = emp_query.all()
    if not employees:
        return AttendanceDashboardResponse(
            success=True,
            message="No employees found",
            date=date_query,
            summary={"total_employees": 0, "present": 0, "absent": 0, "late": 0, "on_break": 0},
            data=[]
        )
        
    employee_ids = [e.id for e in employees]
    
    # 2. Get Attendance Records for the day
    attendance_records = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id.in_(employee_ids),
        AttendanceRecord.attendance_date == date_query
    ).options(joinedload(AttendanceRecord.shift)).all()
    
    record_map = {r.employee_id: r for r in attendance_records}
    
    # 3. Get Rosters for the day
    rosters = db.query(ShiftRoster).filter(
        ShiftRoster.employee_id.in_(employee_ids),
        ShiftRoster.roster_date == date_query,
        ShiftRoster.is_published == True,
        ShiftRoster.is_deleted == False
    ).options(joinedload(ShiftRoster.shift)).all()
    
    roster_map = {ros.employee_id: ros for ros in rosters}
    
    # 4. Detect "On Break" if today
    latest_punches = {}
    if date_query == date.today():
        # Subquery for latest log id per employee today
        subq = db.query(
            AttendanceLog.employee_id,
            func.max(AttendanceLog.id).label('max_id')
        ).filter(
            AttendanceLog.employee_id.in_(employee_ids),
            AttendanceLog.punch_date == date_query
        ).group_by(AttendanceLog.employee_id).subquery()
        
        punches = db.query(AttendanceLog).join(
            subq, AttendanceLog.id == subq.c.max_id
        ).all()
        latest_punches = {p.employee_id: p for p in punches}

    # 5. Build Response
    dashboard_data = []
    summary = {
        "total_employees": len(employees),
        "present": 0,
        "absent": 0,
        "late": 0,
        "on_break": 0,
        "on_leave": 0
    }
    
    for emp in employees:
        record = record_map.get(emp.id)
        roster = roster_map.get(emp.id)
        latest_punch = latest_punches.get(emp.id)
        
        status = AttendanceStatus.ABSENT
        is_checked_in = False
        is_on_break = False
        
        if record:
            status = record.status
            is_checked_in = record.last_check_out is None and record.first_check_in is not None
            
        # Refine status based on live punch if today
        if latest_punch:
            if latest_punch.check_type == CheckType.BREAK_START:
                is_on_break = True
                is_checked_in = True
            elif latest_punch.check_type == CheckType.BREAK_END:
                is_on_break = False
                is_checked_in = True
            elif latest_punch.check_type == CheckType.CHECK_IN:
                is_checked_in = True
            elif latest_punch.check_type == CheckType.CHECK_OUT:
                is_checked_in = False
                
        # Fill data
        dashboard_data.append({
            "employee_uuid": emp.uuid,
            "employee_name": f"{emp.first_name} {emp.last_name}",
            "employee_code": emp.employee_code,
            "department_name": emp.department.department_name if emp.department else None,
            "location_name": emp.location.location_name if emp.location else None,
            "shift_name": roster.shift.shift_name if roster and roster.shift else (record.shift.shift_name if record and record.shift else None),
            "shift_start_time": roster.shift.start_time if roster and roster.shift else (record.shift_start_time if record else None),
            "shift_end_time": roster.shift.end_time if roster and roster.shift else (record.shift_end_time if record else None),
            "status": status,
            "is_checked_in": is_checked_in,
            "is_on_break": is_on_break,
            "first_check_in": record.first_check_in if record else None,
            "last_check_out": record.last_check_out if record else None,
            "is_late": record.is_late if record else False,
            "late_by_minutes": record.late_by_minutes if record else 0,
            "is_early_departure": record.is_early_departure if record else False,
            "early_departure_minutes": record.early_departure_minutes if record else 0
        })
        
        # Aggregate Summary
        if status in [AttendanceStatus.PRESENT, AttendanceStatus.HALF_DAY]:
            summary["present"] += 1
        elif status == AttendanceStatus.ON_LEAVE:
            summary["on_leave"] += 1
        elif status == AttendanceStatus.ABSENT:
            if roster: # Only count as absent if they were supposed to work
                summary["absent"] += 1
                
        if record and record.is_late:
            summary["late"] += 1
        if is_on_break:
            summary["on_break"] += 1

    return AttendanceDashboardResponse(
        success=True,
        message="Attendance dashboard retrieved successfully",
        date=date_query,
        summary=summary,
        data=dashboard_data
    )

@router.get("/logs", response_model=AttendanceLogListResponse)
def list_attendance_logs(
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    _: bool = Depends(deps.check_permission("33")),
    page: Optional[int] = Query(None, ge=1, description="Page number"),
    limit: int = Query(20, ge=1, description="Items per page"),
    employee_uuid: Optional[uuid.UUID] = Query(None, description="Filter by Employee UUID"),
    from_date: Optional[date] = Query(None, description="Start date"),
    to_date: Optional[date] = Query(None, description="End date"),
    device_id: Optional[str] = Query(None, description="Filter by Device ID"),
    is_processed: Optional[bool] = Query(None, description="Filter by processing status"),
    search: Optional[str] = Query(None, description="Search by employee name or code"),
    sort_by: Optional[str] = Query("Recent", description="Sort by employee, punch_time, punch_date, check_type, source, is_processed, is_valid, Recent, Oldest"),
    order: Optional[str] = Query(None, regex="^(asc|desc)$", description="Sort order (asc or desc)")
):
    """
    List raw attendance punches/logs with filtering, search, and robust sorting.
    """
    query = db.query(AttendanceLog).filter(
        AttendanceLog.organization_id == current_org.id
    )
    
    # 1. Joins and Filtering
    joined_employee = False
    
    if employee_uuid:
        query = query.join(Employee, AttendanceLog.employee_id == Employee.id)
        joined_employee = True
        query = query.filter(Employee.uuid == employee_uuid)
    
    if search:
        if not joined_employee:
            query = query.join(Employee, AttendanceLog.employee_id == Employee.id)
            joined_employee = True
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Employee.first_name.ilike(search_term),
                Employee.last_name.ilike(search_term),
                Employee.employee_code.ilike(search_term),
                (Employee.first_name + " " + Employee.last_name).ilike(search_term)
            )
        )
        
    if from_date:
        query = query.filter(AttendanceLog.punch_date >= from_date)
    if to_date:
        query = query.filter(AttendanceLog.punch_date <= to_date)
        
    if device_id:
        query = query.filter(AttendanceLog.device_id == device_id)
        
    if is_processed is not None:
        query = query.filter(AttendanceLog.is_processed == is_processed)
        
    # 2. Advanced Sorting Logic
    # Normalize sorting parameters for robustness
    sort_by_norm = sort_by if sort_by else "Recent"
    order_norm = order.lower() if order else None

    # Handle alias normalization and default orders
    if sort_by_norm.lower() == "oldest":
        sort_by_norm = "Oldest"
        if not order_norm:
            order_norm = "asc"
    elif sort_by_norm.lower() == "recent":
        sort_by_norm = "Recent"
        if not order_norm:
            order_norm = "desc"
    
    # Default to asc for other fields if not specified
    if not order_norm:
        order_norm = "asc"

    # Sort fields mapping (case-insensitive lookup keys)
    sort_map = {
        "employee": [Employee.first_name, Employee.last_name],
        "punch_time": AttendanceLog.punch_time,
        "check_type": AttendanceLog.check_type,
        "source": AttendanceLog.source,
        "is_processed": AttendanceLog.is_processed,
        "is_valid": AttendanceLog.is_valid,
        "punch_date": AttendanceLog.punch_date,
        "recent": AttendanceLog.punch_time,
        "oldest": AttendanceLog.punch_time
    }

    # Ensure Employee is joined if sorting by name
    if sort_by_norm.lower() == "employee" and not joined_employee:
        query = query.join(Employee, AttendanceLog.employee_id == Employee.id)
        joined_employee = True

    # Apply sorting
    target_key = sort_by_norm.lower()
    if target_key in sort_map:
        fields = sort_map[target_key]
        if not isinstance(fields, list):
            fields = [fields]
        
        # Apply all fields to a single order_by to avoid replacing previous ones
        order_criteria = []
        for field in fields:
            order_criteria.append(field.desc() if order_norm == "desc" else field.asc())
        query = query.order_by(*order_criteria)
    else:
        # Fallback to default
        query = query.order_by(AttendanceLog.punch_time.desc())

    # Optimization: Early loading
    query = query.options(joinedload(AttendanceLog.employee))
    
    total_records = query.count()
    
    pagination_data = None
    if page is not None:
        total_pages = (total_records + limit - 1) // limit
        skip = (page - 1) * limit
        records = query.offset(skip).limit(limit).all()
        pagination_data = {
            "total_records": total_records,
            "current_page": page,
            "total_pages": total_pages,
            "page_size": limit
        }
    else:
        records = query.all()
        pagination_data = {
            "total_records": total_records,
            "current_page": 1,
            "total_pages": 1,
            "page_size": total_records
        }
        
    return AttendanceLogListResponse(
        success=True,
        message="Attendance logs retrieved successfully",
        data=records,
        pagination=pagination_data
    )

@router.post("/logs/process", response_model=AttendanceLogProcessResponse)
def process_attendance_logs(
    request: AttendanceLogProcessRequest,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    _: bool = Depends(deps.check_permission("35"))
):
    """
    Process pending attendance logs and update/create daily attendance records.
    """
    # 1. Identify logs to process
    query = db.query(AttendanceLog).filter(
        AttendanceLog.organization_id == current_org.id,
        AttendanceLog.is_processed == False
    )
    
    if request.log_uuids:
        query = query.filter(AttendanceLog.uuid.in_(request.log_uuids))
    
    if request.employee_uuids:
        employees = db.query(Employee).filter(Employee.uuid.in_(request.employee_uuids)).all()
        emp_ids = [e.id for e in employees]
        query = query.filter(AttendanceLog.employee_id.in_(emp_ids))
        
    if request.from_date:
        query = query.filter(AttendanceLog.punch_date >= request.from_date)
    
    if request.to_date:
        query = query.filter(AttendanceLog.punch_date <= request.to_date)
        
    pending_logs = query.all()
    
    if not pending_logs:
         return AttendanceLogProcessResponse(
             success=True,
             message="No pending logs found to process",
             processed_logs_count=0,
             updated_records_count=0
         )

    # 2. Group logs by (employee_id, punch_date)
    from collections import defaultdict
    grouped_logs = defaultdict(list)
    for log in pending_logs:
        grouped_logs[(log.employee_id, log.punch_date)].append(log)
        
    processed_count = 0
    records_updated = 0
    
    # 3. Process each group
    for (emp_id, day), logs in grouped_logs.items():
        # Get or create record
        record = db.query(AttendanceRecord).filter(
            AttendanceRecord.employee_id == emp_id,
            AttendanceRecord.attendance_date == day
        ).first()
        
        # Need to fetch ALL logs for this day to recalculate accurately (including already processed ones)
        # This ensures we get the true first check-in and last check-out even if they were processed earlier
        all_logs_day = db.query(AttendanceLog).filter(
            AttendanceLog.employee_id == emp_id,
            AttendanceLog.punch_date == day
        ).order_by(AttendanceLog.punch_time).all()
        
        check_ins = [l.punch_time for l in all_logs_day if l.check_type == CheckType.CHECK_IN]
        check_outs = [l.punch_time for l in all_logs_day if l.check_type == CheckType.CHECK_OUT]
        
        first_in = min(check_ins) if check_ins else None
        last_out = max(check_outs) if check_outs else None
        
        # If no record exists, creating one
        if not record:
             # Try to find shift for this day
             shift_roster = db.query(ShiftRoster).filter(
                 ShiftRoster.employee_id == emp_id,
                 ShiftRoster.roster_date == day
             ).options(joinedload(ShiftRoster.shift)).first()
             
             shift_id = shift_roster.shift_id if shift_roster else None
             shift_start = shift_roster.shift.start_time if shift_roster and shift_roster.shift else None
             shift_end = shift_roster.shift.end_time if shift_roster and shift_roster.shift else None
             
             record = AttendanceRecord(
                 organization_id=current_org.id,
                 employee_id=emp_id,
                 attendance_date=day,
                 shift_id=shift_id,
                 shift_start_time=shift_start,
                 shift_end_time=shift_end,
                 status=AttendanceStatus.ABSENT # Pending calculation
             )
             db.add(record)
             db.flush() # Flush to assign an ID
             
        # Update record details
        updated = False
        if first_in and record.first_check_in != first_in:
            record.first_check_in = first_in
            updated = True
        if last_out and record.last_check_out != last_out:
            record.last_check_out = last_out
            updated = True
            
        # Basic Status Update (Can be improved with exact shift logic)
        if record.first_check_in:
             # If they checked in, they are at least Present or Late
             # For now, simplistic status:
             if record.status == AttendanceStatus.ABSENT:
                 record.status = AttendanceStatus.PRESENT
                 updated = True

        # Calculate Work Hours (Simple Duration)
        if record.first_check_in and record.last_check_out:
            duration = (record.last_check_out - record.first_check_in).total_seconds() / 3600.0
            if duration > 0:
                record.total_work_hours = round(duration, 2)
                record.net_work_hours = round(duration, 2) # excluding breaks TODO
                updated = True
        
        if updated:
            records_updated += 1
            
        # Mark pending logs as processed
        for log in logs:
            log.is_processed = True
            log.processed_at = datetime.utcnow()
            log.attendance_record_id = record.id
            processed_count += 1
            
    db.commit()
    
    return AttendanceLogProcessResponse(
        success=True,
        message=f"Processed {processed_count} logs, updated {records_updated} records.",
        processed_logs_count=processed_count,
        updated_records_count=records_updated
    )

@router.post("/logs/sync", response_model=AttendanceSyncResponse)
def sync_attendance_logs(
    request: AttendanceSyncRequest,
    db: Session = Depends(deps.get_db),
    # This endpoint might need a special API Key auth for devices, but using organization auth for now or assuming trusted source
    # In real world, devices usually have their own auth token
    current_org: Organization = Depends(deps.get_current_org),
    _: bool = Depends(deps.check_permission("34"))
):
    """
    Sync attendance logs from biometric devices.
    Expects a list of logs with user_id/employee_code and timestamp.
    """
    synced_count = 0
    errors = []
    
    # Pre-fetch employees for mapping (optimize by fetching only needed ones if list is small, or all active if large)
    # Assuming the device sends 'user_id' which maps to 'employee_code'
    
    incoming_codes = set(str(log.get('user_id')) for log in request.logs if log.get('user_id'))
    employees = db.query(Employee).filter(
        Employee.organization_id == current_org.id,
        Employee.employee_code.in_(incoming_codes)
    ).all()
    
    employee_map = {e.employee_code: e.id for e in employees}
    
    for log_data in request.logs:
        try:
            employee_code = str(log_data.get('user_id'))
            timestamp_str = log_data.get('timestamp')
            
            if not employee_code or not timestamp_str:
                continue
                
            employee_id = employee_map.get(employee_code)
            if not employee_id:
                errors.append(f"Employee not found for code: {employee_code}")
                continue
                
            # Parse timestamp
            try:
                punch_time = datetime.fromisoformat(timestamp_str.replace("Z", "+00:00")).replace(tzinfo=None)
            except ValueError:
                errors.append(f"Invalid timestamp format: {timestamp_str}")
                continue
            
            punch_date = punch_time.date()
            
            # Check for duplicate
            exists = db.query(AttendanceLog).filter(
                AttendanceLog.employee_id == employee_id,
                AttendanceLog.punch_time == punch_time,
                AttendanceLog.device_id == request.device_id
            ).first()
            
            if exists:
                continue
                
            # Determine Check Type (Simple logic, can be enhanced)
            # Devices often send status (0: Check-In, 1: Check-Out, etc.)
            device_status = log_data.get('status') 
            check_type = CheckType.CHECK_IN # Default
            
            if device_status == 1 or str(device_status).lower() in ['out', 'checkout', '1']:
                 check_type = CheckType.CHECK_OUT
            elif device_status == 0 or str(device_status).lower() in ['in', 'checkin', '0']:
                 check_type = CheckType.CHECK_IN
            elif str(device_status).lower() in ['break_in', 'breakin']: # device terminology varies
                 check_type = CheckType.BREAK_START
            elif str(device_status).lower() in ['break_out', 'breakout']:
                 check_type = CheckType.BREAK_END
            
            # Create Attendance Log
            new_log = AttendanceLog(
                organization_id=current_org.id,
                employee_id=employee_id,
                punch_time=punch_time,
                punch_date=punch_date,
                check_type=check_type,
                source=AttendanceSource.BIOMETRIC,
                device_id=request.device_id,
                is_processed=True # Mark as processed as we handle logic immediately below
            )
            db.add(new_log)
            db.flush() 

            # --- Synchronize with AttendanceRecord and ShiftRoster ---
            
            # Resolve AttendanceRecord
            attendance_record = db.query(AttendanceRecord).filter(
                AttendanceRecord.employee_id == employee_id,
                AttendanceRecord.attendance_date == punch_date,
                AttendanceRecord.organization_id == current_org.id
            ).first()

            # For Check-Out and Breaks, handle Night Shift (Check for unclosed record from yesterday)
            if check_type in [CheckType.CHECK_OUT, CheckType.BREAK_START, CheckType.BREAK_END]:
                if not attendance_record or (check_type == CheckType.CHECK_OUT and not attendance_record.first_check_in):
                    yesterday_date = punch_date - timedelta(days=1)
                    yesterday_record = db.query(AttendanceRecord).filter(
                        AttendanceRecord.employee_id == employee_id,
                        AttendanceRecord.attendance_date == yesterday_date,
                        AttendanceRecord.organization_id == current_org.id,
                        AttendanceRecord.last_check_out == None
                    ).first()
                    if yesterday_record:
                        attendance_record = yesterday_record

            # Execute Punch-specific Logic
            if check_type == CheckType.CHECK_IN:
                if not attendance_record:
                    # Resolve Shift from Roster
                    roster = db.query(ShiftRoster).filter(
                        ShiftRoster.employee_id == employee_id,
                        ShiftRoster.roster_date == punch_date,
                        ShiftRoster.is_deleted == False
                    ).first()
                    
                    shift_id = roster.shift_id if roster else None
                    shift_start = roster.shift.start_time if roster and roster.shift else None
                    shift_end = roster.shift.end_time if roster and roster.shift else None
                    
                    attendance_record = AttendanceRecord(
                        organization_id=current_org.id,
                        employee_id=employee_id,
                        attendance_date=punch_date,
                        shift_id=shift_id,
                        shift_start_time=shift_start,
                        shift_end_time=shift_end,
                        first_check_in=punch_time,
                        status=AttendanceStatus.PRESENT,
                        check_in_source=AttendanceSource.BIOMETRIC,
                        check_in_device_id=request.device_id
                    )
                    db.add(attendance_record)
                elif not attendance_record.first_check_in:
                    attendance_record.first_check_in = punch_time
                    attendance_record.check_in_source = AttendanceSource.BIOMETRIC
                    attendance_record.check_in_device_id = request.device_id
                    if attendance_record.status == AttendanceStatus.ABSENT:
                        attendance_record.status = AttendanceStatus.PRESENT
                
                # Sync Shift Roster
                roster = db.query(ShiftRoster).filter(
                    ShiftRoster.employee_id == employee_id,
                    ShiftRoster.roster_date == punch_date,
                    ShiftRoster.is_deleted == False
                ).first()
                if roster and not roster.actual_start_time:
                    roster.actual_start_time = punch_time.time()
                    db.add(roster)

            elif check_type == CheckType.CHECK_OUT:
                if attendance_record:
                    attendance_record.last_check_out = punch_time
                    attendance_record.check_out_source = AttendanceSource.BIOMETRIC
                    attendance_record.check_out_device_id = request.device_id
                    
                    # Calculate Hours
                    if attendance_record.first_check_in:
                        duration = punch_time - attendance_record.first_check_in
                        hours = duration.total_seconds() / 3600
                        attendance_record.total_work_hours = round(Decimal(str(hours)), 2)
                        # Net hours = Total - Break (Break handled in Break End)
                        net = hours - float(attendance_record.break_hours or 0)
                        attendance_record.net_work_hours = round(Decimal(str(max(0, net))), 2)

                    # Sync Shift Roster
                    roster = db.query(ShiftRoster).filter(
                        ShiftRoster.employee_id == employee_id,
                        ShiftRoster.roster_date == attendance_record.attendance_date,
                        ShiftRoster.is_deleted == False
                    ).first()
                    if roster:
                        roster.actual_end_time = punch_time.time()
                        db.add(roster)

            elif check_type == CheckType.BREAK_END:
                if attendance_record:
                    # Find last BREAK_START for this employee
                    last_break_start = db.query(AttendanceLog).filter(
                        AttendanceLog.employee_id == employee_id,
                        AttendanceLog.check_type == CheckType.BREAK_START,
                        AttendanceLog.organization_id == current_org.id,
                        AttendanceLog.punch_time < punch_time
                    ).order_by(AttendanceLog.punch_time.desc()).first()
                    
                    if last_break_start:
                        break_sec = (punch_time - last_break_start.punch_time).total_seconds()
                        break_h = break_sec / 3600
                        current_break = float(attendance_record.break_hours or 0)
                        attendance_record.break_hours = round(Decimal(str(current_break + break_h)), 2)
                        
                        # Recalculate Net Work Hours
                        if attendance_record.total_work_hours:
                            net = float(attendance_record.total_work_hours) - (current_break + break_h)
                            attendance_record.net_work_hours = round(Decimal(str(max(0, net))), 2)

            synced_count += 1
            
        except Exception as e:
            errors.append(f"Error processing log for {employee_code}: {str(e)}")
            
    db.commit()
    
    return AttendanceSyncResponse(
        success=True,
        message=f"Synced {synced_count} logs successfully",
        synced_count=synced_count,
        errors=errors if errors else None
    )

@router.get("/regularizations", response_model=AttendanceRegularizationListResponse)
def list_attendance_regularizations(
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    page: Optional[int] = Query(None, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, description="Items per page"),
    employee_uuid: Optional[uuid.UUID] = Query(None, description="Filter by Employee UUID"),
    approver_uuid: Optional[uuid.UUID] = Query(None, description="Filter by Approver UUID"),
    status: Optional[RegularizationStatus] = Query(None, description="Filter by Status"),
    from_date: Optional[date] = Query(None, description="Filter from attendance date"),
    to_date: Optional[date] = Query(None, description="Filter to attendance date"),
    search: Optional[str] = Query(None, description="Search by employee name or code"),
    sort_by: Optional[str] = Query("attendance_date", description="Sort by: employee, attendance_date, status"),
    order: Optional[str] = Query("desc", regex="^(asc|desc)$", description="Sort order: asc or desc")
):
    """
    List attendance regularization requests with filtering, searching, sorting and pagination.
    """
    current_org_id = current_user.id if isinstance(current_user, Organization) else current_user.organization_id
    
    query = db.query(AttendanceRegularization).join(
        Employee, AttendanceRegularization.employee_id == Employee.id
    ).filter(
        AttendanceRegularization.organization_id == current_org_id,
        AttendanceRegularization.is_deleted == False
    )

    # 0. RBAC / Self-Service
    if not isinstance(current_user, Organization) and not deps.has_permission(db, current_user, "41"):
        # If no permission 41, only show own records
        query = query.filter(AttendanceRegularization.employee_id == current_user.id)
    
    # 1. Joins for filtering if UUIDs provided
    if employee_uuid:
        query = query.filter(Employee.uuid == employee_uuid)
        
    if approver_uuid:
        AppAlias = db.query(Employee).filter(
            Employee.uuid == approver_uuid,
            Employee.organization_id == current_org_id
        ).subquery()
        query = query.filter(AttendanceRegularization.approver_id == AppAlias.c.id)
        
    # 2. Other filters
    if status:
        query = query.filter(AttendanceRegularization.status == status)
    if from_date:
        query = query.filter(AttendanceRegularization.attendance_date >= from_date)
    if to_date:
        query = query.filter(AttendanceRegularization.attendance_date <= to_date)
        
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Employee.first_name.ilike(search_term),
                Employee.last_name.ilike(search_term),
                Employee.employee_code.ilike(search_term)
            )
        )

    # 3. Sorting logic
    sort_map = {
        "employee": [Employee.first_name, Employee.last_name],
        "attendance_date": [AttendanceRegularization.attendance_date],
        "status": [AttendanceRegularization.status]
    }
    
    fields = sort_map.get(sort_by.lower(), [AttendanceRegularization.attendance_date])
    
    for field in fields:
        if order.lower() == "desc":
            query = query.order_by(field.desc())
        else:
            query = query.order_by(field.asc())

    # 4. Optimization: Early loading
    query = query.options(
        joinedload(AttendanceRegularization.employee),
        joinedload(AttendanceRegularization.approver)
    )
    
    # 4. Pagination
    total_records = query.count()
    
    pagination_data = None
    if page is not None:
        total_pages = (total_records + limit - 1) // limit
        skip = (page - 1) * limit
        records = query.offset(skip).limit(limit).all()
        pagination_data = {
            "total_records": total_records,
            "current_page": page,
            "total_pages": total_pages,
            "page_size": limit
        }
    else:
        records = query.all()
        pagination_data = {
            "total_records": total_records,
            "current_page": 1,
            "total_pages": 1,
            "page_size": total_records
        }
        
    return AttendanceRegularizationListResponse(
        success=True,
        message="Regularization requests retrieved successfully",
        data=records,
        pagination=pagination_data
    )

@router.post("/regularizations", response_model=AttendanceRegularizationResponse)
def create_attendance_regularization(
    *,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    regularization_in: AttendanceRegularizationCreate
):
    """
    Create a new attendance regularization request.
    """
    current_org_id = current_user.id if isinstance(current_user, Organization) else current_user.organization_id
    
    # 0. RBAC / Self-Service
    if not isinstance(current_user, Organization) and not deps.has_permission(db, current_user, "39"):
        # Not Org and No Perm 39 -> Can only create for self
        if str(regularization_in.employee_uuid) != str(current_user.uuid):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have permission (code 39) to create regularization for others."
            )

    # 1. Validate employee
    employee = db.query(Employee).filter(
        Employee.uuid == regularization_in.employee_uuid,
        Employee.organization_id == current_org_id
    ).first()
    
    if not employee:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"success": False, "message": "Employee not found", "data": None}
        )

    # 2. Check for existing pending request for the same date
    existing_pending = db.query(AttendanceRegularization).filter(
        AttendanceRegularization.employee_id == employee.id,
        AttendanceRegularization.attendance_date == regularization_in.attendance_date,
        AttendanceRegularization.status == RegularizationStatus.PENDING,
        AttendanceRegularization.organization_id == current_org_id
    ).first()
    
    if existing_pending:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "message": "A pending regularization request already exists for this date", "data": None}
        )

    # 3. Fetch existing attendance record to get original values
    attendance_record = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id == employee.id,
        AttendanceRecord.attendance_date == regularization_in.attendance_date,
        AttendanceRecord.organization_id == current_org_id
    ).first()
    
    original_check_in = attendance_record.first_check_in if attendance_record else None
    original_check_out = attendance_record.last_check_out if attendance_record else None

    # 4. Create Regularization Request
    # Default approver is the reporting manager
    approver_id = employee.reporting_manager_id
    
    regularization = AttendanceRegularization(
        organization_id=current_org_id,
        employee_id=employee.id,
        attendance_date=regularization_in.attendance_date,
        original_check_in=original_check_in,
        original_check_out=original_check_out,
        requested_check_in=regularization_in.requested_check_in,
        requested_check_out=regularization_in.requested_check_out,
        reason=regularization_in.reason,
        reason_category=regularization_in.reason_category,
        attachment_urls=regularization_in.attachment_urls,
        status=RegularizationStatus.PENDING,
        approver_id=approver_id
    )
    
    db.add(regularization)
    db.commit()
    db.refresh(regularization)
    
    return AttendanceRegularizationResponse(
        success=True,
        message="Attendance regularization request created successfully",
        data=regularization
    )

@router.get("/regularizations/{regularization_uuid}", response_model=AttendanceRegularizationResponse)
def get_attendance_regularization(
    regularization_uuid: uuid.UUID,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Get detailed information for a specific attendance regularization request.
    """
    current_org_id = current_user.id if isinstance(current_user, Organization) else current_user.organization_id
    
    regularization = db.query(AttendanceRegularization).filter(
        AttendanceRegularization.uuid == regularization_uuid,
        AttendanceRegularization.organization_id == current_org_id,
        AttendanceRegularization.is_deleted == False
    ).options(
        joinedload(AttendanceRegularization.employee),
        joinedload(AttendanceRegularization.approver)
    ).first()
    
    if not regularization:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"success": False, "message": "Regularization request not found", "data": None}
        )

    # RBAC / Self-Service
    if not isinstance(current_user, Organization) and not deps.has_permission(db, current_user, "41"):
        if regularization.employee_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have permission to view other's regularization requests."
            )
        
    return AttendanceRegularizationResponse(
        success=True,
        message="Regularization request retrieved successfully",
        data=regularization
    )

@router.patch("/regularizations/{regularization_uuid}/approve", response_model=AttendanceRegularizationResponse)
def approve_attendance_regularization(
    regularization_uuid: uuid.UUID,
    approval_in: AttendanceRegularizationApproval,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    _: bool = Depends(deps.check_permission("40"))
):
    """
    Approve an attendance regularization request and update the attendance record.
    """
    current_org_id = current_user.id if isinstance(current_user, Organization) else current_user.organization_id
    # 1. Fetch Request
    regularization = db.query(AttendanceRegularization).filter(
        AttendanceRegularization.uuid == regularization_uuid,
        AttendanceRegularization.organization_id == current_org_id,
        AttendanceRegularization.is_deleted == False
    ).first()
    
    if not regularization:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"success": False, "message": "Regularization request not found", "data": None}
        )
        
    if regularization.status != RegularizationStatus.PENDING:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "message": f"Request is already {regularization.status}", "data": None}
        )

    # 2. Update Request Status
    regularization.status = RegularizationStatus.APPROVED
    regularization.approver_comments = approval_in.comments
    regularization.approved_at = datetime.utcnow()
    
    # 3. Handle Attendance Record Update
    record = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id == regularization.employee_id,
        AttendanceRecord.attendance_date == regularization.attendance_date,
        AttendanceRecord.organization_id == current_org_id
    ).first()
    
    if not record:
        # Create record if missing
        # Try to find shift info
        roster = db.query(ShiftRoster).filter(
            ShiftRoster.employee_id == regularization.employee_id,
            ShiftRoster.roster_date == regularization.attendance_date
        ).options(joinedload(ShiftRoster.shift)).first()
        
        record = AttendanceRecord(
            organization_id=current_org_id,
            employee_id=regularization.employee_id,
            attendance_date=regularization.attendance_date,
            shift_id=roster.shift_id if roster else None,
            shift_start_time=roster.shift.start_time if roster and roster.shift else None,
            shift_end_time=roster.shift.end_time if roster and roster.shift else None,
            status=AttendanceStatus.PRESENT
        )
        db.add(record)
        db.flush()
        
    # Apply requested times
    if regularization.requested_check_in:
        record.first_check_in = regularization.requested_check_in
    if regularization.requested_check_out:
        record.last_check_out = regularization.requested_check_out
        
    record.is_regularized = True
    record.regularization_id = regularization.id
    
    # Update Status if it was absent
    if record.status == AttendanceStatus.ABSENT:
        record.status = AttendanceStatus.PRESENT

    # Recalculate Hours
    if record.first_check_in and record.last_check_out:
        duration = (record.last_check_out - record.first_check_in).total_seconds() / 3600.0
        if duration > 0:
            record.total_work_hours = round(Decimal(str(duration)), 2)
            # For now net hours = total hours (until we incorporate breaks better)
            record.net_work_hours = round(Decimal(str(duration)), 2)
            
    db.commit()
    db.refresh(regularization)
    
    return AttendanceRegularizationResponse(
        success=True,
        message="Regularization request approved and attendance record updated",
        data=regularization
    )

@router.patch("/regularizations/{regularization_uuid}/reject", response_model=AttendanceRegularizationResponse)
def reject_attendance_regularization(
    regularization_uuid: uuid.UUID,
    rejection_in: AttendanceRegularizationRejection,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    _: bool = Depends(deps.check_permission("40"))
):
    """
    Reject an attendance regularization request.
    """
    current_org_id = current_user.id if isinstance(current_user, Organization) else current_user.organization_id
    # 1. Fetch Request
    regularization = db.query(AttendanceRegularization).filter(
        AttendanceRegularization.uuid == regularization_uuid,
        AttendanceRegularization.organization_id == current_org_id,
        AttendanceRegularization.is_deleted == False
    ).first()
    
    if not regularization:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"success": False, "message": "Regularization request not found", "data": None}
        )
        
    if regularization.status != RegularizationStatus.PENDING:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "message": f"Request is already {regularization.status}", "data": None}
        )

    # 2. Update Request Status
    regularization.status = RegularizationStatus.REJECTED
    regularization.rejection_reason = rejection_in.rejection_reason
    
    db.commit()
    db.refresh(regularization)
    
    return AttendanceRegularizationResponse(
        success=True,
        message="Regularization request rejected successfully",
        data=regularization
    )

@router.delete("/regularizations/{regularization_uuid}")
def delete_attendance_regularization(
    regularization_uuid: uuid.UUID,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Cancel/Delete an attendance regularization request.
    Only permitted if the request is still PENDING.
    """
    current_org_id = current_user.id if isinstance(current_user, Organization) else current_user.organization_id
    
    regularization = db.query(AttendanceRegularization).filter(
        AttendanceRegularization.uuid == regularization_uuid,
        AttendanceRegularization.organization_id == current_org_id,
        AttendanceRegularization.is_deleted == False
    ).first()
    
    if not regularization:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"success": False, "message": "Regularization request not found", "data": None}
        )
        
    # RBAC: Allow if Org OR Perm 42 OR Owner + Pending
    is_owner = not isinstance(current_user, Organization) and regularization.employee_id == current_user.id
    has_delete_perm = deps.has_permission(db, current_user, "42")
    
    if not isinstance(current_user, Organization) and not has_delete_perm:
        if not is_owner:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have permission to delete this request."
            )
        if regularization.status != RegularizationStatus.PENDING:
             raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Cannot delete a request that is {regularization.status}. Only PENDING requests can be cancelled."
            )
    else:
        # Org or Manager with Perm 42 still must follow the "Only PENDING" rule if desired?
        # User said: "if is pending then allow to delete for logged in user else not"
        # I'll interpret this as: Managers with Perm 42 can delete even if not pending? 
        # Usually, managers can delete anything. But I'll follow the "Pending" rule for self-service strictly.
        if regularization.status != RegularizationStatus.PENDING and not isinstance(current_user, Organization) and not has_delete_perm:
            raise HTTPException(status_code=400, detail="Only pending requests can be deleted.")
    regularization.is_deleted = True
    db.commit()
    
    return {
        "success": True, 
        "message": "Regularization request cancelled and deleted successfully"
    }

@router.get("/devices", response_model=BiometricDeviceListResponse)
def get_biometric_devices(
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, description="Items per page"),
    is_active: Optional[bool] = Query(None, description="Filter by active status"),
    is_online: Optional[bool] = Query(None, description="Filter by online status"),
    search: Optional[str] = Query(None, description="Search by device name or location"),
    sort_by: Optional[str] = Query("device_name", description="Sort by device_name, physical_location, is_online, is_active"),
    order: Optional[str] = Query("asc", regex="^(asc|desc)$", description="Sort order (asc or desc)")
):
    """
    List all registered biometric devices for the organization with filtering, search, and pagination.
    Supports cross-module access for users with Attendance Read permissions.
    """
    # 1. RBAC: Allow if Org or Employee with Permission 36
    if not deps.has_permission(db, current_user, "36"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not have permission to view devices (requires code 36)"
        )

    current_org_id = current_user.id if isinstance(current_user, Organization) else current_user.organization_id

    # 2. Base Query
    query = db.query(BiometricDevice).filter(
        BiometricDevice.organization_id == current_org_id
    )

    # 3. Filtering
    if is_active is not None:
        query = query.filter(BiometricDevice.is_active == is_active)
    if is_online is not None:
        query = query.filter(BiometricDevice.is_online == is_online)
    
    # 4. Search
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                BiometricDevice.device_name.ilike(search_term),
                BiometricDevice.physical_location.ilike(search_term),
                BiometricDevice.device_id.ilike(search_term)
            )
        )

    # 5. Sorting
    sort_map = {
        "device_name": BiometricDevice.device_name,
        "physical_location": BiometricDevice.physical_location,
        "is_online": BiometricDevice.is_online,
        "is_active": BiometricDevice.is_active,
        "created_at": BiometricDevice.created_at
    }
    
    target_field = sort_map.get(sort_by.lower(), BiometricDevice.device_name)
    if order.lower() == "desc":
        query = query.order_by(target_field.desc())
    else:
        query = query.order_by(target_field.asc())

    # 6. Pagination
    total_records = query.count()
    total_pages = (total_records + limit - 1) // limit
    skip = (page - 1) * limit
    
    devices = query.offset(skip).limit(limit).all()
    
    # Manually map location_uuid for the response
    for device in devices:
        if device.location_id:
            location = db.query(Location).filter(Location.id == device.location_id).first()
            if location:
                device.location_uuid = location.uuid
            else:
                device.location_uuid = None
        else:
            device.location_uuid = None
    
    pagination_data = {
        "total_records": total_records,
        "current_page": page,
        "total_pages": total_pages,
        "page_size": limit
    }

    return BiometricDeviceListResponse(
        success=True,
        message="Biometric devices retrieved successfully",
        data=devices,
        pagination=pagination_data
    )

@router.get("/devices/{device_uuid}", response_model=BiometricDeviceResponse)
def get_biometric_device_details(
    device_uuid: uuid.UUID,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    _: bool = Depends(deps.check_permission("36"))
):
    """
    Get detailed information for a specific biometric device.
    """
    current_org_id = current_user.id if isinstance(current_user, Organization) else current_user.organization_id

    device = db.query(BiometricDevice).filter(
        BiometricDevice.uuid == device_uuid,
        BiometricDevice.organization_id == current_org_id
    ).first()
    
    if not device:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Biometric device not found"
        )
    
    # Map location_uuid for response
    if device.location_id:
        location = db.query(Location).filter(Location.id == device.location_id).first()
        device.location_uuid = location.uuid if location else None
    else:
        device.location_uuid = None
        
    return BiometricDeviceResponse(
        success=True,
        message="Biometric device details retrieved successfully",
        data=device
    )

@router.post("/devices", response_model=BiometricDeviceResponse)
def register_biometric_device(
    device_in: BiometricDeviceCreate,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    _: bool = Depends(deps.check_permission("37"))
):
    """
    Register a new biometric device for the organization.
    """
    current_org_id = current_user.id if isinstance(current_user, Organization) else current_user.organization_id

    # 1. Check for duplicate device_id
    existing = db.query(BiometricDevice).filter(
        BiometricDevice.device_id == device_in.device_id
    ).first()
    
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Device with ID {device_in.device_id} is already registered."
        )
    
    # 2. Resolve location_uuid if provided
    device_data = device_in.model_dump()
    location_uuid = device_data.pop('location_uuid', None)
    location_id = None
    
    if location_uuid:
        location = db.query(Location).filter(
            Location.uuid == location_uuid,
            Location.organization_id == current_org_id
        ).first()
        if not location:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Location with UUID {location_uuid} not found."
            )
        location_id = location.id

    # 3. Create device
    device = BiometricDevice(
        organization_id=current_org_id,
        location_id=location_id,
        **device_data
    )
    
    db.add(device)
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error registering device: {str(e)}"
        )
    
    db.refresh(device)
    
    # Map location_uuid for response
    if device.location_id:
        location = db.query(Location).filter(Location.id == device.location_id).first()
        device.location_uuid = location.uuid if location else None
    else:
        device.location_uuid = None
    
    return BiometricDeviceResponse(
        success=True,
        message="Biometric device registered successfully",
        data=device
    )

@router.put("/devices/{device_uuid}", response_model=BiometricDeviceResponse)
def update_biometric_device(
    device_uuid: uuid.UUID,
    device_in: BiometricDeviceUpdate,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    _: bool = Depends(deps.check_permission("38"))
):
    """
    Update biometric device details.
    """
    current_org_id = current_user.id if isinstance(current_user, Organization) else current_user.organization_id

    device = db.query(BiometricDevice).filter(
        BiometricDevice.uuid == device_uuid,
        BiometricDevice.organization_id == current_org_id
    ).first()
    
    if not device:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Biometric device not found"
        )
    
    update_data = device_in.model_dump(exclude_unset=True)
    
    if 'location_uuid' in update_data:
        location_uuid = update_data.pop('location_uuid')
        if location_uuid:
            location = db.query(Location).filter(
                Location.uuid == location_uuid,
                Location.organization_id == current_org_id
            ).first()
            if not location:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Location with UUID {location_uuid} not found."
                )
            device.location_id = location.id
        else:
            device.location_id = None

    for field, value in update_data.items():
        setattr(device, field, value)
        
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error updating device: {str(e)}"
        )
    
    db.refresh(device)
    
    # Map location_uuid for response
    if device.location_id:
        location = db.query(Location).filter(Location.id == device.location_id).first()
        device.location_uuid = location.uuid if location else None
    else:
        device.location_uuid = None
    
    return BiometricDeviceResponse(
        success=True,
        message="Biometric device updated successfully",
        data=device
    )

@router.get("/devices/{device_id}/status", response_model=BiometricDeviceStatusResponse)
def get_biometric_device_status(
    device_id: str,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    _: bool = Depends(deps.check_permission("36"))
):
    """
    Check real-time status of a biometric device.
    """
    current_org_id = current_user.id if isinstance(current_user, Organization) else current_user.organization_id

    device = db.query(BiometricDevice).filter(
        BiometricDevice.device_id == device_id,
        BiometricDevice.organization_id == current_org_id
    ).first()
    
    if not device:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Biometric device not found"
        )
        
    return BiometricDeviceStatusResponse(
        success=True,
        message="Device status retrieved successfully",
        data={
            "device_id": device.device_id,
            "is_online": device.is_online,
            "last_sync_at": device.last_sync_at,
            "last_sync_status": device.last_sync_status,
            "last_heartbeat": device.last_heartbeat
        }
    )


@router.post("/payroll-export")
def export_attendance_for_payroll(
    request: PayrollExportRequest,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    _: bool = Depends(deps.check_permission("33"))
):
    """
    Export attendance summary for payroll processing to Excel (.xlsx).
    Groups data by employee and aggregates attendance statuses and hours.
    """
    # 1. Fetch Employees
    emp_query = db.query(Employee).filter(
        Employee.organization_id == current_org.id,
        Employee.is_deleted == False
    )
    
    if request.department_uuid:
        emp_query = emp_query.join(Department).filter(Department.uuid == request.department_uuid)
        
    employees = emp_query.all()
    emp_ids = [e.id for e in employees]
    
    # 2. Fetch Attendance Records for the date range
    records = db.query(AttendanceRecord).filter(
        AttendanceRecord.organization_id == current_org.id,
        AttendanceRecord.employee_id.in_(emp_ids),
        AttendanceRecord.attendance_date >= request.from_date,
        AttendanceRecord.attendance_date <= request.to_date
    ).all()
    
    # 3. Group records by employee
    from collections import defaultdict
    emp_records = defaultdict(list)
    for rec in records:
        emp_records[rec.employee_id].append(rec)
        
    # 4. Process each employee
    payroll_data = []
    total_days = (request.to_date - request.from_date).days + 1
    
    for emp in employees:
        recs = emp_records.get(emp.id, [])
        
        present_days = 0.0
        absent_days = 0.0
        half_days = 0
        leave_days = 0.0
        holiday_days = 0
        overtime_hours = 0.0
        late_minutes = 0
        early_departure_minutes = 0
        net_work_hours = 0.0
        
        for rec in recs:
            if rec.status == AttendanceStatus.PRESENT:
                present_days += 1.0
            elif rec.status == AttendanceStatus.HALF_DAY:
                present_days += 0.5
                half_days += 1
            elif rec.status == AttendanceStatus.ABSENT:
                absent_days += 1.0
            elif rec.status == AttendanceStatus.ON_LEAVE:
                leave_days += 1.0
            elif rec.status == AttendanceStatus.HOLIDAY:
                holiday_days += 1
                
            overtime_hours += float(rec.overtime_hours or 0)
            late_minutes += rec.late_by_minutes or 0
            early_departure_minutes += rec.early_departure_minutes or 0
            net_work_hours += float(rec.net_work_hours or 0)
            
        payroll_data.append({
            "employee_code": emp.employee_code,
            "employee_name": f"{emp.first_name} {emp.last_name}",
            "total_days": total_days,
            "present_days": present_days,
            "absent_days": absent_days,
            "half_days": half_days,
            "leave_days": leave_days,
            "holiday_days": holiday_days,
            "overtime_hours": round(overtime_hours, 2),
            "late_minutes": late_minutes,
            "early_departure_minutes": early_departure_minutes,
            "net_work_hours": round(net_work_hours, 2)
        })
        
    # 5. Generate Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Attendance"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center")
    
    headers = [
        "Employee Code", "Employee Name", "Total Days", "Present Days", 
        "Absent Days", "Half Days", "Leave Days", "Holiday Days", 
        "Overtime Hours", "Late Mins", "Early Mins", "Net Work Hours"
    ]
    
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    for item in payroll_data:
        ws.append([
            item["employee_code"],
            item["employee_name"],
            item["total_days"],
            item["present_days"],
            item["absent_days"],
            item["half_days"],
            item["leave_days"],
            item["holiday_days"],
            item["overtime_hours"],
            item["late_minutes"],
            item["early_departure_minutes"],
            item["net_work_hours"]
        ])
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 3

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"payroll_attendance_{request.from_date}_to_{request.to_date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

