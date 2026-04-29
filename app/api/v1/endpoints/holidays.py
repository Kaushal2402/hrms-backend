import uuid
from datetime import date, datetime
from typing import List, Optional, Union
from fastapi import APIRouter, Depends, Query, status, HTTPException, File, UploadFile
from sqlalchemy.orm import Session
from sqlalchemy import extract, or_
from app.models.employee import Employee, Location, Department

from app.api import deps
from app.models.attendance import Holiday, HolidayType
from app.models.organization import Organization
from app.schemas.holiday import (
    HolidayListResponse, HolidaySchema, HolidayCreate, 
    HolidayResponse, HolidayUpdate, HolidayBulkImportResponse, HolidayImportError,
    BulkHolidayCreateRequest, HolidayBulkCreateResponse
)
import io
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname
from openpyxl.styles import Font, PatternFill
from fastapi.responses import StreamingResponse
import csv
import io

from app.core.permissions import HolidayPermissions

router = APIRouter()

@router.get("/", response_model=HolidayListResponse)
def list_holidays(
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    year: Optional[int] = Query(None, description="Filter by year (YYYY)"),
    location_id: Optional[int] = Query(None, description="Filter by location ID"),
    holiday_type: Optional[str] = Query(None, description="Filter by holiday type (public, restricted, etc.)"),
    is_active: Optional[bool] = Query(True, description="Filter by active status"),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    authorized: bool = Depends(deps.check_permission(HolidayPermissions.READ)),
    search: Optional[str] = Query(None, description="Search by holiday name or description"),
    sort_by: str = Query("holiday_date", description="Sort by 'holiday_date' or 'holiday_name'"),
    order: str = Query("asc", description="Sort order ('asc' or 'desc')"),
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, description="Items per page")
):
    """
    Retrieve a list of holidays with various filters and pagination.
    """
    query = db.query(Holiday).filter(
        Holiday.organization_id == current_org.id,
        Holiday.is_deleted == False
    )

    if year:
        query = query.filter(Holiday.holiday_year == year)
    
    if location_id:
        # Since location_ids is a JSON field in the model
        # We need to check if location_id is in the JSON array
        from sqlalchemy import func
        query = query.filter(
            or_(
                Holiday.is_location_specific == False,
                func.json_contains(Holiday.location_ids, func.cast(location_id, func.JSON))
            )
        )

    if holiday_type:
        query = query.filter(Holiday.holiday_type == holiday_type)
    
    if is_active is not None:
        query = query.filter(Holiday.is_active == is_active)

    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Holiday.holiday_name.ilike(search_term),
                Holiday.description.ilike(search_term)
            )
        )

    # Sorting
    if sort_by == "holiday_name":
        sort_attr = Holiday.holiday_name
    else: # default to holiday_date
        sort_attr = Holiday.holiday_date

    if order.lower() == "desc":
        query = query.order_by(sort_attr.desc())
    else:
        query = query.order_by(sort_attr.asc())

    # Pagination
    total_records = query.count()
    total_pages = (total_records + limit - 1) // limit
    skip = (page - 1) * limit

    holidays = query.offset(skip).limit(limit).all()

    # 4. Resolve Scoping Names for Response
    location_ids_all = set()
    department_ids_all = set()
    for h in holidays:
        if h.location_ids:
            location_ids_all.update(h.location_ids)
        if h.department_ids:
            department_ids_all.update(h.department_ids)
    
    location_map = {}
    if location_ids_all:
        locs = db.query(Location).filter(Location.id.in_(list(location_ids_all))).all()
        location_map = {l.id: {"uuid": l.uuid, "location_name": l.location_name} for l in locs}
        
    department_map = {}
    if department_ids_all:
        depts = db.query(Department).filter(Department.id.in_(list(department_ids_all))).all()
        department_map = {d.id: {"uuid": d.uuid, "department_name": d.department_name} for d in depts}
        
    for h in holidays:
        h.locations = [location_map[lid] for lid in h.location_ids if lid in location_map] if h.location_ids else []
        h.departments = [department_map[did] for did in h.department_ids if did in department_map] if h.department_ids else []

    return HolidayListResponse(
        success=True,
        message="Holidays retrieved successfully",
        data=holidays,
        pagination={
            "total_records": total_records,
            "current_page": page,
            "total_pages": total_pages,
            "page_size": limit
        }
    )


@router.post("/", response_model=HolidayResponse)
def create_holiday(
    holiday_in: HolidayCreate,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    authorized: bool = Depends(deps.check_permission("60"))
):
    """
    Create a new holiday record.
    """
    # 1. Check if holiday already exists for this date in the organization
    existing_holiday = db.query(Holiday).filter(
        Holiday.organization_id == current_org.id,
        Holiday.holiday_date == holiday_in.holiday_date
    ).first()

    if existing_holiday:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"A holiday already exists for the date {holiday_in.holiday_date}"
        )

    # 2. Extract year from date
    holiday_year = holiday_in.holiday_date.year

    # 3. Resolve Scoping UUIDs to IDs
    location_ids = []
    if holiday_in.is_location_specific and holiday_in.location_uuids:
        location_records = db.query(Location.id).filter(
            Location.uuid.in_(holiday_in.location_uuids),
            Location.organization_id == current_org.id
        ).all()
        location_ids = [r[0] for r in location_records]

    department_ids = []
    if holiday_in.is_department_specific and holiday_in.department_uuids:
        department_records = db.query(Department.id).filter(
            Department.uuid.in_(holiday_in.department_uuids),
            Department.organization_id == current_org.id
        ).all()
        department_ids = [r[0] for r in department_records]

    # 4. Create holiday record
    holiday = Holiday(
        organization_id=current_org.id,
        holiday_name=holiday_in.holiday_name,
        holiday_date=holiday_in.holiday_date,
        holiday_year=holiday_year,
        holiday_type=holiday_in.holiday_type,
        description=holiday_in.description,
        is_location_specific=holiday_in.is_location_specific,
        location_ids=location_ids if holiday_in.is_location_specific else None,
        is_department_specific=holiday_in.is_department_specific,
        department_ids=department_ids if holiday_in.is_department_specific else None,
        is_optional=holiday_in.is_optional,
        optional_quota_required=holiday_in.optional_quota_required,
        is_restricted=holiday_in.is_restricted,
        max_employees_allowed=holiday_in.max_employees_allowed,
        is_active=holiday_in.is_active
    )

    db.add(holiday)
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create holiday"
        )
    
    db.refresh(holiday)

    # Resolve Scoping Names for details
    if holiday.location_ids:
        locs = db.query(Location).filter(Location.id.in_(holiday.location_ids)).all()
        holiday.locations = [{"uuid": l.uuid, "location_name": l.location_name} for l in locs]
    else:
        holiday.locations = []

    if holiday.department_ids:
        depts = db.query(Department).filter(Department.id.in_(holiday.department_ids)).all()
        holiday.departments = [{"uuid": d.uuid, "department_name": d.department_name} for d in depts]
    else:
        holiday.departments = []

    return HolidayResponse(
        success=True,
        message="Holiday created successfully",
        data=holiday
    )

@router.get("/import-template")
def get_holiday_import_template(
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    authorized: bool = Depends(deps.check_permission("59"))
):
    """
    Generate an Excel template for bulk holiday import.
    """
    wb = Workbook()
    
    # --- LOOKUPS SHEET (Hidden) ---
    ws_lookup = wb.create_sheet("Lookups")
    ws_lookup.sheet_state = 'hidden'
    
    # 1. Holiday Types
    ws_lookup.cell(row=1, column=1, value="Holiday Types")
    valid_types = [t.value for t in HolidayType]
    for idx, t in enumerate(valid_types, start=2):
        ws_lookup.cell(row=idx, column=1, value=t)
    
    # 2. Locations
    ws_lookup.cell(row=1, column=2, value="Locations")
    locations = db.query(Location.location_name).filter(
        Location.organization_id == current_org.id, 
        Location.is_active == True
    ).all()
    location_names = [l[0] for l in locations]
    for idx, name in enumerate(location_names, start=2):
        ws_lookup.cell(row=idx, column=2, value=name)
        
    # 3. Departments
    ws_lookup.cell(row=1, column=3, value="Departments")
    departments = db.query(Department.department_name).filter(
        Department.organization_id == current_org.id, 
        Department.is_active == True
    ).all()
    department_names = [d[0] for d in departments]
    for idx, name in enumerate(department_names, start=2):
        ws_lookup.cell(row=idx, column=3, value=name)

    # 4. Yes/No
    ws_lookup.cell(row=1, column=4, value="YesNo")
    ws_lookup.cell(row=2, column=4, value="Yes")
    ws_lookup.cell(row=3, column=4, value="No")

    # --- MAIN SHEET ---
    ws = wb.active
    ws.title = "Import Holidays"
    
    headers = [
        "Holiday Name*", "Date (YYYY-MM-DD)*", "Holiday Type*", "Description",
        "Location Specific?", "Locations (Names)", "Department Specific?", "Departments (Names)",
        "Is Optional?", "Is Restricted?", "Max Employees"
    ]
    ws.append(headers)
    
    # Styling headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = 20

    # Data Validation
    # Type dropdown: C2:C1000
    type_dv = DataValidation(type="list", formula1=f"Lookups!$A$2:$A${len(valid_types)+1}")
    ws.add_data_validation(type_dv)
    type_dv.add("C2:C1000")
    
    # Yes/No dropdowns
    yesno_dv = DataValidation(type="list", formula1="Lookups!$D$2:$D$3")
    ws.add_data_validation(yesno_dv)
    yesno_dv.add("E2:E1000") # Location Specific?
    yesno_dv.add("G2:G1000") # Department Specific?
    yesno_dv.add("I2:I1000") # Is Optional?
    yesno_dv.add("J2:J1000") # Is Restricted?

    # Locations dropdown: F2:F1000
    if location_names:
        loc_dv = DataValidation(type="list", formula1=f"Lookups!$B$2:$B${len(location_names)+1}")
        ws.add_data_validation(loc_dv)
        loc_dv.add("F2:F1000")

    # Departments dropdown: H2:H1000
    if department_names:
        dept_dv = DataValidation(type="list", formula1=f"Lookups!$C$2:$C${len(department_names)+1}")
        ws.add_data_validation(dept_dv)
        dept_dv.add("H2:H1000")

    # Help text
    ws['L1'] = "Instructions:"
    ws['L1'].font = Font(bold=True)
    ws['L2'] = "* Mandatory fields"
    ws['L3'] = "Dates must be YYYY-MM-DD"
    ws['L4'] = "For multiple locations/departments, separate names with commas."

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=holiday_import_template.xlsx"}
    )

@router.get("/{holiday_id}", response_model=HolidayResponse)
def get_holiday(
    holiday_id: str,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    authorized: bool = Depends(deps.check_permission("59"))
):
    """
    Get detailed information about a specific holiday.
    Supports both internal ID and UUID.
    """
    holiday = None
    try:
        # Try resolving by UUID first
        if "-" in holiday_id:
            uuid_obj = uuid.UUID(holiday_id)
            holiday = db.query(Holiday).filter(
                Holiday.uuid == uuid_obj,
                Holiday.organization_id == current_org.id,
                Holiday.is_deleted == False
            ).first()
        else:
            # Try resolving by internal ID
            holiday = db.query(Holiday).filter(
                Holiday.id == int(holiday_id),
                Holiday.organization_id == current_org.id,
                Holiday.is_deleted == False
            ).first()
    except (ValueError, TypeError):
        pass

    if not holiday:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Holiday not found"
        )

    # Resolve Scoping Names
    if holiday.location_ids:
        locs = db.query(Location).filter(Location.id.in_(holiday.location_ids)).all()
        holiday.locations = [{"uuid": l.uuid, "location_name": l.location_name} for l in locs]
    else:
        holiday.locations = []

    if holiday.department_ids:
        depts = db.query(Department).filter(Department.id.in_(holiday.department_ids)).all()
        holiday.departments = [{"uuid": d.uuid, "department_name": d.department_name} for d in depts]
    else:
        holiday.departments = []

    return HolidayResponse(
        success=True,
        message="Holiday details retrieved successfully",
        data=holiday
    )

@router.put("/{holiday_id}", response_model=HolidayResponse)
def update_holiday(
    holiday_id: str,
    holiday_in: HolidayUpdate,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    authorized: bool = Depends(deps.check_permission(HolidayPermissions.UPDATE))
):
    """
    Update an existing holiday's details.
    """
    # 1. Fetch holiday
    holiday = None
    try:
        if "-" in holiday_id:
            uuid_obj = uuid.UUID(holiday_id)
            holiday = db.query(Holiday).filter(
                Holiday.uuid == uuid_obj,
                Holiday.organization_id == current_org.id,
                Holiday.is_deleted == False
            ).first()
        else:
            holiday = db.query(Holiday).filter(
                Holiday.id == int(holiday_id),
                Holiday.organization_id == current_org.id,
                Holiday.is_deleted == False
            ).first()
    except (ValueError, TypeError):
        pass

    if not holiday:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Holiday not found"
        )

    # 2. Prepare update data
    update_data = holiday_in.model_dump(exclude_unset=True)
    
    # Special handling for date
    if 'holiday_date' in update_data:
        # Check for duplicates on the new date
        duplicate = db.query(Holiday).filter(
            Holiday.organization_id == current_org.id,
            Holiday.holiday_date == update_data['holiday_date'],
            Holiday.id != holiday.id,
            Holiday.is_deleted == False
        ).first()
        if duplicate:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"A holiday already exists for the date {update_data['holiday_date']}"
            )
        # Update year if date changed
        holiday.holiday_year = update_data['holiday_date'].year

    # 3. Apply updates
    update_data = holiday_in.model_dump(exclude_unset=True)
    
    # Resolve Scoping UUIDs if provided
    if 'location_uuids' in update_data:
        location_uuids = update_data.pop('location_uuids')
        if holiday.is_location_specific and location_uuids:
            location_records = db.query(Location.id).filter(
                Location.uuid.in_(location_uuids),
                Location.organization_id == current_org.id
            ).all()
            holiday.location_ids = [r[0] for r in location_records]
        else:
            holiday.location_ids = None

    if 'department_uuids' in update_data:
        department_uuids = update_data.pop('department_uuids')
        if holiday.is_department_specific and department_uuids:
            department_records = db.query(Department.id).filter(
                Department.uuid.in_(department_uuids),
                Department.organization_id == current_org.id
            ).all()
            holiday.department_ids = [r[0] for r in department_records]
        else:
            holiday.department_ids = None

    for field, value in update_data.items():
        setattr(holiday, field, value)

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to update holiday"
        )
    
    db.refresh(holiday)
    
    # Resolve Scoping Names for details
    if holiday.location_ids:
        locs = db.query(Location).filter(Location.id.in_(holiday.location_ids)).all()
        holiday.locations = [{"uuid": l.uuid, "location_name": l.location_name} for l in locs]
    else:
        holiday.locations = []

    if holiday.department_ids:
        depts = db.query(Department).filter(Department.id.in_(holiday.department_ids)).all()
        holiday.departments = [{"uuid": d.uuid, "department_name": d.department_name} for d in depts]
    else:
        holiday.departments = []

    return HolidayResponse(
        success=True,
        message="Holiday updated successfully",
        data=holiday
    )

@router.delete("/{holiday_id}", response_model=HolidayResponse)
def delete_holiday(
    holiday_id: str,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    authorized: bool = Depends(deps.check_permission(HolidayPermissions.DELETE))
):
    """
    Soft delete a holiday record.
    """
    # 1. Fetch holiday
    holiday = None
    try:
        if "-" in holiday_id:
            uuid_obj = uuid.UUID(holiday_id)
            holiday = db.query(Holiday).filter(
                Holiday.uuid == uuid_obj,
                Holiday.organization_id == current_org.id,
                Holiday.is_deleted == False
            ).first()
        else:
            holiday = db.query(Holiday).filter(
                Holiday.id == int(holiday_id),
                Holiday.organization_id == current_org.id,
                Holiday.is_deleted == False
            ).first()
    except (ValueError, TypeError):
        pass

    if not holiday:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Holiday not found"
        )

    # 2. Perform soft delete
    holiday.is_deleted = True
    holiday.deleted_at = datetime.utcnow()

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to delete holiday"
        )

    return HolidayResponse(
        success=True,
        message="Holiday deleted successfully",
        data=None
    )


@router.post("/bulk-import", response_model=HolidayBulkImportResponse)
def bulk_import_holidays(
    file: UploadFile = File(...),
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    authorized: bool = Depends(deps.check_permission("60"))
):
    """
    Import holidays from an Excel or CSV file.
    """
    is_excel = file.filename.endswith(('.xlsx', '.xls'))
    is_csv = file.filename.endswith('.csv')
    
    if not is_excel and not is_csv:
        raise HTTPException(status_code=400, detail="Only Excel (.xlsx) or CSV files are allowed")

    rows = []
    try:
        if is_excel:
            contents = file.file.read()
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                rows.append(dict(zip(headers, row)))
        else:
            contents = file.file.read().decode("utf-8")
            csv_reader = csv.DictReader(io.StringIO(contents))
            rows = list(csv_reader)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading file: {str(e)}")
    
    total_processed = 0
    successful_count = 0
    failed_count = 0
    errors = []

    # Pre-fetch lookup maps for resolution
    location_map = {l.location_name.strip().lower(): l.id for l in db.query(Location).filter(Location.organization_id == current_org.id).all()}
    department_map = {d.department_name.strip().lower(): d.id for d in db.query(Department).filter(Department.organization_id == current_org.id).all()}

    def str_to_bool(val):
        if not val: return False
        return str(val).strip().lower() in ['yes', 'true', '1', 'y']

    for i, row in enumerate(rows, start=1):
        total_processed += 1
        
        # Column mapping (support both CSV header and Excel header)
        name = row.get('Holiday Name*', row.get('holiday_name'))
        date_val = row.get('Date (YYYY-MM-DD)*', row.get('holiday_date'))
        type_str = row.get('Holiday Type*', row.get('holiday_type', 'public'))
        desc = row.get('Description', row.get('description'))
        
        # Scoping
        is_loc_spec = str_to_bool(row.get('Location Specific?', False))
        loc_names_str = row.get('Locations (Names)', '')
        is_dept_spec = str_to_bool(row.get('Department Specific?', False))
        dept_names_str = row.get('Departments (Names)', '')
        
        # Flags
        is_opt = str_to_bool(row.get('Is Optional?', False))
        is_restr = str_to_bool(row.get('Is Restricted?', False))
        max_emp = row.get('Max Employees')

        if not name or not date_val:
            errors.append(HolidayImportError(row=i, name=str(name), error="Name and Date are required"))
            failed_count += 1
            continue

        # Parse Date
        holiday_date = None
        if isinstance(date_val, datetime):
            holiday_date = date_val.date()
        elif isinstance(date_val, date):
            holiday_date = date_val
        else:
            try:
                holiday_date = datetime.strptime(str(date_val).strip(), '%Y-%m-%d').date()
            except ValueError:
                errors.append(HolidayImportError(row=i, name=name, error=f"Invalid date format: {date_val}. Use YYYY-MM-DD"))
                failed_count += 1
                continue

        # Validate Type
        type_str = str(type_str).strip().lower()
        valid_types = [t.value for t in HolidayType]
        if type_str not in valid_types:
            errors.append(HolidayImportError(row=i, name=name, error=f"Invalid type: {type_str}"))
            failed_count += 1
            continue

        # Resolve Locations
        final_loc_ids = []
        if is_loc_spec and loc_names_str:
            names = [n.strip().lower() for n in str(loc_names_str).split(',') if n.strip()]
            for n in names:
                if n in location_map:
                    final_loc_ids.append(location_map[n])
                else:
                    errors.append(HolidayImportError(row=i, name=name, error=f"Location not found: {n}"))
            if not final_loc_ids:
                failed_count += 1
                continue

        # Resolve Departments
        final_dept_ids = []
        if is_dept_spec and dept_names_str:
            names = [n.strip().lower() for n in str(dept_names_str).split(',') if n.strip()]
            for n in names:
                if n in department_map:
                    final_dept_ids.append(department_map[n])
                else:
                    errors.append(HolidayImportError(row=i, name=name, error=f"Department not found: {n}"))
            if not final_dept_ids:
                failed_count += 1
                continue

        # Check for duplicates
        existing = db.query(Holiday).filter(
            Holiday.organization_id == current_org.id,
            Holiday.holiday_date == holiday_date,
            Holiday.is_deleted == False
        ).first()

        if existing:
            errors.append(HolidayImportError(row=i, name=name, error=f"Duplicate holiday for date {holiday_date}"))
            failed_count += 1
            continue

        # Create
        holiday = Holiday(
            organization_id=current_org.id,
            holiday_name=name,
            holiday_date=holiday_date,
            holiday_year=holiday_date.year,
            holiday_type=type_str,
            description=desc,
            is_location_specific=is_loc_spec,
            location_ids=final_loc_ids if is_loc_spec else None,
            is_department_specific=is_dept_spec,
            department_ids=final_dept_ids if is_dept_spec else None,
            is_optional=is_opt,
            is_restricted=is_restr,
            max_employees_allowed=int(max_emp) if max_emp and str(max_emp).isdigit() else None,
            is_active=True,
            is_deleted=False
        )
        db.add(holiday)
        successful_count += 1

    if successful_count > 0:
        try:
            db.commit()
        except Exception as e:
            db.rollback()
            return HolidayBulkImportResponse(
                success=False,
                message=f"Database error: {str(e)}",
                total_processed=total_processed, successful_count=0, failed_count=total_processed,
                errors=[HolidayImportError(row=0, error="Commit failed")]
            )

    return HolidayBulkImportResponse(
        success=True,
        message=f"Imported {successful_count} holidays. {failed_count} failed.",
        total_processed=total_processed, successful_count=successful_count, failed_count=failed_count,
        errors=errors
    )

@router.post("/bulk-create", response_model=HolidayBulkCreateResponse)
def bulk_create_holidays(
    request: BulkHolidayCreateRequest,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    authorized: bool = Depends(deps.check_permission("60"))
):
    """
    Bulk create holiday records from a JSON array.
    """
    total_processed = 0
    successful_count = 0
    failed_count = 0
    errors = []

    for i, holiday_in in enumerate(request.holidays):
        total_processed += 1
        try:
            # Check if holiday already exists for this date in the organization
            existing_holiday = db.query(Holiday).filter(
                Holiday.organization_id == current_org.id,
                Holiday.holiday_date == holiday_in.holiday_date,
                Holiday.is_deleted == False
            ).first()

            if existing_holiday:
                errors.append(HolidayImportError(
                    row=i, 
                    name=holiday_in.holiday_name, 
                    error=f"A holiday already exists for the date {holiday_in.holiday_date}"
                ))
                failed_count += 1
                continue

            # Resolve Scoping UUIDs to IDs
            location_ids = []
            if holiday_in.is_location_specific and holiday_in.location_uuids:
                location_records = db.query(Location.id).filter(
                    Location.uuid.in_(holiday_in.location_uuids),
                    Location.organization_id == current_org.id
                ).all()
                location_ids = [r[0] for r in location_records]

            department_ids = []
            if holiday_in.is_department_specific and holiday_in.department_uuids:
                department_records = db.query(Department.id).filter(
                    Department.uuid.in_(holiday_in.department_uuids),
                    Department.organization_id == current_org.id
                ).all()
                department_ids = [r[0] for r in department_records]

            # Create holiday record
            holiday = Holiday(
                organization_id=current_org.id,
                holiday_name=holiday_in.holiday_name,
                holiday_date=holiday_in.holiday_date,
                holiday_year=holiday_in.holiday_date.year,
                holiday_type=holiday_in.holiday_type,
                description=holiday_in.description,
                is_location_specific=holiday_in.is_location_specific,
                location_ids=location_ids if holiday_in.is_location_specific else None,
                is_department_specific=holiday_in.is_department_specific,
                department_ids=department_ids if holiday_in.is_department_specific else None,
                is_optional=holiday_in.is_optional,
                optional_quota_required=holiday_in.optional_quota_required,
                is_restricted=holiday_in.is_restricted,
                max_employees_allowed=holiday_in.max_employees_allowed,
                is_active=holiday_in.is_active
            )

            db.add(holiday)
            successful_count += 1
        except Exception as e:
            errors.append(HolidayImportError(
                row=i, 
                name=holiday_in.holiday_name, 
                error=str(e)
            ))
            failed_count += 1

    if successful_count > 0:
        try:
            db.commit()
        except Exception as e:
            db.rollback()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to commit bulk holidays: {str(e)}"
            )

    return HolidayBulkCreateResponse(
        success=True,
        message=f"Bulk creation completed. {successful_count} holidays created, {failed_count} failed.",
        total_processed=total_processed,
        successful_count=successful_count,
        failed_count=failed_count,
        errors=errors
    )
