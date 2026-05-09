from typing import List, Optional, Union
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form, Request, BackgroundTasks, status
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, or_, and_
from decimal import Decimal
from app.api import deps
from app.models.employee import Employee, JobTitle, Department, Location, EmployeePersonalInfo, EmployeeAddress, EmployeeEmergencyContact, EmployeeEducation, EmployeeCertification, EmployeeWorkExperience, EmployeeHistory, EmployeeDocument
from app.schemas.employee import (
    EmployeeSchema,
    EmployeeCreate,
    EmployeeUpdate,
    EmployeeResponse,
    EmployeeListResponse,
    EmployeeDetailResponse,
    EmployeeDetailSchema,
    EmployeeDeleteResponse,
    EmployeeImportResponse,
    EmployeeImportError,
    EmploymentStatus,
    EmploymentType,
    PersonalInfoCreate,
    PersonalInfoUpdate,
    PersonalInfoResponse,
    AddressCreate,
    AddressUpdate,
    AddressResponse,
    AddressListResponse,
    EmergencyContactCreate,
    EmergencyContactUpdate,
    EmergencyContactResponse,
    EmergencyContactListResponse,
    EducationCreate,
    EducationUpdate,
    EducationResponse,
    EducationListResponse,
    CertificationCreate,
    CertificationUpdate,
    CertificationResponse,
    CertificationListResponse,
    WorkExperienceCreate,
    WorkExperienceUpdate,
    WorkExperienceResponse,
    WorkExperienceCreate,
    WorkExperienceUpdate,
    WorkExperienceResponse,
    WorkExperienceListResponse,
    ReportingStructureData,
    ReportingStructureResponse,
    EmployeeSummarySchema,
    EmployeeHistorySchema,
    EmployeeHistoryListResponse,
    EmployeeHistoryCreate,
    EmployeeHistoryUpdate,
    EmployeeHistoryResponse,
    EmployeeDocumentSchema,
    EmployeeDocumentResponse,
    EmployeeDocumentListResponse,
    EmployeeDocumentUpdate,
    DocumentVerification,
    ChangeType,
    DocumentType,
    EmployeeSummaryCardResponse
)
from app.schemas.leave import (
    LeaveBalanceListResponse, LeaveAccrualHistoryListResponse,
    LeaveApplicationListResponse, LeaveCalendarResponse, LeaveCalendarEvent,
    EmployeeCompOffResponse, EmployeeCompOffListResponse,
    HolidayCalendarEvent, LeaveCalendarData
)
from app.schemas.holiday import EmployeeOptionalHolidayListResponse, HolidayListResponse
from app.schemas.attendance import (
    ShiftRosterListResponse, OvertimeSummaryResponse,
    ApprovalDelegationListResponse, ApprovalDelegationSchema,
    ShiftRosterCalendarResponse, CalendarHoliday, CalendarLeave
)
from app.models.attendance import (
    ShiftRoster, ShiftMaster, OvertimeRequest, OvertimeStatus, CompensationType,
    LeaveBalance, LeaveType, LeaveAccrualHistory, LeaveApplication, LeaveStatus,
    CompensatoryOff, Holiday, OptionalHolidaySelection, HolidayType,
    ApprovalDelegation, DelegationType
)
from app.models.rbac import Role, UserRole
from app.schemas.department import PaginationData
from app.models.organization import Organization
from app.core import security
from app.core.permissions import EmployeePermissions
from app.utils.email import send_set_password_email
from datetime import datetime, date, timedelta
import uuid
import csv
import io
import shutil
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname
from fastapi import UploadFile, File

router = APIRouter()

@router.get("/me", response_model=EmployeeDetailResponse)
def get_own_profile(
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Get the currently logged-in employee's profile.
    Only works for Employee accounts.
    """
    if isinstance(current_user, Organization):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, 
            detail="Organizations do not have an employee profile. Use this endpoint with an employee account."
        )
    
    # current_user is already the Employee object
    # Populate role manually as it's not a direct relationship on the Employee model
    from app.models.rbac import UserRole
    primary_user_role = db.query(UserRole).filter(
        UserRole.user_id == current_user.id,
        UserRole.is_primary == True,
        UserRole.is_active == True
    ).first()
    
    if primary_user_role:
        current_user.role = primary_user_role.role
    
    return EmployeeDetailResponse(
        success=True,
        message="Profile retrieved successfully",
        data=current_user
    )

@router.get("/me/shift-roster", response_model=ShiftRosterCalendarResponse)
def get_my_shift_roster(
    from_date: Optional[date] = Query(None, description="Filter from roster date"),
    to_date: Optional[date] = Query(None, description="Filter to roster date"),
    page: Optional[int] = Query(None, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, description="Items per page"),
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Get the currently logged-in employee's published shift roster, holidays, and leaves.
    """
    if isinstance(current_user, Organization):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, 
            detail="Organizations do not have a shift roster. Use this endpoint with an employee account."
        )
    
    query = db.query(ShiftRoster).filter(
        ShiftRoster.employee_id == current_user.id,
        ShiftRoster.organization_id == current_user.organization_id,
        ShiftRoster.is_deleted == False,
        ShiftRoster.is_published == True
    )

    if from_date:
        query = query.filter(ShiftRoster.roster_date >= from_date)
    if to_date:
        query = query.filter(ShiftRoster.roster_date <= to_date)
        
    query = query.options(
        joinedload(ShiftRoster.shift)
    ).order_by(ShiftRoster.roster_date.asc())

    total_records = query.count()
    
    pagination_data = None
    if page is not None:
        total_pages = (total_records + limit - 1) // limit
        skip = (page - 1) * limit
        rosters = query.offset(skip).limit(limit).all()
        pagination_data = {
            "total_records": total_records,
            "current_page": page,
            "total_pages": total_pages,
            "page_size": limit
        }
    else:
        rosters = query.all()
        pagination_data = {
            "total_records": total_records,
            "current_page": 1,
            "total_pages": 1,
            "page_size": total_records if total_records > 0 else 0
        }
    # 1. Fetch Holidays
    cal_from = from_date or date.today().replace(day=1)
    cal_to = to_date or (date.today() + timedelta(days=31))
    
    holidays = db.query(Holiday).filter(
        Holiday.organization_id == current_user.organization_id,
        Holiday.holiday_date >= cal_from,
        Holiday.holiday_date <= cal_to,
        Holiday.is_active == True
    ).all()
    
    calendar_holidays = [
        CalendarHoliday(
            holiday_name=h.holiday_name,
            holiday_date=h.holiday_date,
            holiday_type=h.holiday_type.value if hasattr(h.holiday_type, 'value') else str(h.holiday_type)
        ) for h in holidays
    ]
    
    # 2. Fetch Leaves
    leaves = db.query(LeaveApplication).filter(
        LeaveApplication.employee_id == current_user.id,
        LeaveApplication.status.in_([LeaveStatus.APPROVED, LeaveStatus.REJECTED, LeaveStatus.CANCELLED]),
        LeaveApplication.from_date <= cal_to,
        LeaveApplication.to_date >= cal_from
    ).options(joinedload(LeaveApplication.leave_type)).all()
    
    calendar_leaves = [
        CalendarLeave(
            leave_type_name=l.leave_type.leave_name,
            from_date=l.from_date,
            to_date=l.to_date,
            status=l.status.value if hasattr(l.status, 'value') else str(l.status),
            total_days=float(l.total_days)
        ) for l in leaves
    ]

    return ShiftRosterCalendarResponse(
        success=True,
        message="My calendar data retrieved successfully",
        rosters=rosters,
        holidays=calendar_holidays,
        leaves=calendar_leaves,
        pagination=pagination_data
    )

@router.get("/", response_model=EmployeeListResponse)
def list_employees(
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    page: Optional[int] = Query(None, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, description="Items per page"),
    sort_by: Optional[str] = Query(None, description="Sort by field"),
    order: Optional[str] = Query("desc", description="Sort order (asc/desc)"),
    department_id: Optional[uuid.UUID] = Query(None, description="Filter by Department UUID"),
    location_id: Optional[uuid.UUID] = Query(None, description="Filter by Location UUID"),
    employment_status: Optional[EmploymentStatus] = Query(None, description="Filter by Employment Status"),
    employment_type: Optional[EmploymentType] = Query(None, description="Filter by Employment Type"),
    search: Optional[str] = Query(None, description="Search term (Name, Email, Code)"),
    is_active: Optional[bool] = Query(None, description="Filter by Active Status"),
    date_of_joining_from: Optional[date] = Query(None, description="Date of Joining From"),
    date_of_joining_to: Optional[date] = Query(None, description="Date of Joining To"),
    authorized: bool = Depends(deps.check_permission("1"))
):
    """
    List all employees with filtering and pagination.
    """
    query = db.query(Employee).filter(
        Employee.organization_id == current_org.id,
        Employee.is_deleted == False
    )

    # Exclude currently logged-in employee if applicable
    if isinstance(current_user, Employee):
        query = query.filter(Employee.id != current_user.id)
    
    # 1. Filters
    if is_active is not None:
        query = query.filter(Employee.is_active == is_active)
        
    if employment_status:
        query = query.filter(Employee.employment_status == employment_status)
        
    if employment_type:
        query = query.filter(Employee.employment_type == employment_type)
        
    if department_id:
        dept = db.query(Department).filter(Department.uuid == department_id, Department.organization_id == current_org.id).first()
        if dept:
            query = query.filter(Employee.department_id == dept.id)
        else:
            # If invalid department UUID, return empty or error? Converting to empty result is safer for filters.
            query = query.filter(Employee.id == -1) 
            
    if location_id:
        loc = db.query(Location).filter(Location.uuid == location_id, Location.organization_id == current_org.id).first()
        if loc:
            query = query.filter(Employee.location_id == loc.id)
        else:
            query = query.filter(Employee.id == -1)

    if date_of_joining_from:
        query = query.filter(Employee.date_of_joining >= date_of_joining_from)
    
    if date_of_joining_to:
        query = query.filter(Employee.date_of_joining <= date_of_joining_to)
        
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Employee.first_name.ilike(search_term),
                Employee.last_name.ilike(search_term),
                Employee.work_email.ilike(search_term),
                Employee.employee_code.ilike(search_term),
                # TODO: Concatenated full name search if needed
            )
        )

    # 2. Sorting
    sort_mapping = {
        "employee_code": Employee.employee_code,
        "first_name": Employee.first_name,
        "last_name": Employee.last_name,
        "employment_type": Employee.employment_type,
        "employment_status": Employee.employment_status,
        "is_active": Employee.is_active,
        "date_of_joining": Employee.date_of_joining,
        "email": Employee.work_email,
        "Recent": Employee.id,
        "Oldest": Employee.id
    }
    
    # Default to created_at if sort_by is not in mapping
    sort_field = sort_mapping.get(sort_by, Employee.created_at)
    
    # Handle Oldest specifically
    effective_order = order.lower() if order else "desc"
    if sort_by == "Oldest":
         effective_order = "asc"

    if effective_order == "asc":
        query = query.order_by(sort_field.asc())
    else:
        query = query.order_by(sort_field.desc())

    # 3. Pagination
    total_records = query.count()
    
    pagination_data = None
    if page is not None:
        total_pages = (total_records + limit - 1) // limit
        skip = (page - 1) * limit
        employees = query.offset(skip).limit(limit).all()
        pagination_data = {
            "total_records": total_records,
            "current_page": page,
            "total_pages": total_pages,
            "page_size": limit
        }
    else:
        employees = query.all()
        pagination_data = {
            "total_records": total_records,
            "current_page": 1,
            "total_pages": 1,
            "page_size": total_records if total_records > 0 else 0
        }
        
    if not employees:
        return EmployeeListResponse(
            success=False,
            message="No employees found"
        )

    # Populate role for list view too, though expensive n+1. Optimization: join UserRole and Role.
    # For now iterate
    for emp in employees:
         primary_user_role = db.query(UserRole).filter(
            UserRole.user_id == emp.id,
            UserRole.is_primary == True,
            UserRole.is_active == True
        ).first()
         if primary_user_role:
             emp.role = primary_user_role.role

    return EmployeeListResponse(
        success=True,
        message="Employees retrieved successfully",
        data=employees,
        pagination=pagination_data
    )

@router.get("/lookup")
def lookup_employees(
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    search: Optional[str] = Query(None, description="Search by name or code"),
    limit: int = Query(100, ge=1, description="Limit lookup results")
):
    """
    Lite lookup endpoint for employees (UUID, Full Name, Code, Email).
    Accessible to all authenticated users for filters/dropdowns.
    """
    query = db.query(Employee).filter(
        Employee.organization_id == current_org.id,
        Employee.is_active == True,
        Employee.is_deleted == False
    )
    
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Employee.first_name.ilike(search_term),
                Employee.last_name.ilike(search_term),
                Employee.employee_code.ilike(search_term),
                Employee.work_email.ilike(search_term)
            )
        )
        
    employees = query.order_by(Employee.first_name.asc()).limit(limit).all()
    
    return {
        "success": True,
        "data": [
            {
                "uuid": emp.uuid, 
                "full_name": f"{emp.first_name} {emp.last_name}",
                "employee_code": emp.employee_code,
                "work_email": emp.work_email
            } for emp in employees
        ]
    }

@router.post("/", response_model=EmployeeResponse)
def create_employee(
    employee_in: EmployeeCreate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    authorized: bool = Depends(deps.check_permission("2"))
):
    """
    Create a new employee.
    """
    # 1. Check uniqueness of emails (Global check across Org and Employee)
    emails_to_check = [employee_in.work_email]
    if employee_in.personal_email:
        emails_to_check.append(employee_in.personal_email)
    
    for email in emails_to_check:
        if db.query(Organization).filter(Organization.email == email).first() or \
           db.query(Employee).filter(or_(Employee.work_email == email, Employee.personal_email == email)).first():
            raise HTTPException(status_code=400, detail=f"The email address '{email}' is already in use.")
    
    # 2. Check Employee Code if provided, else generate
    if employee_in.employee_code:
        if db.query(Employee).filter(
            Employee.organization_id == current_org.id,
            Employee.employee_code == employee_in.employee_code
        ).first():
            raise HTTPException(status_code=400, detail="Employee code already exists.")
        final_emp_code = employee_in.employee_code
    else:
        # Generate code: EMP-{max_id + 1}
        # Note: This is not race-condition safe but sufficient for now.
        # Ideally use a sequence or retry logic.
        max_id = db.query(func.max(Employee.id)).scalar() or 0
        final_emp_code = f"EMP-{max_id + 1:04d}"
        
        # Double check generation (paranoid check)
        while db.query(Employee).filter(
            Employee.organization_id == current_org.id,
            Employee.employee_code == final_emp_code
        ).first():
             max_id += 1
             final_emp_code = f"EMP-{max_id + 1:04d}"

    # 3. Resolve Related Entities (UUID -> ID)
    # Job Title
    job_title = db.query(JobTitle).filter(
        JobTitle.uuid == employee_in.job_title_id,
        JobTitle.organization_id == current_org.id
    ).first()
    if not job_title:
        raise HTTPException(status_code=400, detail="Invalid Job Title UUID")
        
    # Department
    department = db.query(Department).filter(
        Department.uuid == employee_in.department_id,
        Department.organization_id == current_org.id
    ).first()
    if not department:
        raise HTTPException(status_code=400, detail="Invalid Department UUID")
        
    # Location
    location = db.query(Location).filter(
        Location.uuid == employee_in.location_id,
        Location.organization_id == current_org.id
    ).first()
    if not location:
        raise HTTPException(status_code=400, detail="Invalid Location UUID")
        
    # Reporting Manager
    manager_id = None
    if employee_in.reporting_manager_id:
        manager = db.query(Employee).filter(
            Employee.uuid == employee_in.reporting_manager_id,
            Employee.organization_id == current_org.id
        ).first()
        if not manager:
            raise HTTPException(status_code=400, detail="Invalid Reporting Manager UUID")
        manager_id = manager.id

    # 4. Create Employee
    # Exclude UUID fields and auto-generated ones

    data = employee_in.model_dump(exclude={
        'job_title_id', 'department_id', 'location_id', 'reporting_manager_id', 'employee_code', 'role_uuid'
    })
    
    db_obj = Employee(
        **data,
        employee_code=final_emp_code,
        organization_id=current_org.id,
        job_title_id=job_title.id,
        department_id=department.id,
        location_id=location.id,
        reporting_manager_id=manager_id
    )
    
    db.add(db_obj)
    db.flush()

    if employee_in.role_uuid:
        role = db.query(Role).filter(
            Role.uuid == employee_in.role_uuid,
            Role.organization_id == current_org.id
        ).first()

        if role:
             user_role = UserRole(
                user_id=db_obj.id,
                role_id=role.id,
                is_primary=True,
                is_active=True,
                valid_from=datetime.utcnow()
            )
             db.add(user_role)

    try:
        db.commit()
        db.refresh(db_obj)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    # Generate invitation token and schedule email
    invite_token = security.create_access_token(subject=db_obj.uuid, expires_delta=timedelta(hours=24))
    db_obj.reset_password_token = invite_token
    db_obj.reset_password_token_expires_at = datetime.utcnow() + timedelta(hours=24)
    db.commit()

    background_tasks.add_task(send_set_password_email, db_obj.work_email, invite_token, db_obj.first_name)

    # Attach role for response
    if employee_in.role_uuid and role:
         db_obj.role = role

    return EmployeeResponse(
        success=True,
        message="Employee created successfully",
        data=db_obj
    )

@router.post("/{employee_uuid}/send-invitation", response_model=EmployeeResponse)
def send_employee_invitation(
    employee_uuid: uuid.UUID,
    background_tasks: BackgroundTasks,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    authorized: bool = Depends(deps.check_permission("1"))
):
    """
    Send invitation email to a specific employee to set their password.
    """
    employee = db.query(Employee).filter(
        Employee.uuid == employee_uuid,
        Employee.organization_id == current_org.id,
        Employee.is_deleted == False
    ).first()
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    if employee.hashed_password:
         raise HTTPException(status_code=400, detail="Employee has already set their password")

    # Generate invitation token and schedule email
    invite_token = security.create_access_token(subject=str(employee.uuid), expires_delta=timedelta(hours=24))
    employee.reset_password_token = invite_token
    employee.reset_password_token_expires_at = datetime.utcnow() + timedelta(hours=24)
    db.commit()

    background_tasks.add_task(send_set_password_email, employee.work_email, invite_token, employee.first_name)

    return EmployeeResponse(
        success=True,
        message="Invitation email sent successfully",
        data=employee
    )

@router.post("/send-invitations-all", response_model=EmployeeDeleteResponse)
def send_all_employee_invitations(
    background_tasks: BackgroundTasks,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    authorized: bool = Depends(deps.check_permission("1"))
):
    """
    Send invitation email to all employees who haven't set their password.
    """
    employees = db.query(Employee).filter(
        Employee.organization_id == current_org.id,
        Employee.is_deleted == False,
        or_(Employee.hashed_password == None, Employee.hashed_password == "")
    ).all()
    
    if not employees:
        return EmployeeDeleteResponse(
            success=True, 
            message="No employees found without passwords", 
            data=None
        )
        
    for employee in employees:
        invite_token = security.create_access_token(subject=str(employee.uuid), expires_delta=timedelta(hours=24))
        employee.reset_password_token = invite_token
        employee.reset_password_token_expires_at = datetime.utcnow() + timedelta(hours=24)
        background_tasks.add_task(send_set_password_email, employee.work_email, invite_token, employee.first_name)
    
    db.commit()

    return EmployeeDeleteResponse(
        success=True, 
        message=f"Invitations sent to {len(employees)} employees", 
        data={"count": len(employees)}
    )

@router.get("/export")
def export_employees(
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    department_id: Optional[uuid.UUID] = Query(None, description="Filter by Department UUID"),
    location_id: Optional[uuid.UUID] = Query(None, description="Filter by Location UUID"),
    employment_status: Optional[EmploymentStatus] = Query(None, description="Filter by Employment Status"),
    employment_type: Optional[EmploymentType] = Query(None, description="Filter by Employment Type"),
    search: Optional[str] = Query(None, description="Search term (Name, Email, Code)"),
    is_active: Optional[bool] = Query(None, description="Filter by Active Status"),
    date_of_joining_from: Optional[date] = Query(None, description="Date of Joining From"),
    date_of_joining_to: Optional[date] = Query(None, description="Date of Joining To"),
    format: str = Query("csv", description="Export format (csv)"),
    authorized: bool = Depends(deps.check_permission("1"))
):
    """
    Export employees to CSV.
    """
    query = db.query(Employee).filter(
        Employee.organization_id == current_org.id,
        Employee.is_deleted == False
    )
    
    # Filters
    if is_active is not None:
        query = query.filter(Employee.is_active == is_active)
        
    if employment_status:
        query = query.filter(Employee.employment_status == employment_status)
        
    if employment_type:
        query = query.filter(Employee.employment_type == employment_type)
        
    if department_id:
        dept = db.query(Department).filter(Department.uuid == department_id, Department.organization_id == current_org.id).first()
        if dept:
            query = query.filter(Employee.department_id == dept.id)
        else:
            query = query.filter(Employee.id == -1) 
            
    if location_id:
        loc = db.query(Location).filter(Location.uuid == location_id, Location.organization_id == current_org.id).first()
        if loc:
            query = query.filter(Employee.location_id == loc.id)
        else:
            query = query.filter(Employee.id == -1)

    if date_of_joining_from:
        query = query.filter(Employee.date_of_joining >= date_of_joining_from)
    
    if date_of_joining_to:
        query = query.filter(Employee.date_of_joining <= date_of_joining_to)
        
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Employee.first_name.ilike(search_term),
                Employee.last_name.ilike(search_term),
                Employee.work_email.ilike(search_term),
                Employee.employee_code.ilike(search_term),
            )
        )
        
    # Eager load
    query = query.options(
        joinedload(Employee.job_title),
        joinedload(Employee.department),
        joinedload(Employee.location),
        joinedload(Employee.reporting_manager)
    )
    
    employees = query.all()
    
    # 1. Headers
    headers = [
        'Employee Code', 'First Name', 'Last Name', 'Work Email', 'Personal Email', 'Mobile Number', 
        'Gender', 'Employment Type', 'Employment Status', 'Date of Birth', 'Date of Joining', 
        'Date of Confirmation', 'Job Title Code', 'Department Code', 'Location Code', 
        'Reporting Manager Email'
    ]
    
    # 2. Collect Data
    data_rows = []
    for emp in employees:
        data_rows.append([
            emp.employee_code,
            emp.first_name,
            emp.last_name,
            emp.work_email,
            emp.personal_email,
            emp.mobile_number,
            emp.gender.value if emp.gender else '',
            emp.employment_type.value if emp.employment_type else '',
            emp.employment_status.value if emp.employment_status else '',
            emp.date_of_birth.isoformat() if isinstance(emp.date_of_birth, (date, datetime)) else str(emp.date_of_birth or ''),
            emp.date_of_joining.isoformat() if isinstance(emp.date_of_joining, (date, datetime)) else str(emp.date_of_joining or ''),
            emp.date_of_confirmation.isoformat() if isinstance(emp.date_of_confirmation, (date, datetime)) else str(emp.date_of_confirmation or ''),
            emp.job_title.title_code if emp.job_title else '',
            emp.department.department_code if emp.department else '',
            emp.location.location_code if emp.location else '',
            emp.reporting_manager.work_email if emp.reporting_manager else ''
        ])
        
    # 3. Export as requested format
    if format.lower() == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"
        
        # Add headers
        ws.append(headers)
        
        # Add data
        for r in data_rows:
            ws.append(r)
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=employees_export_{date.today()}.xlsx"}
        )
    else: # Default to CSV
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for r in data_rows:
            writer.writerow(r)
        
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8')),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=employees_export_{date.today()}.csv"}
        )

@router.get("/import-template")
def get_import_template(
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    authorized: bool = Depends(deps.check_permission("1"))
):
    """
    Download Excel template for bulk import with dropdowns populated from database.
    """
    # 1. Fetch Data
    # 1. Fetch Data
    job_titles = [r[0] for r in db.query(JobTitle.title_code).filter(JobTitle.organization_id == current_org.id, JobTitle.is_active == True).all()]
    departments = [r[0] for r in db.query(Department.department_code).filter(Department.organization_id == current_org.id, Department.is_active == True).all()]
    locations = [r[0] for r in db.query(Location.location_code).filter(Location.organization_id == current_org.id, Location.is_active == True).all()]
    roles = [r[0] for r in db.query(Role.role_code).filter(Role.organization_id == current_org.id, Role.is_active == True).all()]
    
    # Enums
    # Hardcoding enums from schema definitions to avoid reflection complexity if simple
    genders = ["male", "female", "other", "prefer_not_to_say"]
    emp_types = [e.value for e in EmploymentType]
    emp_statuses = [e.value for e in EmploymentStatus]
    
    wb = Workbook()
    ws_template = wb.active
    ws_template.title = "Template"
    
    # Headers
    headers = [
        'First Name', 'Last Name', 'Work Email', 'Mobile Number', 'Gender', 'Date of Birth',
        'Employment Type', 'Employment Status', 'Date of Joining', 'Date of Confirmation', 
        'Personal Email', 'Job Title Code', 'Department Code', 'Location Code', 
        'Reporting Manager Email', 'Role Code'
    ]
    ws_template.append(headers)
    
    # Create Data Sheet for validations
    ws_data = wb.create_sheet("Data")
    ws_data.sheet_state = "hidden" # Hide it
    
    # Helper to populate column
    def populate_col(data_list, col_idx):
        for i, val in enumerate(data_list, start=1):
            # val is likely a tuple or scalar. query(Col) returns tuples
            v = val[0] if isinstance(val, tuple) else val
            ws_data.cell(row=i, column=col_idx, value=v)
        return len(data_list)
        
    jt_count = populate_col(job_titles, 1)
    dept_count = populate_col(departments, 2)
    loc_count = populate_col(locations, 3)
    gender_count = populate_col(genders, 4)
    type_count = populate_col(emp_types, 5)
    status_count = populate_col(emp_statuses, 6)
    role_count = populate_col(roles, 7)
    
    # Data Validations
    # Gender (Col 5 -> E)
    if gender_count > 0:
        dv_gen = DataValidation(type="list", formula1=f"'Data'!$D$1:$D${gender_count}", allow_blank=True)
        ws_template.add_data_validation(dv_gen)
        dv_gen.add("E2:E1000")
        
    # Emp Type (Col 7 -> G)
    if type_count > 0:
        dv_type = DataValidation(type="list", formula1=f"'Data'!$E$1:$E${type_count}", allow_blank=True)
        ws_template.add_data_validation(dv_type)
        dv_type.add("G2:G1000")
        
    # Emp Status (Col 8 -> H)
    if status_count > 0:
        dv_stat = DataValidation(type="list", formula1=f"'Data'!$F$1:$F${status_count}", allow_blank=True)
        ws_template.add_data_validation(dv_stat)
        dv_stat.add("H2:H1000")

    # Job Title (Col 12 -> L)
    if jt_count > 0:
        dv_jt = DataValidation(type="list", formula1=f"'Data'!$A$1:$A${jt_count}", allow_blank=True)
        ws_template.add_data_validation(dv_jt)
        dv_jt.add("L2:L1000")
        
    # Department (Col 13 -> M)
    if dept_count > 0:
        dv_dept = DataValidation(type="list", formula1=f"'Data'!$B$1:$B${dept_count}", allow_blank=True)
        ws_template.add_data_validation(dv_dept)
        dv_dept.add("M2:M1000")
        
    # Location (Col 14 -> N)
    if loc_count > 0:
        dv_loc = DataValidation(type="list", formula1=f"'Data'!$C$1:$C${loc_count}", allow_blank=True)
        ws_template.add_data_validation(dv_loc)
        dv_loc.add("N2:N1000")
        
    # Role Code (Col 16 -> P)
    if role_count > 0:
        dv_role = DataValidation(type="list", formula1=f"'Data'!$G$1:$G${role_count}", allow_blank=True)
        ws_template.add_data_validation(dv_role)
        dv_role.add("P2:P1000")
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=employee_import_template.xlsx"}
    )

@router.get("/{employee_id}", response_model=EmployeeDetailResponse)
def get_employee(
    employee_id: str,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    authorized: bool = Depends(deps.check_permission("1"))
):
    """
    Get employee details by ID (UUID or Integer ID).
    """
    error = None
    employee = None
    
    # 1. Try as UUID
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj,
            Employee.organization_id == current_org.id,
            Employee.is_deleted == False
        ).first()
    except ValueError:
        pass
        
    # 2. Try as Integer ID
    if not employee:
        try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(
                Employee.id == int_id,
                Employee.organization_id == current_org.id,
                Employee.is_deleted == False
            ).first()
        except ValueError:
            pass
            
    # 3. Try as Employee Code ? (Bonus/Robustness)
    if not employee:
         employee = db.query(Employee).filter(
            Employee.employee_code == employee_id,
            Employee.organization_id == current_org.id,
            Employee.is_deleted == False
        ).first()
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # Ensure relationships are loaded (eagar load usually better but lazy works with active session)
    
    # Manually attach role for response if schema has it
    # Find primary role
    if employee:
        primary_user_role = db.query(UserRole).filter(
            UserRole.user_id == employee.id,
            UserRole.is_primary == True,
            UserRole.is_active == True
        ).first()
        
        if primary_user_role:
             employee.role = primary_user_role.role

    return EmployeeDetailResponse(
        success=True,
        message="Employee details retrieved successfully",
        data=employee
    )

@router.get("/{employee_id}/summary", response_model=EmployeeSummaryCardResponse)
def get_employee_summary_card(
    employee_id: str,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    authorized: bool = Depends(deps.check_permission("1"))
):
    """
    Get condensed employee information for a summary card (quick view).
    """
    employee = None
    
    # 1. Try as UUID
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).options(
            joinedload(Employee.job_title),
            joinedload(Employee.department),
            joinedload(Employee.location),
            joinedload(Employee.reporting_manager)
        ).filter(
            Employee.uuid == uuid_obj,
            Employee.organization_id == current_org.id,
            Employee.is_deleted == False
        ).first()
    except ValueError:
        pass
        
    # 2. Try as Integer ID
    if not employee:
        try:
            int_id = int(employee_id)
            employee = db.query(Employee).options(
                joinedload(Employee.job_title),
                joinedload(Employee.department),
                joinedload(Employee.location),
                joinedload(Employee.reporting_manager)
            ).filter(
                Employee.id == int_id,
                Employee.organization_id == current_org.id,
                Employee.is_deleted == False
            ).first()
        except ValueError:
            pass
            
    # 3. Try as Employee Code
    if not employee:
         employee = db.query(Employee).options(
            joinedload(Employee.job_title),
            joinedload(Employee.department),
            joinedload(Employee.location),
            joinedload(Employee.reporting_manager)
        ).filter(
            Employee.employee_code == employee_id,
            Employee.organization_id == current_org.id,
            Employee.is_deleted == False
        ).first()
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    return EmployeeSummaryCardResponse(
        success=True,
        message="Employee summary card retrieved successfully",
        data=employee
    )


@router.put("/{employee_id}", response_model=EmployeeDetailResponse)
def update_employee(
    employee_id: str,
    employee_in: EmployeeUpdate,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    authorized: bool = Depends(deps.check_permission("3"))
):
    """
    Update employee information.
    """
    employee = None
    
    # 1. Resolve Employee
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj,
            Employee.organization_id == current_org.id,
            Employee.is_deleted == False
        ).first()
    except ValueError:
        pass
        
    if not employee:
        try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(
                Employee.id == int_id,
                Employee.organization_id == current_org.id,
                Employee.is_deleted == False
            ).first()
        except ValueError:
            pass

    if not employee:
         employee = db.query(Employee).filter(
            Employee.employee_code == employee_id,
            Employee.organization_id == current_org.id,
            Employee.is_deleted == False
        ).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Validation & FK Resolution
    update_data = employee_in.model_dump(exclude_unset=True)
            
    # Resolve FKs
    if 'job_title_id' in update_data:
        jt = db.query(JobTitle).filter(JobTitle.uuid == update_data['job_title_id'], JobTitle.organization_id == current_org.id).first()
        if not jt: raise HTTPException(status_code=400, detail="Invalid Job Title UUID")
        update_data['job_title_id'] = jt.id # Replace UUID with ID
        
    if 'department_id' in update_data:
        dept = db.query(Department).filter(Department.uuid == update_data['department_id'], Department.organization_id == current_org.id).first()
        if not dept: raise HTTPException(status_code=400, detail="Invalid Department UUID")
        update_data['department_id'] = dept.id
        
    if 'location_id' in update_data:
        loc = db.query(Location).filter(Location.uuid == update_data['location_id'], Location.organization_id == current_org.id).first()
        if not loc: raise HTTPException(status_code=400, detail="Invalid Location UUID")
        update_data['location_id'] = loc.id
        
    if 'reporting_manager_id' in update_data:
        if update_data['reporting_manager_id'] is None:
             update_data['reporting_manager_id'] = None
        else:
            mgr = db.query(Employee).filter(Employee.uuid == update_data['reporting_manager_id'], Employee.organization_id == current_org.id).first()
            if not mgr: raise HTTPException(status_code=400, detail="Invalid Reporting Manager UUID")
            # prevent self-referential manager loop (simple check: can't be self)
            if mgr.id == employee.id:
                 raise HTTPException(status_code=400, detail="Employee cannot report to themselves")
            update_data['reporting_manager_id'] = mgr.id
            
    # 3. Apply Updates
    for field, value in update_data.items():
        setattr(employee, field, value)
        
    db.add(employee)
    try:
        db.commit()
        db.refresh(employee)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    
    # Manually attach role for response if schema has it
    if employee:
        primary_user_role = db.query(UserRole).filter(
            UserRole.user_id == employee.id,
            UserRole.is_primary == True,
            UserRole.is_active == True
        ).first()

        if primary_user_role:
             employee.role = primary_user_role.role

    return EmployeeDetailResponse(
        success=True,
        message="Employee updated successfully",
        data=employee
    )

@router.delete("/{employee_id}", response_model=EmployeeDeleteResponse)
def delete_employee(
    employee_id: str,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    authorized: bool = Depends(deps.check_permission("4"))
):
    """
    Soft delete an employee.
    """
    employee = None
    
    # 1. Resolve Employee
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj,
            Employee.organization_id == current_org.id,
            Employee.is_deleted == False
        ).first()
    except ValueError:
        pass
        
    if not employee:
        try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(
                Employee.id == int_id,
                Employee.organization_id == current_org.id,
                Employee.is_deleted == False
            ).first()
        except ValueError:
            pass
            
    if not employee:
        # Code check
         employee = db.query(Employee).filter(
            Employee.employee_code == employee_id,
            Employee.organization_id == current_org.id,
            Employee.is_deleted == False
        ).first()
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # Check if this employee is a manager for active employees?
    # Optional business logic: prevent deletion if they manage others.
    # For now, allowing sort delete.
    
    try:
        employee.is_active = False # Also deactivate
        employee.is_deleted = True
        employee.deleted_at = datetime.utcnow()
        db.add(employee)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return EmployeeDeleteResponse(
        success=True,
        message="Employee deleted successfully",
        data=None
    )

@router.post("/{employee_uuid}/activate", response_model=EmployeeDetailResponse)
def activate_employee(
    employee_uuid: str,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Reactivate a deactivated employee.
    """
    employee = None
    
    # 1. Resolve Employee
    try:
        # uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == employee_uuid,
            Employee.organization_id == current_org.id,
            Employee.is_deleted == False
        ).first()
    except ValueError:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    try:
        if employee.is_active:
            employee.is_active = False
        else:
            employee.is_active = True
        db.add(employee)
        db.commit()
        db.refresh(employee)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return EmployeeDetailResponse(
        success=True,
        message="Employee activated successfully",
        data=employee
    )

@router.get("/{employee_id}/shift-roster", response_model=ShiftRosterListResponse)
def get_employee_shift_roster(
    employee_id: str,
    from_date: Optional[date] = Query(None, description="Filter from roster date"),
    to_date: Optional[date] = Query(None, description="Filter to roster date"),
    page: Optional[int] = Query(None, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, description="Items per page"),
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Get shift roster for a specific employee.
    """
    # 0. RBAC & Identity
    current_org_id = current_user.id if isinstance(current_user, Organization) else current_user.organization_id
    is_authorized_all = deps.has_permission(db, current_user, "25")

    # 1. Resolve employee
    try:
        if "-" in employee_id:
            employee = db.query(Employee).filter(Employee.uuid == employee_id, Employee.organization_id == current_org_id).first()
        else:
            employee = db.query(Employee).filter(Employee.id == int(employee_id), Employee.organization_id == current_org_id).first()
    except (ValueError, TypeError):
        employee = None

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # 2. Ownership & Permission check
    is_owner = isinstance(current_user, Employee) and current_user.id == employee.id
    
    if not (isinstance(current_user, Organization) or is_authorized_all or is_owner):
         raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not have permission to view this employee's shift roster."
        )

    # 3. Query roster
    query = db.query(ShiftRoster).filter(
        ShiftRoster.employee_id == employee.id,
        ShiftRoster.organization_id == current_org_id,
        ShiftRoster.is_deleted == False
    )

    # If employee viewing own roster, force published only
    if is_owner and not (isinstance(current_user, Organization) or is_authorized_all):
        query = query.filter(ShiftRoster.is_published == True)

    if from_date:
        query = query.filter(ShiftRoster.roster_date >= from_date)
    if to_date:
        query = query.filter(ShiftRoster.roster_date <= to_date)

    # 3. Pagination & Execution
    total_records = query.count()
    
    pagination_data = None
    if page is not None:
        total_pages = (total_records + limit - 1) // limit
        skip = (page - 1) * limit
        rosters = query.options(
            joinedload(ShiftRoster.shift)
        ).order_by(ShiftRoster.roster_date.asc()).offset(skip).limit(limit).all()
        pagination_data = {
            "total_records": total_records,
            "current_page": page,
            "total_pages": total_pages,
            "page_size": limit
        }
    else:
        rosters = query.options(
            joinedload(ShiftRoster.shift)
        ).order_by(ShiftRoster.roster_date.asc()).all()
        pagination_data = {
            "total_records": total_records,
            "current_page": 1,
            "total_pages": 1,
            "page_size": total_records if total_records > 0 else 0
        }

    if not rosters:
        return ShiftRosterListResponse(
            success=False,
            message="No roster records found for this employee",
            data=[]
        )

    return ShiftRosterListResponse(
        success=True,
        message="Employee roster retrieved successfully",
        data=rosters,
        pagination=pagination_data
    )

@router.get("/{employee_id}/leave-balances", response_model=LeaveBalanceListResponse)
def get_employee_leave_balances(
    employee_id: str,
    year: Optional[int] = Query(None, description="Filter by year (YYYY)"),
    leave_type_uuid: Optional[uuid.UUID] = Query(None, description="Filter by leave type UUID"),
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Get leave balances for a specific employee.
    """
    # 1. Resolve employee
    try:
        if "-" in employee_id:
            employee = db.query(Employee).filter(Employee.uuid == employee_id, Employee.organization_id == current_org.id).first()
        else:
            employee = db.query(Employee).filter(Employee.id == int(employee_id), Employee.organization_id == current_org.id).first()
    except (ValueError, TypeError):
        employee = None

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    # 2. RBAC/Identity Check
    # If Organization, bypass
    # If Employee, only allow if they are checking their own balance
    if isinstance(current_user, Employee) and current_user.id != employee.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied. You can only check your own leave balance."
        )

    # 3. Query Balance
    query = db.query(LeaveBalance).filter(
        LeaveBalance.employee_id == employee.id,
        LeaveBalance.organization_id == current_org.id
    )

    if year:
        query = query.filter(LeaveBalance.balance_year == year)
    else:
        # Default to current year if not specified
        current_year = datetime.utcnow().year
        query = query.filter(LeaveBalance.balance_year == current_year)

    if leave_type_uuid:
        # Resolve leave type
        leave_type = db.query(LeaveType).filter(
            LeaveType.uuid == leave_type_uuid,
            LeaveType.organization_id == current_org.id
        ).first()
        if leave_type:
            query = query.filter(LeaveBalance.leave_type_id == leave_type.id)
        else:
            # If leave type requested but not found, return empty results for safety
            return LeaveBalanceListResponse(
                success=True,
                message="Leave type not found",
                data=[]
            )

    balances = query.options(joinedload(LeaveBalance.leave_type)).all()

    return LeaveBalanceListResponse(
        success=True,
        message="Employee leave balances retrieved successfully",
        data=balances
    )

@router.get("/{employee_id}/leave-accrual-history", response_model=LeaveAccrualHistoryListResponse)
def get_employee_leave_accrual_history(
    employee_id: str,
    year: Optional[int] = Query(None, description="Filter by year (YYYY)"),
    leave_type_uuid: Optional[uuid.UUID] = Query(None, description="Filter by leave type UUID"),
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, description="Items per page"),
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Get leave accrual history for a specific employee.
    """
    # 1. Resolve employee
    try:
        if "-" in employee_id:
            employee = db.query(Employee).filter(Employee.uuid == employee_id, Employee.organization_id == current_org.id).first()
        else:
            employee = db.query(Employee).filter(Employee.id == int(employee_id), Employee.organization_id == current_org.id).first()
    except (ValueError, TypeError):
        employee = None

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # 2. RBAC/Identity Check
    # If Organization, bypass
    # If Employee, only allow if they are checking their own history
    if isinstance(current_user, Employee) and current_user.id != employee.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied. You can only check your own leave accrual history."
        )

    # 3. Query Accrual History
    query = db.query(LeaveAccrualHistory).filter(
        LeaveAccrualHistory.employee_id == employee.id
    )

    if year:
        # Filter by year using date range for better cross-DB compatibility
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        query = query.filter(LeaveAccrualHistory.accrual_date >= start_date, LeaveAccrualHistory.accrual_date <= end_date)

    if leave_type_uuid:
        # Resolve leave type
        leave_type = db.query(LeaveType).filter(
            LeaveType.uuid == leave_type_uuid,
            LeaveType.organization_id == current_org.id
        ).first()
        if leave_type:
            query = query.filter(LeaveAccrualHistory.leave_type_id == leave_type.id)
        else:
            return LeaveAccrualHistoryListResponse(
                success=True,
                message="Leave type not found",
                data=[]
            )

    # Pagination
    total_records = query.count()
    total_pages = (total_records + limit - 1) // limit
    skip = (page - 1) * limit

    history = query.options(joinedload(LeaveAccrualHistory.leave_type)).order_by(LeaveAccrualHistory.accrual_date.desc()).offset(skip).limit(limit).all()

    return LeaveAccrualHistoryListResponse(
        success=True,
        message="Accrual history retrieved successfully",
        data=history,
        pagination={
            "total_records": total_records,
            "current_page": page,
            "total_pages": total_pages,
            "page_size": limit
        }
    )

@router.get("/{employee_id}/leave-history", response_model=LeaveApplicationListResponse)
def get_employee_leave_history(
    employee_id: str,
    year: Optional[int] = Query(None, description="Filter by year (YYYY)"),
    leave_type_uuid: Optional[uuid.UUID] = Query(None, description="Filter by leave type UUID"),
    status: Optional[str] = Query(None, description="Filter by application status"),
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, description="Items per page"),
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Get leave application history for a specific employee.
    """
    # 1. Resolve employee
    try:
        if "-" in employee_id:
            employee = db.query(Employee).filter(Employee.uuid == employee_id, Employee.organization_id == current_org.id).first()
        else:
            employee = db.query(Employee).filter(Employee.id == int(employee_id), Employee.organization_id == current_org.id).first()
    except (ValueError, TypeError):
        employee = None

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # RBAC Check: Employee can only view their own history
    if isinstance(current_user, Employee) and current_user.id != employee.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied. You can only view your own leave history."
        )

    # 2. Query Leave Applications
    query = db.query(LeaveApplication).filter(
        LeaveApplication.employee_id == employee.id,
        LeaveApplication.organization_id == current_org.id
    )

    if year:
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        query = query.filter(LeaveApplication.from_date >= start_date, LeaveApplication.to_date <= end_date)

    if leave_type_uuid:
        leave_type = db.query(LeaveType).filter(
            LeaveType.uuid == leave_type_uuid,
            LeaveType.organization_id == current_org.id
        ).first()
        if leave_type:
            query = query.filter(LeaveApplication.leave_type_id == leave_type.id)
        else:
            return LeaveApplicationListResponse(
                success=True,
                message="Leave type not found",
                data=[],
                pagination={"total_records": 0, "current_page": 1, "total_pages": 0, "page_size": 10}
            )

    if status:
        query = query.filter(LeaveApplication.status == status)

    # Pagination
    total_records = query.count()
    total_pages = (total_records + limit - 1) // limit
    skip = (page - 1) * limit

    applications = query.options(
        joinedload(LeaveApplication.leave_type),
        joinedload(LeaveApplication.current_approver)
    ).order_by(LeaveApplication.created_at.desc()).offset(skip).limit(limit).all()

    return LeaveApplicationListResponse(
        success=True,
        message="Leave history retrieved successfully",
        data=applications,
        pagination={
            "total_records": total_records,
            "current_page": page,
            "total_pages": total_pages,
            "page_size": limit
        }
    )

@router.get("/{employee_id}/leave-calendar", response_model=LeaveCalendarResponse)
def get_employee_leave_calendar(
    employee_id: str,
    from_date: Optional[date] = Query(None, description="Start date of the range"),
    to_date: Optional[date] = Query(None, description="End date of the range"),
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Get leave applications for a specific employee in calendar format.
    """
    # 1. Resolve employee
    try:
        if "-" in employee_id:
            employee = db.query(Employee).filter(Employee.uuid == employee_id, Employee.organization_id == current_org.id).first()
        else:
            employee = db.query(Employee).filter(Employee.id == int(employee_id), Employee.organization_id == current_org.id).first()
    except (ValueError, TypeError):
        employee = None

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # RBAC Check: Employee can only view their own calendar
    if isinstance(current_user, Employee) and current_user.id != employee.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied. You can only view your own leave calendar."
        )

    # 2. Base Query
    query = db.query(LeaveApplication).filter(
        LeaveApplication.employee_id == employee.id,
        LeaveApplication.organization_id == current_org.id,
        LeaveApplication.status.in_([LeaveStatus.PENDING, LeaveStatus.APPROVED])
    )

    # 3. Date Range Filter
    if from_date and to_date:
        query = query.filter(
            or_(
                and_(LeaveApplication.from_date <= from_date, LeaveApplication.to_date >= from_date),
                and_(LeaveApplication.from_date <= to_date, LeaveApplication.to_date >= to_date),
                and_(LeaveApplication.from_date >= from_date, LeaveApplication.to_date <= to_date)
            )
        )
    elif from_date:
        query = query.filter(LeaveApplication.to_date >= from_date)
    elif to_date:
        query = query.filter(LeaveApplication.from_date <= to_date)

    # 4. Fetch Results
    applications = query.options(
        joinedload(LeaveApplication.leave_type)
    ).all()

    # 4.5 Fetch Holidays
    from sqlalchemy import JSON
    holiday_query = db.query(Holiday).filter(
        Holiday.organization_id == current_org.id,
        Holiday.is_active == True,
        Holiday.is_deleted == False
    )
    if from_date:
        holiday_query = holiday_query.filter(Holiday.holiday_date >= from_date)
    if to_date:
        holiday_query = holiday_query.filter(Holiday.holiday_date <= to_date)
    
    # Filter for Location/Department
    location_conditions = [Holiday.is_location_specific == False]
    if employee.location_id is not None:
        location_conditions.append(
            and_(
                Holiday.is_location_specific == True,
                Holiday.location_ids.isnot(None),
                func.json_contains(Holiday.location_ids, func.cast(employee.location_id, JSON)) == 1
            )
        )
    location_filter = or_(*location_conditions)
    
    department_conditions = [Holiday.is_department_specific == False]
    if employee.department_id is not None:
        department_conditions.append(
            and_(
                Holiday.is_department_specific == True,
                Holiday.department_ids.isnot(None),
                func.json_contains(Holiday.department_ids, func.cast(employee.department_id, JSON)) == 1
            )
        )
    department_filter = or_(*department_conditions)
    
    holiday_query = holiday_query.filter(location_filter, department_filter)
    holidays = holiday_query.all()

    # 5. Transform to Calendar Events
    leave_events = []
    for app in applications:
        leave_events.append(LeaveCalendarEvent(
            uuid=app.uuid,
            employee_name=f"{employee.first_name} {employee.last_name}",
            employee_uuid=employee.uuid,
            leave_type_name=app.leave_type.leave_name,
            from_date=app.from_date,
            to_date=app.to_date,
            status=app.status,
            is_half_day=app.is_half_day,
            total_days=app.total_days
        ))

    holiday_events = []
    for h in holidays:
        holiday_events.append(HolidayCalendarEvent(
            uuid=h.uuid,
            holiday_name=h.holiday_name,
            holiday_date=h.holiday_date,
            holiday_type=h.holiday_type,
            description=h.description,
            is_optional=h.is_optional,
            is_restricted=h.is_restricted
        ))

    return LeaveCalendarResponse(
        success=True,
        message="Employee leave calendar retrieved successfully",
        data=LeaveCalendarData(
            leaves=leave_events,
            holidays=holiday_events
        )
    )

@router.get("/{employee_id}/compensatory-offs", response_model=EmployeeCompOffListResponse)
def get_employee_compensatory_offs(
    employee_id: str,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    is_utilized: Optional[bool] = Query(None, description="Filter by utilization status"),
    is_expired: Optional[bool] = Query(None, description="Filter by expiry status"),
    sort_by: str = Query("worked_date", description="Sort by 'worked_date' or 'expiry_date'"),
    order: str = Query("desc", description="Sort order ('asc' or 'desc')"),
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, description="Items per page")
):
    """
    Get compensatory off balance and expiring credits for a specific employee.
    """
    # 1. Resolve employee
    employee = None
    try:
        if "-" in employee_id:
            employee = db.query(Employee).filter(Employee.uuid == employee_id, Employee.organization_id == current_org.id).first()
        else:
            employee = db.query(Employee).filter(Employee.id == int(employee_id), Employee.organization_id == current_org.id).first()
    except (ValueError, TypeError):
        pass
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # 2. RBAC Policy: Employee can only access their own summary
    if isinstance(current_user, Employee) and current_user.id != employee.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied. You can only view your own compensatory off summary."
        )

    # 2. Base Query
    base_query = db.query(CompensatoryOff).filter(
        CompensatoryOff.employee_id == employee.id,
        CompensatoryOff.organization_id == current_org.id
    )

    # 3. Calculate Summary (Scoping to ALL employee records before pagination/filtering)
    all_employee_records = base_query.all()

    total_earned = Decimal('0.0')
    total_utilized = Decimal('0.0')
    total_expired = Decimal('0.0')
    available_balance = Decimal('0.0')
    for r in all_employee_records:
        total_earned += r.comp_off_days
        total_utilized += r.utilized_days
        
        if r.is_expired or r.is_lapsed:
            total_expired += (r.comp_off_days - r.utilized_days)
        else:
            available_balance += r.remaining_days

    # 4. Apply Filters for the List view
    query = base_query
    if is_utilized is not None:
        query = query.filter(CompensatoryOff.is_utilized == is_utilized)
    if is_expired is not None:
        query = query.filter(CompensatoryOff.is_expired == is_expired)

    # 5. Sorting
    if sort_by == "expiry_date":
        if order.lower() == "asc":
            query = query.order_by(CompensatoryOff.expiry_date.asc())
        else:
            query = query.order_by(CompensatoryOff.expiry_date.desc())
    else: # Default: worked_date
        if order.lower() == "asc":
            query = query.order_by(CompensatoryOff.worked_date.asc())
        else:
            query = query.order_by(CompensatoryOff.worked_date.desc())

    # 6. Pagination
    total_records = query.count()
    total_pages = (total_records + limit - 1) // limit
    records = query.offset((page - 1) * limit).limit(limit).all()

    return EmployeeCompOffListResponse(
        success=True,
        message="Employee compensatory off summary and filtered records retrieved successfully",
        data=records,
        pagination={
            "total_records": total_records,
            "current_page": page,
            "total_pages": total_pages,
            "page_size": limit
        },
        summary={
            "total_earned": total_earned,
            "total_utilized": total_utilized,
            "total_expired": total_expired,
            "available_balance": available_balance
        }
    )

@router.get("/{employee_id}/optional-holidays", response_model=EmployeeOptionalHolidayListResponse)
def get_employee_optional_holidays(
    employee_id: str,
    year: Optional[int] = Query(None, description="Filter by year (YYYY)"),
    search: Optional[str] = Query(None, description="Search by holiday name"),
    sort_by: str = Query("holiday_date", description="Sort by: holiday_name, holiday_date"),
    order: str = Query("asc", description="Sort order: asc, desc"),
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, description="Items per page"),
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Get optional/restricted holidays selected by a specific employee.
    """
    # 1. Fetch Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj,
            Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
        
    if not employee:
        try:
            employee = db.query(Employee).filter(
                Employee.id == int(employee_id),
                Employee.organization_id == current_org.id
            ).first()
        except ValueError:
            pass

    if not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Employee not found"
        )

    # 1.5 RBAC Policy: Employee can only access their own optional holidays
    if isinstance(current_user, Employee) and current_user.id != employee.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied. You can only view your own holiday selections."
        )

    # 2. Query Selections
    query = db.query(OptionalHolidaySelection).filter(
        OptionalHolidaySelection.employee_id == employee.id
    ).join(Holiday).options(joinedload(OptionalHolidaySelection.holiday))

    if year:
        query = query.filter(OptionalHolidaySelection.selection_year == year)

    if search:
        search_term = f"%{search}%"
        query = query.filter(Holiday.holiday_name.ilike(search_term))
    
    # 3. Sorting
    sort_column = getattr(Holiday, sort_by, Holiday.holiday_date)
    if order.lower() == "desc":
        query = query.order_by(sort_column.desc())
    else:
        query = query.order_by(sort_column.asc())
    
    # 4. Pagination
    total_records = query.count()
    total_pages = (total_records + limit - 1) // limit
    skip = (page - 1) * limit
    selections = query.offset(skip).limit(limit).all()

    return EmployeeOptionalHolidayListResponse(
        success=True,
        message="Optional holidays retrieved successfully",
        data=selections,
        pagination={
            "total_records": total_records,
            "current_page": page,
            "total_pages": total_pages,
            "page_size": limit
        }
    )

@router.get("/{employee_id}/holidays", response_model=HolidayListResponse)
def get_employee_holidays(
    employee_id: str,
    year: Optional[int] = Query(None, description="Filter by year (YYYY)"),
    search: Optional[str] = Query(None, description="Search by holiday name"),
    sort_by: str = Query("holiday_date", description="Sort by: holiday_name, holiday_date"),
    order: str = Query("asc", description="Sort order: asc, desc"),
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, description="Items per page"),
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Get all holidays (mandatory and optional) applicable to a specific employee.
    """
    # 1. Resolve employee
    employee = None
    try:
        if "-" in employee_id:
            employee = db.query(Employee).filter(Employee.uuid == employee_id, Employee.organization_id == current_org.id).first()
        else:
            employee = db.query(Employee).filter(Employee.id == int(employee_id), Employee.organization_id == current_org.id).first()
    except (ValueError, TypeError):
        pass
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # 2. RBAC Policy: Employee can only access their own holiday list
    if isinstance(current_user, Employee) and current_user.id != employee.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied. You can only view your own applicable holidays."
        )

    # 3. Build Holiday Query
    query = db.query(Holiday).filter(
        Holiday.organization_id == current_org.id,
        Holiday.is_active == True,
        Holiday.is_deleted == False
    )

    if year:
        query = query.filter(Holiday.holiday_year == year)

    # 4. Filter by Location/Department logic
    # Holidays are applicable if they are NOT specific OR if the employee's ID is in the specific list
    from sqlalchemy import func
    
    # Filter for Location
    from sqlalchemy import JSON
    location_conditions = [Holiday.is_location_specific == False]
    if employee.location_id is not None:
        location_conditions.append(
            and_(
                Holiday.is_location_specific == True,
                Holiday.location_ids.isnot(None),
                func.json_contains(Holiday.location_ids, func.cast(employee.location_id, JSON)) == 1
            )
        )
    location_filter = or_(*location_conditions)
    
    # Filter for Department
    department_conditions = [Holiday.is_department_specific == False]
    if employee.department_id is not None:
        department_conditions.append(
            and_(
                Holiday.is_department_specific == True,
                Holiday.department_ids.isnot(None),
                func.json_contains(Holiday.department_ids, func.cast(employee.department_id, JSON)) == 1
            )
        )
    department_filter = or_(*department_conditions)

    query = query.filter(location_filter, department_filter)

    if search:
        search_term = f"%{search}%"
        query = query.filter(Holiday.holiday_name.ilike(search_term))

    # 4. Sorting
    sort_column = getattr(Holiday, sort_by, Holiday.holiday_date)
    if order.lower() == "desc":
        query = query.order_by(sort_column.desc())
    else:
        query = query.order_by(sort_column.asc())

    # 5. Pagination
    total_records = query.count()
    total_pages = (total_records + limit - 1) // limit
    skip = (page - 1) * limit
    holidays = query.offset(skip).limit(limit).all()

    return HolidayListResponse(
        success=True,
        message="Applicable holidays retrieved successfully",
        data=holidays,
        pagination={
            "total_records": total_records,
            "current_page": page,
            "total_pages": total_pages,
            "page_size": limit
        }
    )

@router.get("/{employee_id}/overtime-summary", response_model=OvertimeSummaryResponse)
def get_employee_overtime_summary(
    employee_id: str,
    from_date: Optional[date] = Query(None, description="Filter from attendance date"),
    to_date: Optional[date] = Query(None, description="Filter to attendance date"),
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Get overtime summary for a specific employee.
    """
    current_org_id = current_user.id if isinstance(current_user, Organization) else current_user.organization_id
    
    # 1. Fetch Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj,
            Employee.organization_id == current_org_id
        ).first()
    except ValueError:
        pass
        
    if not employee:
        try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(
                Employee.id == int_id,
                Employee.organization_id == current_org_id
            ).first()
        except ValueError:
            pass
            
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # RBAC: Only logged-in user can access their own summary
    if not isinstance(current_user, Organization):
        # If current_user is an Employee, they can only access their own summary
        if employee.id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only access your own overtime summary."
            )
    # Organization/Admin access is allowed as they pass the above check by bypass

    # 2. Build Query
    query = db.query(OvertimeRequest).filter(
        OvertimeRequest.employee_id == employee.id,
        OvertimeRequest.organization_id == current_org_id
    )
    
    if from_date:
        query = query.filter(OvertimeRequest.attendance_date >= from_date)
    if to_date:
        query = query.filter(OvertimeRequest.attendance_date <= to_date)
        
    overtime_requests = query.all()
    
    # 3. Calculate Summary
    total_hours = Decimal('0.00')
    paid_hours = Decimal('0.00')
    comp_off_hours = Decimal('0.00')
    pending_hours = Decimal('0.00')
    approved_hours = Decimal('0.00')
    rejected_hours = Decimal('0.00')
    
    for req in overtime_requests:
        hours = Decimal(str(req.overtime_hours))
        
        if req.status == OvertimeStatus.PENDING:
            pending_hours += hours
        elif req.status in [OvertimeStatus.APPROVED, OvertimeStatus.PAID, OvertimeStatus.COMP_OFF]:
            approved_hours += hours
            total_hours += hours
            
            # Determine if it was paid or comp-off
            # Status takes precedence if it's explicitly PAID or COMP_OFF
            if req.status == OvertimeStatus.PAID:
                paid_hours += hours
            elif req.status == OvertimeStatus.COMP_OFF:
                comp_off_hours += hours
            else:
                # Use compensation_type field
                if req.compensation_type == CompensationType.paid:
                    paid_hours += hours
                elif req.compensation_type == CompensationType.comp_off:
                    comp_off_hours += hours
                elif req.compensation_type == CompensationType.both:
                    paid_hours += hours / 2
                    comp_off_hours += hours / 2
        elif req.status == OvertimeStatus.REJECTED:
            rejected_hours += hours
            
    summary = {
        "total_hours": total_hours,
        "paid_hours": paid_hours,
        "comp_off_hours": comp_off_hours,
        "pending_hours": pending_hours,
        "approved_hours": approved_hours,
        "rejected_hours": rejected_hours,
        "request_count": len(overtime_requests)
    }
    
    return OvertimeSummaryResponse(
        success=True,
        message="Overtime summary retrieved successfully",
        data=summary
    )



@router.post("/bulk-import", response_model=EmployeeImportResponse)
async def bulk_import_employees(
    file: UploadFile = File(...),
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    authorized: bool = Depends(deps.check_permission("2"))
):
    """
    Bulk import employees from CSV.
    Input columns expected: first_name, last_name, work_email, mobile_number, gender, 
    employment_type, employment_status, date_of_joining, date_of_confirmation, 
    job_title_code, department_code, location_code, reporting_manager_email (optional).
    """
    
    
    # Supported formats
    fname = file.filename.lower()
    if not (fname.endswith('.csv') or fname.endswith('.xlsx')):
        raise HTTPException(status_code=400, detail="Only CSV and Excel (.xlsx) files are supported.")

    # 1. Read File Content Once
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="The uploaded file is empty.")

    normalized_rows = []
    
    def map_fuzzy_header(h_raw):
        if not h_raw: return ""
        h = str(h_raw).strip().lower().replace('\t', ' ') # Replace tabs with space for mapping
        if "first" in h and ("name" in h or "nam" in h): return "first_name"
        if "last" in h and ("name" in h or "nam" in h): return "last_name"
        if "work" in h and ("email" in h or "ema" in h): return "work_email"
        if ("mobile" in h or "phone" in h) and ("num" in h or "mobile" in h): return "mobile_number"
        if "gender" in h: return "gender"
        if "birth" in h or "dob" in h: return "date_of_birth"
        if "employ" in h and "type" in h: return "employment_type"
        if "status" in h: return "employment_status"
        if "join" in h or "doj" in h: return "date_of_joining"
        if "confirm" in h or "doc" in h or "confir" in h: return "date_of_confirmation"
        if "person" in h and ("email" in h or "ema" in h): return "personal_email"
        if "title" in h or "designa" in h or "job" in h: 
            return "job_title_code"
        if "depart" in h or "dept" in h: return "department_code"
        if "locat" in h: return "location_code"
        if "manag" in h or "report" in h: return "reporting_manager_email"
        if "role" in h: return "role_code"
        return h.replace(' ', '_')

    # 2. Parse and Normalize Headers (Fuzzy Mapping)
    if fname.endswith('.csv'):
        try:
            decoded = content.decode('utf-8-sig')
            
            # Detect delimiter
            dialect = csv.excel
            try:
                sniffer = csv.Sniffer()
                # Use a portion of the file for sniffing to avoid issues with large files
                sample = decoded[:4096]
                if sniffer.has_header(sample):
                    dialect = sniffer.sniff(sample)
            except:
                # Default to comma if sniffing fails
                pass
                
            f = io.StringIO(decoded)
            reader = csv.reader(f, dialect=dialect)
            
            # Find first non-empty row for headers
            raw_headers = []
            rows_data = []
            for r in reader:
                if any(str(v).strip() for v in r if v is not None):
                    if not raw_headers:
                        raw_headers = r
                    else:
                        rows_data.append(r)
            
            headers = [map_fuzzy_header(h) for h in raw_headers]
            
            for row_vals in rows_data:
                r_dict = {}
                for i, val in enumerate(row_vals):
                    if i < len(headers) and headers[i]:
                         r_dict[headers[i]] = val
                normalized_rows.append(r_dict)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error parsing CSV: {str(e)}")
            
    else: # Excel
        try:
            wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True)
            # Use 'Template' sheet if available
            ws = wb['Template'] if 'Template' in wb.sheetnames else wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                raise HTTPException(status_code=400, detail="Empty Excel file")
                
            # Find first non-empty row for headers
            raw_headers = []
            header_row_idx = -1
            for idx, r in enumerate(all_rows):
                if any(str(v).strip() for v in r if v is not None):
                    raw_headers = r
                    header_row_idx = idx
                    break
            
            if header_row_idx == -1:
                raise HTTPException(status_code=400, detail="No headers found in Excel file")
                
            headers = [map_fuzzy_header(h) for h in raw_headers]
            
            for r in all_rows[header_row_idx+1:]:
                # Skip truly empty or whitespace-only rows
                if not any(str(v).strip() for v in r if v is not None):
                    continue
                    
                r_dict = {}
                for i, val in enumerate(r):
                    if i < len(headers) and headers[i]:
                         # Handle Dates and convert everything to string for consistency
                         if val is None:
                             v_str = ''
                         elif isinstance(val, (datetime, date)): 
                             v_str = val.isoformat()
                             if isinstance(val, datetime): v_str = val.date().isoformat()
                         else:
                             # Handle Excel numeric formatting (e.g., 4 becomes 4.0)
                             if isinstance(val, float) and val.is_integer():
                                 v_str = str(int(val))
                             else:
                                 v_str = str(val).strip()
                         r_dict[headers[i]] = v_str
                normalized_rows.append(r_dict)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error parsing Excel: {str(e)}")

    # 3. Main Processing Loop
    success_count = 0
    error_count = 0
    errors = []
    row_num = 0

    for row in normalized_rows:
        row_num += 1
        row_error = None
        
        try:
            # 1. Clean keys and values
            row = {str(k): str(v).strip() for k, v in row.items()}
            
            # 2. Mandatory Fields Check
            required_fields = [
                'first_name', 'last_name', 'work_email', 'mobile_number', 'gender', 'date_of_birth',
                'employment_type', 'employment_status', 'date_of_joining', 'date_of_confirmation',
                'personal_email', 'job_title_code', 'department_code', 'location_code'
            ]
            missing = [f for f in required_fields if f not in row]
            if missing:
                found_keys = list(row.keys())
                raise ValueError(f"Missing fields: {', '.join(missing)}. Found columns: {', '.join(found_keys)}")
                
            # 3. Validation & Parsing
            # Enums values will be validated by Pydantic Model
            pass
                
            # Dates
            # Pydantic handles ISO format (YYYY-MM-DD). If different format, need parsing.
            # Assuming ISO for now.
            
            # 4. Resolve FKs (Job Title, Dept, Loc)
            # Job Title
            job_title_id = None
            if 'job_title_code' in row:
                jt = db.query(JobTitle).filter(
                    JobTitle.title_code == row['job_title_code'], 
                    JobTitle.organization_id == current_org.id
                ).first()
                if not jt: raise ValueError(f"Invalid Job Title Code: {row['job_title_code']}")
                job_title_id = jt.uuid
                
            # Department
            department_id = None
            if 'department_code' in row:
                dept = db.query(Department).filter(
                    Department.department_code == row['department_code'],
                    Department.organization_id == current_org.id
                ).first()
                if not dept:
                    dept = db.query(Department).filter(
                        Department.department_name == row['department_code'], # Treating as Name
                        Department.organization_id == current_org.id
                    ).first()
                if not dept: raise ValueError(f"Invalid Department Code/Name: {row.get('department_code')}")
                department_id = dept.uuid
                
            # Location
            location_id = None
            if 'location_code' in row:
                loc = db.query(Location).filter(
                    Location.location_code == row['location_code'],
                    Location.organization_id == current_org.id
                ).first()
                if not loc: raise ValueError(f"Invalid Location Code: {row['location_code']}")
                location_id = loc.uuid
                
            # Manager
            manager_id = None
            if row.get('reporting_manager_email'):
                mgr = db.query(Employee).filter(
                    Employee.work_email == row['reporting_manager_email'],
                    Employee.organization_id == current_org.id
                ).first()
                if not mgr: 
                    raise ValueError(f"Manager email not found: {row['reporting_manager_email']}")
                manager_id = mgr.uuid
            
            # Role
            role_uuid = None
            if row.get('role_code'):
                role = db.query(Role).filter(
                    Role.role_code == row['role_code'],
                    Role.organization_id == current_org.id
                ).first()
                if not role:
                    raise ValueError(f"Invalid Role Code: {row['role_code']}")
                role_uuid = role.uuid
            else:
                # Assign default role for organization
                default_role = db.query(Role).filter(
                    Role.organization_id == current_org.id,
                    Role.is_active == True
                ).order_by(Role.is_default.desc()).first()
                if not default_role:
                    # Fallback to 'Employee' role name if no default flag is set
                    default_role = db.query(Role).filter(
                        Role.organization_id == current_org.id,
                        Role.role_name.ilike("Employee"),
                        Role.is_active == True
                    ).first()
                if default_role:
                    role_uuid = default_role.uuid
                else:
                    raise ValueError("No default role found for organization. Please specify a Role Code.")
            
            # 5. Construct EmployeeCreate
            # We construct dict then validate with Pydantic
            emp_data = {
                "first_name": row['first_name'],
                "last_name": row['last_name'],
                "work_email": row['work_email'],
                "mobile_number": row['mobile_number'],
                "gender": row['gender'],
                "date_of_birth": row.get('date_of_birth'),
                "employment_type": row['employment_type'],
                "employment_status": row['employment_status'],
                "date_of_joining": row['date_of_joining'],
                "date_of_confirmation": row.get('date_of_confirmation'),
                "personal_email": row.get('personal_email'),
                # if missing, let Pydantic raise error
                
                "job_title_id": job_title_id,
                "department_id": department_id,
                "location_id": location_id,
                "reporting_manager_id": str(manager_id) if manager_id else None,
                "role_uuid": role_uuid
            }
            
            # Generate Code if needed? create_employee handles it.
            # But we are calling create_employee logic? create_employee is an endpoint function.
            # We should reuse the logic or call function directly?
            # Calling endpoint function `create_employee(emp_in, ...)` is possible but tricky with deps.
            # Better to instantiate EmployeeCreate and use DB.
            
            # Validation
            emp_in = EmployeeCreate(**emp_data)
            
            # Uniqueness Check (Email)
            if db.query(Employee).filter(Employee.work_email == emp_in.work_email, Employee.organization_id == current_org.id).first():
                raise ValueError("Work email already exists")
            
            # Helper to generate code (duplicate of create logic - should refactor to service method, but inline for now)
            max_id = db.query(func.max(Employee.id)).scalar() or 0
            final_emp_code = f"EMP-{max_id + 1:04d}" # +success_count to increment in same batch? No, DB autoinc ID is not reserved yet.
            # Only reliable way is to flush or use UUIDs or re-query max_id every time (slow).
            # I will query max_id every time or simple increment.
            # Race condition risk in concurrent imports.
            
            # Create Object
            # Need to map UUIDs back to IDs for Model creation
            # (Assuming we have IDs from objects queried earlier)
            
            # Re-querying IDs?
            # `emp_in` has UUIDs.
            # I can just use the objects I queried: `jt`, `dept`, `loc`, `mgr`.
            # But I need to handle if they are mandatory or optional.
            # EmployeeCreate validators run on `emp_in`.
            
            # Model creation
            
            # Resolving IDs (Int) from the UUIDs in emp_in
            jt = db.query(JobTitle).filter(JobTitle.uuid == emp_in.job_title_id).first()
            dp = db.query(Department).filter(Department.uuid == emp_in.department_id).first()
            lo = db.query(Location).filter(Location.uuid == emp_in.location_id).first()
            mg_id = None
            if emp_in.reporting_manager_id:
                mg = db.query(Employee).filter(Employee.uuid == emp_in.reporting_manager_id).first()
                if mg: mg_id = mg.id

            db_obj = Employee(
                **emp_in.model_dump(exclude={'job_title_id', 'department_id', 'location_id', 'reporting_manager_id', 'employee_code', 'role_uuid'}),
                employee_code=final_emp_code,
                organization_id=current_org.id,
                job_title_id=jt.id,
                department_id=dp.id,
                location_id=lo.id,
                reporting_manager_id=mg_id
            )
            
            db.add(db_obj)
            db.flush() # Get id

            # Role Assignment
            if emp_in.role_uuid:
                role = db.query(Role).filter(Role.uuid == emp_in.role_uuid).first()
                if role:
                    user_role = UserRole(
                        user_id=db_obj.id,
                        role_id=role.id,
                        is_primary=True,
                        is_active=True,
                        valid_from=datetime.utcnow()
                    )
                    db.add(user_role)

            db.commit()
            success_count += 1
            
        except Exception as e:
            db.rollback()
            error_count += 1
            errors.append(EmployeeImportError(row=row_num, error=str(e)))
            
    return EmployeeImportResponse(
        success=True,
        message="Import processing completed",
        success_count=success_count,
        error_count=error_count,
        errors=errors
    )

@router.post("/{employee_id}/personal-info", response_model=PersonalInfoResponse)
def create_personal_info(
    employee_id: str,
    info_in: PersonalInfoCreate,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Create personal information record for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 2 (Create)
    if not deps.has_permission(db, current_user, "2"):
        # Case 3: Self-access for create
        if not (isinstance(current_user, Employee) and current_user.id == employee.id):
             raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="You do not have permission to create personal information for this employee"
            )
        
    # Check if exists
    existing_info = db.query(EmployeePersonalInfo).filter(EmployeePersonalInfo.employee_id == employee.id).first()
    if existing_info:
        raise HTTPException(status_code=400, detail="Personal information record already exists. Use PUT to update.")
        
    # Create
    db_obj = EmployeePersonalInfo(
        employee_id=employee.id,
        **info_in.model_dump(exclude_unset=True)
    )
    db.add(db_obj)
    try:
        db.commit()
        db.refresh(db_obj)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return PersonalInfoResponse(
        success=True,
        message="Personal information created successfully",
        data=db_obj
    )

@router.put("/{employee_id}/personal-info", response_model=PersonalInfoResponse)
def update_personal_info(
    employee_id: str,
    info_in: PersonalInfoUpdate,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Update personal information for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 3 (Update)
    if not deps.has_permission(db, current_user, "3"):
        # Even for self, user said: "once details add I can't update my self" without permission
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="You do not have permission to update personal information"
        )
        
    # Check if exists
    db_obj = db.query(EmployeePersonalInfo).filter(EmployeePersonalInfo.employee_id == employee.id).first()
    if not db_obj:
        raise HTTPException(status_code=404, detail="Personal information record not found. Use POST to create.")
        
    # Update
    update_data = info_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_obj, field, value)
        
    try:
        db.commit()
        db.refresh(db_obj)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return PersonalInfoResponse(
        success=True,
        message="Personal information updated successfully",
        data=db_obj
    )

@router.get("/{employee_id}/personal-info", response_model=PersonalInfoResponse)
def get_personal_info(
    employee_id: str,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Get personal information for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 1 (Read)
    if not deps.has_permission(db, current_user, "1"):
        # Case 3: Self-access for read
        if not (isinstance(current_user, Employee) and current_user.id == employee.id):
             raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="You do not have permission to view this employee's personal information"
            )
        
    # Query Info
    info = db.query(EmployeePersonalInfo).filter(EmployeePersonalInfo.employee_id == employee.id).first()
    
    if not info:
        raise HTTPException(status_code=404, detail="Personal information not found")
        
    return PersonalInfoResponse(
        success=True,
        message="Personal information retrieved successfully",
        data=info
    )

@router.post("/{employee_id}/addresses", response_model=AddressResponse)
def create_employee_address(
    employee_id: str,
    address_in: AddressCreate,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Add a new address for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 2 (Create)
    if not deps.has_permission(db, current_user, "2"):
        # No self-access for create in address module as per request
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="You do not have permission to add addresses for this employee"
        )
    
    # Check if duplicate type exists (e.g., only one Current address allowed?)
    # User didn't specify constraint, but 'address_type' usually implies uniqueness per type or multiple.
    # Model doesn't enforce unique type. I'll allow multiple for now unless it conflicts.
    
    # Create Address
    db_obj = EmployeeAddress(
        employee_id=employee.id,
        **address_in.model_dump()
    )
    db.add(db_obj)
    try:
        db.commit()
        db.refresh(db_obj)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return AddressResponse(
        success=True,
        message="Address created successfully",
        data=db_obj
    )

@router.put("/{employee_id}/addresses/{address_id}", response_model=AddressResponse)
def update_employee_address(
    employee_id: str,
    address_id: int,
    address_in: AddressUpdate,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Update an address for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 3 (Update)
    if not deps.has_permission(db, current_user, "3"):
        # No self-access for update in address module
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="You do not have permission to update addresses"
        )
        
    # 3. Resolve Address
    address = db.query(EmployeeAddress).filter(
        EmployeeAddress.id == address_id,
        EmployeeAddress.employee_id == employee.id
    ).first()
    
    if not address:
        raise HTTPException(status_code=404, detail="Address not found")
        
    # 4. Update
    update_data = address_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(address, field, value)
        
    try:
        db.commit()
        db.refresh(address)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return AddressResponse(
        success=True,
        message="Address updated successfully",
        data=address
    )

@router.get("/{employee_id}/addresses", response_model=AddressListResponse)
def get_employee_addresses(
    employee_id: str,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Get all addresses for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 1 (Read)
    if not deps.has_permission(db, current_user, "1"):
        # Case 3: Self-access for read
        if not (isinstance(current_user, Employee) and current_user.id == employee.id):
             raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="You do not have permission to view these addresses"
            )
            
    # Query Addresses
    addresses = db.query(EmployeeAddress).filter(EmployeeAddress.employee_id == employee.id).all()
    
    return AddressListResponse(
        success=True,
        message="Addresses retrieved successfully",
        data=addresses
    )

@router.delete("/{employee_id}/addresses/{address_id}", response_model=AddressResponse)
def delete_employee_address(
    employee_id: str,
    address_id: int,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Delete an address for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 4 (Delete)
    if not deps.has_permission(db, current_user, "4"):
        # No self-access for delete in address module
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="You do not have permission to delete addresses"
        )
        
    # 3. Resolve Address
    address = db.query(EmployeeAddress).filter(
        EmployeeAddress.id == address_id,
        EmployeeAddress.employee_id == employee.id
    ).first()
    
    if not address:
        raise HTTPException(status_code=404, detail="Address not found")
        
    # 4. Delete
    try:
        db.delete(address)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return AddressResponse(
        success=True,
        message="Address deleted successfully",
        data=None
    )

@router.patch("/{employee_id}/addresses/{address_id}/set-primary", response_model=AddressResponse)
def set_primary_address(
    employee_id: str,
    address_id: int,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Set an address as primary for an employee. Unsets primary for all other addresses.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 3 (Update)
    if not deps.has_permission(db, current_user, "3"):
        # No self-access for update in address module
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="You do not have permission to update addresses"
        )
        
    # 3. Resolve Address
    address = db.query(EmployeeAddress).filter(
        EmployeeAddress.id == address_id,
        EmployeeAddress.employee_id == employee.id
    ).first()
    
    if not address:
        raise HTTPException(status_code=404, detail="Address not found")
        
    # 3. Update Primary Status
    try:
        # Unset others
        db.query(EmployeeAddress).filter(
            EmployeeAddress.employee_id == employee.id,
            EmployeeAddress.id != address_id
        ).update({"is_primary": False})
        
        # Set target
        address.is_primary = True
        
        db.commit()
        db.refresh(address)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return AddressResponse(
        success=True,
        message="Address set as primary successfully",
        data=address
    )

@router.post("/{employee_id}/emergency-contacts", response_model=EmergencyContactResponse)
def create_emergency_contact(
    employee_id: str,
    contact_in: EmergencyContactCreate,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Add a new emergency contact for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 2 (Create)
    if not deps.has_permission(db, current_user, "2"):
        # No self-access for create in emergency contact module
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="You do not have permission to add emergency contacts for this employee"
        )
        
    # Create Contact
    db_obj = EmployeeEmergencyContact(
        employee_id=employee.id,
        **contact_in.model_dump()
    )
    db.add(db_obj)
    try:
        db.commit()
        db.refresh(db_obj)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return EmergencyContactResponse(
        success=True,
        message="Emergency contact created successfully",
        data=db_obj
    )

@router.put("/{employee_id}/emergency-contacts/{contact_id}", response_model=EmergencyContactResponse)
def update_emergency_contact(
    employee_id: str,
    contact_id: int,
    contact_in: EmergencyContactUpdate,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Update an emergency contact for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 3 (Update)
    if not deps.has_permission(db, current_user, "3"):
        # No self-access for update in emergency contact module
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="You do not have permission to update emergency contacts"
        )
        
    # 3. Resolve Contact
    contact = db.query(EmployeeEmergencyContact).filter(
        EmployeeEmergencyContact.id == contact_id,
        EmployeeEmergencyContact.employee_id == employee.id
    ).first()
    
    if not contact:
        raise HTTPException(status_code=404, detail="Emergency contact not found")
        
    # 4. Update
    update_data = contact_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(contact, field, value)
        
    try:
        db.commit()
        db.refresh(contact)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return EmergencyContactResponse(
        success=True,
        message="Emergency contact updated successfully",
        data=contact
    )

@router.get("/{employee_id}/emergency-contacts", response_model=EmergencyContactListResponse)
def get_emergency_contacts(
    employee_id: str,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Get all emergency contacts for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 1 (Read)
    if not deps.has_permission(db, current_user, "1"):
        # Case 3: Self-access for read
        if not (isinstance(current_user, Employee) and current_user.id == employee.id):
             raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="You do not have permission to view these emergency contacts"
            )
            
    # Query Contacts
    contacts = db.query(EmployeeEmergencyContact).filter(
        EmployeeEmergencyContact.employee_id == employee.id
    ).all()
    
    return EmergencyContactListResponse(
        success=True,
        message="Emergency contacts retrieved successfully",
        data=contacts
    )

@router.delete("/{employee_id}/emergency-contacts/{contact_id}", response_model=EmergencyContactResponse)
def delete_emergency_contact(
    employee_id: str,
    contact_id: int,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Delete an emergency contact for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 4 (Delete)
    if not deps.has_permission(db, current_user, "4"):
        # No self-access for delete in emergency contact module
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="You do not have permission to delete emergency contacts"
        )
        
    # 3. Resolve Contact
    contact = db.query(EmployeeEmergencyContact).filter(
        EmployeeEmergencyContact.id == contact_id,
        EmployeeEmergencyContact.employee_id == employee.id
    ).first()
    
    if not contact:
        raise HTTPException(status_code=404, detail="Emergency contact not found")
        
    # 4. Delete
    try:
        db.delete(contact)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return EmergencyContactResponse(
        success=True,
        message="Emergency contact deleted successfully",
        data=None
    )

@router.patch("/{employee_id}/emergency-contacts/{contact_id}/set-primary", response_model=EmergencyContactResponse)
def set_primary_emergency_contact(
    employee_id: str,
    contact_id: int,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Set an emergency contact as primary for an employee. Unsets primary for all other contacts.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 3 (Update)
    if not deps.has_permission(db, current_user, "3"):
        # No self-access for update in emergency contact module
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="You do not have permission to update emergency contacts"
        )
        
    # 3. Resolve Contact
    contact = db.query(EmployeeEmergencyContact).filter(
        EmployeeEmergencyContact.id == contact_id,
        EmployeeEmergencyContact.employee_id == employee.id
    ).first()
    
    if not contact:
        raise HTTPException(status_code=404, detail="Emergency contact not found")
        
    # 3. Update Primary Status
    try:
        # Unset others
        db.query(EmployeeEmergencyContact).filter(
            EmployeeEmergencyContact.employee_id == employee.id,
            EmployeeEmergencyContact.id != contact_id
        ).update({"is_primary": False})
        
        # Set target
        contact.is_primary = True
        
        db.commit()
        db.refresh(contact)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return EmergencyContactResponse(
        success=True,
        message="Emergency contact set as primary successfully",
        data=contact
    )

@router.post("/{employee_id}/education", response_model=EducationResponse)
def create_education(
    employee_id: str,
    education_in: EducationCreate,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Add a new education record for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 2 (Create)
    if not deps.has_permission(db, current_user, "2"):
        # Case 3: Self-access for create
        if not (isinstance(current_user, Employee) and current_user.id == employee.id):
             raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="You do not have permission to add education record for this employee"
            )
            
    # Create Education
    db_obj = EmployeeEducation(
        employee_id=employee.id,
        **education_in.model_dump()
    )
    db.add(db_obj)
    try:
        db.commit()
        db.refresh(db_obj)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return EducationResponse(
        success=True,
        message="Education record created successfully",
        data=db_obj
    )

@router.put("/{employee_id}/education/{education_id}", response_model=EducationResponse)
def update_education(
    employee_id: str,
    education_id: int,
    education_in: EducationUpdate,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Update an education record for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 3 (Update)
    if not deps.has_permission(db, current_user, "3"):
        # Case 3: Self-access for update
        if not (isinstance(current_user, Employee) and current_user.id == employee.id):
             raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="You do not have permission to update education records"
            )
            
    # 3. Resolve Education
    education = db.query(EmployeeEducation).filter(
        EmployeeEducation.id == education_id,
        EmployeeEducation.employee_id == employee.id
    ).first()
    
    if not education:
        raise HTTPException(status_code=404, detail="Education record not found")
        
    # 4. Update
    update_data = education_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(education, field, value)
        
    try:
        db.commit()
        db.refresh(education)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return EducationResponse(
        success=True,
        message="Education record updated successfully",
        data=education
    )

@router.get("/{employee_id}/education", response_model=EducationListResponse)
def get_education(
    employee_id: str,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Get educational qualifications for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 1 (Read)
    if not deps.has_permission(db, current_user, "1"):
        # Case 3: Self-access for read
        if not (isinstance(current_user, Employee) and current_user.id == employee.id):
             raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="You do not have permission to view these education records"
            )
            
    # Query Education
    education_records = db.query(EmployeeEducation).filter(
        EmployeeEducation.employee_id == employee.id
    ).all()
    
    return EducationListResponse(
        success=True,
        message="Education records retrieved successfully",
        data=education_records
    )

@router.delete("/{employee_id}/education/{education_id}", response_model=EducationResponse)
def delete_education(
    employee_id: str,
    education_id: int,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Delete an education record for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 4 (Delete)
    if not deps.has_permission(db, current_user, "4"):
        # Case 3: Self-access for delete
        if not (isinstance(current_user, Employee) and current_user.id == employee.id):
             raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="You do not have permission to delete education records"
            )
            
    # 3. Resolve Education
    education = db.query(EmployeeEducation).filter(
        EmployeeEducation.id == education_id,
        EmployeeEducation.employee_id == employee.id
    ).first()
    
    if not education:
        raise HTTPException(status_code=404, detail="Education record not found")
        
    # 4. Delete
    try:
        db.delete(education)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return EducationResponse(
        success=True,
        message="Education record deleted successfully",
        data=None
    )

@router.post("/{employee_id}/certifications", response_model=CertificationResponse)
def create_certification(
    employee_id: str,
    cert_in: CertificationCreate,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Add a new certification for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 2 (Create)
    if not deps.has_permission(db, current_user, "2"):
        # Case 3: Self-access for create
        if not (isinstance(current_user, Employee) and current_user.id == employee.id):
             raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="You do not have permission to add certification record for this employee"
            )
            
    # Create Certification
    db_obj = EmployeeCertification(
        employee_id=employee.id,
        **cert_in.model_dump()
    )
    db.add(db_obj)
    try:
        db.commit()
        db.refresh(db_obj)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return CertificationResponse(
        success=True,
        message="Certification created successfully",
        data=db_obj
    )

@router.put("/{employee_id}/certifications/{certification_id}", response_model=CertificationResponse)
def update_certification(
    employee_id: str,
    certification_id: int,
    cert_in: CertificationUpdate,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Update a certification for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 3 (Update)
    if not deps.has_permission(db, current_user, "3"):
        # Case 3: Self-access for update
        if not (isinstance(current_user, Employee) and current_user.id == employee.id):
             raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="You do not have permission to update certifications"
            )
            
    # 3. Resolve Certification
    cert = db.query(EmployeeCertification).filter(
        EmployeeCertification.id == certification_id,
        EmployeeCertification.employee_id == employee.id
    ).first()
    
    if not cert:
        raise HTTPException(status_code=404, detail="Certification record not found")
        
    # 4. Update
    update_data = cert_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(cert, field, value)
        
    try:
        db.commit()
        db.refresh(cert)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return CertificationResponse(
        success=True,
        message="Certification updated successfully",
        data=cert
    )

@router.get("/{employee_id}/certifications", response_model=CertificationListResponse)
def get_certifications(
    employee_id: str,
    is_active: Optional[bool] = None,
    expiring_soon: bool = False,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Get professional certifications for an employee.
    
    Query parameters:
    - is_active: Filter by active/inactive certifications
    - expiring_soon: Filter for certifications expiring within the next 30 days
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 1 (Read)
    if not deps.has_permission(db, current_user, "1"):
        # Case 3: Self-access for read
        if not (isinstance(current_user, Employee) and current_user.id == employee.id):
             raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="You do not have permission to view these certifications"
            )
            
    # 3. Query
    query = db.query(EmployeeCertification).filter(
        EmployeeCertification.employee_id == employee.id
    )
    
    if is_active is not None:
        query = query.filter(EmployeeCertification.is_active == is_active)
        
    if expiring_soon:
        today = datetime.utcnow().date()
        thirty_days_later = today + timedelta(days=30)
        query = query.filter(
            EmployeeCertification.expiry_date >= today,
            EmployeeCertification.expiry_date <= thirty_days_later
        )
        
    certifications = query.order_by(EmployeeCertification.issue_date.desc()).all()
    
    return CertificationListResponse(
        success=True,
        message="Certifications retrieved successfully",
        data=certifications
    )

@router.delete("/{employee_id}/certifications/{certification_id}", response_model=CertificationResponse)
def delete_certification(
    employee_id: str,
    certification_id: int,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Delete a certification for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 4 (Delete)
    if not deps.has_permission(db, current_user, "4"):
        # Case 3: Self-access for delete
        if not (isinstance(current_user, Employee) and current_user.id == employee.id):
             raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="You do not have permission to delete certification records"
            )
            
    # 3. Resolve Certification
    cert = db.query(EmployeeCertification).filter(
        EmployeeCertification.id == certification_id,
        EmployeeCertification.employee_id == employee.id
    ).first()
    
    if not cert:
        raise HTTPException(status_code=404, detail="Certification record not found")
        
    # 4. Delete
    try:
        db.delete(cert)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return CertificationResponse(
        success=True,
        message="Certification deleted successfully",
        data=None
    )

@router.post("/{employee_id}/work-experience", response_model=WorkExperienceResponse)
def create_work_experience(
    employee_id: str,
    work_in: WorkExperienceCreate,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Add a new work experience record for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 2 (Create)
    if not deps.has_permission(db, current_user, "2"):
        # Case 3: Self-access for create
        if not (isinstance(current_user, Employee) and current_user.id == employee.id):
             raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="You do not have permission to add work experience record for this employee"
            )
            
    # Create Work Experience
    db_obj = EmployeeWorkExperience(
        employee_id=employee.id,
        **work_in.model_dump()
    )
    db.add(db_obj)
    try:
        db.commit()
        db.refresh(db_obj)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return WorkExperienceResponse(
        success=True,
        message="Work experience created successfully",
        data=db_obj
    )

@router.put("/{employee_id}/work-experience/{experience_id}", response_model=WorkExperienceResponse)
def update_work_experience(
    employee_id: str,
    experience_id: int,
    work_in: WorkExperienceUpdate,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Update a work experience record for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 3 (Update)
    if not deps.has_permission(db, current_user, "3"):
        # Case 3: Self-access for update
        if not (isinstance(current_user, Employee) and current_user.id == employee.id):
             raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="You do not have permission to update work experience records"
            )
            
    # 3. Resolve Work Experience
    work_exp = db.query(EmployeeWorkExperience).filter(
        EmployeeWorkExperience.id == experience_id,
        EmployeeWorkExperience.employee_id == employee.id
    ).first()
    
    if not work_exp:
        raise HTTPException(status_code=404, detail="Work experience record not found")
        
    # 4. Update
    update_data = work_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(work_exp, field, value)
        
    try:
        db.commit()
        db.refresh(work_exp)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return WorkExperienceResponse(
        success=True,
        message="Work experience updated successfully",
        data=work_exp
    )

@router.get("/{employee_id}/work-experience", response_model=WorkExperienceListResponse)
def get_work_experience(
    employee_id: str,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Get previous work experience for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 1 (Read)
    if not deps.has_permission(db, current_user, "1"):
        # Case 3: Self-access for read
        if not (isinstance(current_user, Employee) and current_user.id == employee.id):
             raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="You do not have permission to view these work experience records"
            )
            
    # Query Work Experience
    work_experience = db.query(EmployeeWorkExperience).filter(
        EmployeeWorkExperience.employee_id == employee.id
    ).order_by(EmployeeWorkExperience.start_date.desc()).all()
    
    return WorkExperienceListResponse(
        success=True,
        message="Work experience records retrieved successfully",
        data=work_experience
    )

@router.delete("/{employee_id}/work-experience/{experience_id}", response_model=WorkExperienceResponse)
def delete_work_experience(
    employee_id: str,
    experience_id: int,
    db: Session = Depends(deps.get_db),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Delete a work experience record for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Access Control
    # Case 1 & 2: Org or Permission 4 (Delete)
    if not deps.has_permission(db, current_user, "4"):
        # Case 3: Self-access for delete
        if not (isinstance(current_user, Employee) and current_user.id == employee.id):
             raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, 
                detail="You do not have permission to delete work experience records"
            )
            
    # 3. Resolve Work Experience
    work_exp = db.query(EmployeeWorkExperience).filter(
        EmployeeWorkExperience.id == experience_id,
        EmployeeWorkExperience.employee_id == employee.id
    ).first()
    
    if not work_exp:
        raise HTTPException(status_code=404, detail="Work experience record not found")
        
    # 4. Delete
    try:
        db.delete(work_exp)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return WorkExperienceResponse(
        success=True,
        message="Work experience deleted successfully",
        data=None
    )

@router.get("/{employee_id}/reporting-structure", response_model=ReportingStructureResponse)
def get_reporting_structure(
    employee_id: str,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Get reporting hierarchy (upward chain and direct reports).
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).options(
             joinedload(Employee.job_title),
             joinedload(Employee.department)
        ).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).options(
                 joinedload(Employee.job_title),
                 joinedload(Employee.department)
            ).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).options(
             joinedload(Employee.job_title),
             joinedload(Employee.department)
        ).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. Build Upward Chain
    reporting_chain = []
    current = employee
    while current.reporting_manager_id:
        manager = db.query(Employee).options(
             joinedload(Employee.job_title),
             joinedload(Employee.department)
        ).filter(Employee.id == current.reporting_manager_id).first()
        
        if manager:
            reporting_chain.append(EmployeeSummarySchema.model_validate(manager))
            current = manager
        else:
            break
            
    # 3. Get Direct Reports (Downward)
    direct_reports = db.query(Employee).options(
         joinedload(Employee.job_title),
         joinedload(Employee.department)
    ).filter(
        Employee.reporting_manager_id == employee.id,
        Employee.is_active == True
    ).all()
    
    reports_data = [EmployeeSummarySchema.model_validate(r) for r in direct_reports]
    
    return ReportingStructureResponse(
        success=True,
        message="Reporting structure retrieved successfully",
        data=ReportingStructureData(
            employee=EmployeeSummarySchema.model_validate(employee),
            reporting_chain=reporting_chain,
            direct_reports=reports_data
        )
    )

@router.get("/{employee_id}/direct-reports", response_model=EmployeeListResponse)
def get_direct_reports(
    employee_id: str,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Get direct reportees for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # Get Direct Reports
    direct_reports = db.query(Employee).options(
         joinedload(Employee.job_title),
         joinedload(Employee.department),
         joinedload(Employee.location)
    ).filter(
        Employee.reporting_manager_id == employee.id,
        Employee.is_active == True,
        Employee.organization_id == current_org.id
    ).all()
    
    return EmployeeListResponse(
        success=True,
        message="Direct reports retrieved successfully",
        data=direct_reports
    )

@router.get("/{employee_id}/team-members", response_model=EmployeeListResponse)
def get_team_members(
    employee_id: str,
    include_indirect: bool = Query(False, description="Include indirect reportees (recursive)"),
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org)
):
    """
    Get all team members (direct and optional indirect reportees).
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    if not include_indirect:
        # Simple direct reports query
        team = db.query(Employee).options(
             joinedload(Employee.job_title),
             joinedload(Employee.department),
             joinedload(Employee.location)
        ).filter(
            Employee.reporting_manager_id == employee.id,
            Employee.is_active == True,
            Employee.organization_id == current_org.id
        ).all()
        return EmployeeListResponse(
            success=True,
            message="Team members retrieved successfully",
            data=team
        )
    
    # Recursive fetch
    # 1. Fetch all active employees in org
    all_emps = db.query(Employee).options(
         joinedload(Employee.job_title),
         joinedload(Employee.department),
         joinedload(Employee.location)
    ).filter(
        Employee.organization_id == current_org.id,
        Employee.is_active == True
    ).all()
    
    # 2. Build adjacency list: manager_id -> [employees]
    manager_map = {}
    for emp in all_emps:
        if emp.reporting_manager_id:
            if emp.reporting_manager_id not in manager_map:
                manager_map[emp.reporting_manager_id] = []
            manager_map[emp.reporting_manager_id].append(emp)
            
    # 3. BFS traversal
    team_members = []
    queue = [employee.id]
    
    # We use a set to avoid loops if any
    visited = set([employee.id])
    
    while queue:
        current_manager_id = queue.pop(0)
        if current_manager_id in manager_map:
            directs = manager_map[current_manager_id]
            for direct in directs:
                if direct.id not in visited:
                    visited.add(direct.id)
                    team_members.append(direct)
                    queue.append(direct.id)
                    
    return EmployeeListResponse(
        success=True,
        message=f"Team members retrieved successfully (Found {len(team_members)})",
        data=team_members
    )

@router.get("/{employee_uuid}/history", response_model=EmployeeHistoryListResponse)
def get_employee_history(
    employee_uuid: str,
    change_type: Optional[ChangeType] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Get change history for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        employee = db.query(Employee).filter(
            Employee.uuid == employee_uuid, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        raise HTTPException(status_code=404, detail="Employee not found")

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # 2. RBAC & Ownership Check
    is_manager = deps.has_permission(db, current_user, EmployeePermissions.READ)
    is_owner = isinstance(current_user, Employee) and current_user.id == employee.id
    
    if not (isinstance(current_user, Organization) or is_manager or is_owner):
        raise HTTPException(
            status_code=403, 
            detail="Access denied. You can only view your own history or requires permission."
        )
        
    # Build Query
    query = db.query(EmployeeHistory).filter(
        EmployeeHistory.employee_id == employee.id
    )
    
    if change_type:
        query = query.filter(EmployeeHistory.change_type == change_type)
        
    if from_date:
        query = query.filter(EmployeeHistory.effective_date >= from_date)
        
    if to_date:
        query = query.filter(EmployeeHistory.effective_date <= to_date)
        
    # Order by effective date desc
    query = query.order_by(EmployeeHistory.effective_date.desc(), EmployeeHistory.created_at.desc())
    
    # Pagination
    total_records = query.count()
    total_pages = (total_records + limit - 1) // limit
    
    offset = (page - 1) * limit
    
    history_records = query.options(
        joinedload(EmployeeHistory.previous_job_title),
        joinedload(EmployeeHistory.new_job_title),
        joinedload(EmployeeHistory.previous_department),
        joinedload(EmployeeHistory.new_department),
        joinedload(EmployeeHistory.previous_location),
        joinedload(EmployeeHistory.new_location),
        joinedload(EmployeeHistory.previous_reporting_manager),
        joinedload(EmployeeHistory.new_reporting_manager)
    ).offset(offset).limit(limit).all()
    
    return EmployeeHistoryListResponse(
        success=True,
        message="History records retrieved successfully",
        data=history_records,
        pagination=PaginationData(
            total_records=total_records,
            current_page=page,
            total_pages=total_pages,
            page_size=limit
        )
    )

@router.post("/{employee_uuid}/history", response_model=EmployeeHistoryResponse)
def create_employee_history(
    employee_uuid: str,
    history_in: EmployeeHistoryCreate,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Record a new change history for an employee.
    If effective_date is today or earlier, optionally update the Employee record? (Not implemented here to keep side effects explicit, user should update Employee separately or we should have a 'promote' endpoint)
    For this 'record history' endpoint, we just insert the record.
    """
    # 1. Resolve Employee
    employee = None
    try:
        employee = db.query(Employee).filter(
            Employee.uuid == employee_uuid, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            employee = db.query(Employee).filter(Employee.uuid == employee_uuid, Employee.organization_id == current_org.id).first()
         except ValueError:
            raise HTTPException(status_code=404, detail="Employee not found")
        
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # 2. RBAC Check
    if not deps.has_permission(db, current_user, EmployeePermissions.CREATE):
        raise HTTPException(
            status_code=403, 
            detail="You do not have permission to record employee history"
        )

    # Helper to resolve UUID to ID
    def resolve_id(model, uuid_val):
        if not uuid_val:
            return None
        obj = db.query(model).filter(model.uuid == uuid_val, model.organization_id == current_org.id).first()
        return obj.id if obj else None

    # Resolve FKs
    data = history_in.model_dump(exclude={
        'previous_job_title_uuid', 'new_job_title_uuid',
        'previous_department_uuid', 'new_department_uuid',
        'previous_location_uuid', 'new_location_uuid',
        'previous_reporting_manager_uuid', 'new_reporting_manager_uuid'
    })

    data['previous_job_title_id'] = resolve_id(JobTitle, history_in.previous_job_title_uuid)
    data['new_job_title_id'] = resolve_id(JobTitle, history_in.new_job_title_uuid)
    
    data['previous_department_id'] = resolve_id(Department, history_in.previous_department_uuid)
    data['new_department_id'] = resolve_id(Department, history_in.new_department_uuid)
    
    data['previous_location_id'] = resolve_id(Location, history_in.previous_location_uuid)
    data['new_location_id'] = resolve_id(Location, history_in.new_location_uuid)
    
    data['previous_reporting_manager_id'] = resolve_id(Employee, history_in.previous_reporting_manager_uuid)
    data['new_reporting_manager_id'] = resolve_id(Employee, history_in.new_reporting_manager_uuid)
        
    # Create History
    db_obj = EmployeeHistory(
        employee_id=employee.id,
        **data
    )
    db.add(db_obj)
    try:
        db.commit()
        db.refresh(db_obj)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return EmployeeHistoryResponse(
        success=True,
        message="History record created successfully",
        data=db_obj
    )

@router.get("/{employee_uuid}/history/{history_id}", response_model=EmployeeHistoryResponse)
def get_employee_history_detail(
    employee_uuid: str,
    history_id: int,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Get specific history record.
    """
    # 1. Resolve Employee
    employee = None
    try:
        employee = db.query(Employee).filter(
            Employee.uuid == employee_uuid, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            employee = db.query(Employee).filter(Employee.uuid == employee_uuid, Employee.organization_id == current_org.id).first()
         except ValueError:
            raise HTTPException(status_code=404, detail="Employee not found")

    if not employee:
         raise HTTPException(status_code=404, detail="Employee not found")
         
    # 2. RBAC & Ownership Check
    is_manager = deps.has_permission(db, current_user, EmployeePermissions.READ)
    is_owner = isinstance(current_user, Employee) and current_user.id == employee.id
    
    if not (isinstance(current_user, Organization) or is_manager or is_owner):
        raise HTTPException(
            status_code=403, 
            detail="Access denied. You can only view your own history or requires permission."
        )
         
    # 2. Get History Record
    history_record = db.query(EmployeeHistory).options(
        joinedload(EmployeeHistory.previous_job_title),
        joinedload(EmployeeHistory.new_job_title),
        joinedload(EmployeeHistory.previous_department),
        joinedload(EmployeeHistory.new_department),
        joinedload(EmployeeHistory.previous_location),
        joinedload(EmployeeHistory.new_location),
        joinedload(EmployeeHistory.previous_reporting_manager),
        joinedload(EmployeeHistory.new_reporting_manager)
    ).filter(
        EmployeeHistory.id == history_id,
        EmployeeHistory.employee_id == employee.id
    ).first()
    
    return EmployeeHistoryResponse(
        success=True,
        message="History record retrieved successfully",
        data=history_record
    )
    
@router.put("/{employee_uuid}/history/{history_id}", response_model=EmployeeHistoryResponse)
def update_employee_history(
    employee_uuid: str,
    history_id: int,
    history_in: EmployeeHistoryUpdate,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Update an existing history record.
    """
    # 1. Resolve Employee
    employee = None
    try:
        employee = db.query(Employee).filter(
            Employee.uuid == employee_uuid, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # 2. RBAC Check (Permission 3: UPDATE)
    if not deps.has_permission(db, current_user, "3"):
         raise HTTPException(
            status_code=403, 
            detail="You do not have permission to update employee history"
        )
        
    # 3. Get History Record
    history_record = db.query(EmployeeHistory).filter(
        EmployeeHistory.id == history_id,
        EmployeeHistory.employee_id == employee.id
    ).first()
    
    if not history_record:
        raise HTTPException(status_code=404, detail="History record not found")
        
    # 4. Resolve UUIDs to IDs
    def resolve_id(model, uuid_val):
        if not uuid_val:
            return None
        obj = db.query(model).filter(model.uuid == uuid_val, model.organization_id == current_org.id).first()
        return obj.id if obj else None

    update_data = history_in.model_dump(exclude_unset=True)
    
    # Map UUID fields to ID fields
    uuid_map = {
        'previous_job_title_uuid': ('previous_job_title_id', JobTitle),
        'new_job_title_uuid': ('new_job_title_id', JobTitle),
        'previous_department_uuid': ('previous_department_id', Department),
        'new_department_uuid': ('new_department_id', Department),
        'previous_location_uuid': ('previous_location_id', Location),
        'new_location_uuid': ('new_location_id', Location),
        'previous_reporting_manager_uuid': ('previous_reporting_manager_id', Employee),
        'new_reporting_manager_uuid': ('new_reporting_manager_id', Employee),
    }
    
    for uuid_field, (id_field, model) in uuid_map.items():
        if uuid_field in update_data:
            update_data[id_field] = resolve_id(model, update_data.pop(uuid_field))
            
    # Update Fields
    for field, value in update_data.items():
        setattr(history_record, field, value)
        
    try:
        db.commit()
        db.refresh(history_record)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return EmployeeHistoryResponse(
        success=True,
        message="History record updated successfully",
        data=history_record
    )

@router.get("/{employee_uuid}/promotions", response_model=EmployeeHistoryListResponse)
def get_employee_promotions(
    employee_uuid: str,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Get promotion history for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        employee = db.query(Employee).filter(
            Employee.uuid == employee_uuid, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            employee = db.query(Employee).filter(Employee.uuid == employee_uuid, Employee.organization_id == current_org.id).first()
         except ValueError:
            raise HTTPException(status_code=404, detail="Employee not found")

    if not employee:
         raise HTTPException(status_code=404, detail="Employee not found")
         
    # 2. RBAC & Ownership Check
    is_manager = deps.has_permission(db, current_user, EmployeePermissions.READ)
    is_owner = isinstance(current_user, Employee) and current_user.id == employee.id
    
    if not (isinstance(current_user, Organization) or is_manager or is_owner):
        raise HTTPException(
            status_code=403, 
            detail="Access denied. You can only view your own history or requires permission."
        )
         
    # Query Promotions
    promotions = db.query(EmployeeHistory).options(
        joinedload(EmployeeHistory.previous_job_title),
        joinedload(EmployeeHistory.new_job_title),
        joinedload(EmployeeHistory.previous_department),
        joinedload(EmployeeHistory.new_department)
    ).filter(
        EmployeeHistory.employee_id == employee.id,
        EmployeeHistory.change_type == ChangeType.PROMOTION
    ).order_by(EmployeeHistory.effective_date.desc()).all()
    
    return EmployeeHistoryListResponse(
        success=True,
        message="Promotion history retrieved successfully",
        data=promotions
    )

@router.get("/{employee_uuid}/transfers", response_model=EmployeeHistoryListResponse)
def get_employee_transfers(
    employee_uuid: str,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Get transfer history (Transfer, Department Change, Location Change).
    """
    # 1. Resolve Employee
    employee = None
    try:
        employee = db.query(Employee).filter(
            Employee.uuid == employee_uuid, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            employee = db.query(Employee).filter(Employee.uuid == employee_uuid, Employee.organization_id == current_org.id).first()
         except ValueError:
            raise HTTPException(status_code=404, detail="Employee not found")

    if not employee:
         raise HTTPException(status_code=404, detail="Employee not found")
         
    # 2. RBAC & Ownership Check
    is_manager = deps.has_permission(db, current_user, EmployeePermissions.READ)
    is_owner = isinstance(current_user, Employee) and current_user.id == employee.id
    
    if not (isinstance(current_user, Organization) or is_manager or is_owner):
        raise HTTPException(
            status_code=403, 
            detail="Access denied. You can only view your own history or requires permission."
        )
         
    # Query Transfers (Broad definition)
    transfer_types = [ChangeType.TRANSFER, ChangeType.DEPARTMENT_CHANGE, ChangeType.LOCATION_CHANGE]
    
    transfers = db.query(EmployeeHistory).options(
        joinedload(EmployeeHistory.previous_job_title),
        joinedload(EmployeeHistory.new_job_title),
        joinedload(EmployeeHistory.previous_department),
        joinedload(EmployeeHistory.new_department)
    ).filter(
        EmployeeHistory.employee_id == employee.id,
        EmployeeHistory.change_type.in_(transfer_types)
    ).order_by(EmployeeHistory.effective_date.desc()).all()
    
    return EmployeeHistoryListResponse(
        success=True,
        message="Transfer history retrieved successfully",
        data=transfers
    )

@router.get("/{employee_uuid}/salary-history", response_model=EmployeeHistoryListResponse)
def get_employee_salary_history(
    employee_uuid: str,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Get salary revision history.
    """
    # 1. Resolve Employee
    employee = None
    try:
        employee = db.query(Employee).filter(
            Employee.uuid == employee_uuid, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            employee = db.query(Employee).filter(Employee.uuid == employee_uuid, Employee.organization_id == current_org.id).first()
         except ValueError:
            raise HTTPException(status_code=404, detail="Employee not found")

    if not employee:
         raise HTTPException(status_code=404, detail="Employee not found")
         
    # 2. RBAC & Ownership Check
    is_manager = deps.has_permission(db, current_user, EmployeePermissions.READ)
    is_owner = isinstance(current_user, Employee) and current_user.id == employee.id
    
    if not (isinstance(current_user, Organization) or is_manager or is_owner):
        raise HTTPException(
            status_code=403, 
            detail="Access denied. You can only view your own history or requires permission."
        )
         
    # Query Salary Revisions
    revisions = db.query(EmployeeHistory).filter(
        EmployeeHistory.employee_id == employee.id,
        EmployeeHistory.change_type == ChangeType.SALARY_REVISION
    ).order_by(EmployeeHistory.effective_date.desc()).all()
    
    return EmployeeHistoryListResponse(
        success=True,
        message="Salary history retrieved successfully",
        data=revisions
    )

@router.post("/{employee_id}/documents", response_model=EmployeeDocumentResponse)
def create_employee_document(
    employee_id: str,
    file: UploadFile = File(...),
    document_type: DocumentType = Form(...),
    document_name: str = Form(...),
    document_number: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    issue_date: Optional[date] = Form(None),
    expiry_date: Optional[date] = Form(None),
    issuing_authority: Optional[str] = Form(None),
    request: Request = None,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Upload a new employee document.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # 2. RBAC & Ownership Check
    is_manager = deps.has_permission(db, current_user, "2")
    is_owner = isinstance(current_user, Employee) and current_user.id == employee.id
    
    if not (isinstance(current_user, Organization) or is_manager or is_owner):
        raise HTTPException(
            status_code=403, 
            detail="You do not have permission to add documents for this employee"
        )
        
    # 2. Save File
    # Path: uploads/{org_uuid}/employee/{emp_uuid}/documents/{filename}
    upload_dir = os.path.join("uploads", str(current_org.uuid), "employee", str(employee.uuid), "documents")
    os.makedirs(upload_dir, exist_ok=True)
    
    file_ext = os.path.splitext(file.filename)[1]
    safe_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = os.path.join(upload_dir, safe_filename)
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    file_size_kb = os.path.getsize(file_path) // 1024
    
    # URL generation
    base_url = str(request.base_url).rstrip("/") if request else ""
    file_url = f"{base_url}/static/{current_org.uuid}/employee/{employee.uuid}/documents/{safe_filename}"
    
    # 3. Create DB Record
    db_obj = EmployeeDocument(
        employee_id=employee.id,
        document_type=document_type,
        document_name=document_name,
        document_number=document_number,
        description=description,
        file_name=file.filename,
        file_url=file_url,
        file_size_kb=file_size_kb,
        mime_type=file.content_type,
        issue_date=issue_date,
        expiry_date=expiry_date,
        issuing_authority=issuing_authority
    )
    db.add(db_obj)
    try:
        db.commit()
        db.refresh(db_obj)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return EmployeeDocumentResponse(
        success=True,
        message="Document uploaded successfully",
        data=db_obj
    )



@router.get("/{employee_id}/documents", response_model=EmployeeDocumentListResponse)
def get_employee_documents(
    employee_id: str,
    document_type: Optional[DocumentType] = None,
    is_verified: Optional[bool] = None,
    expiring_soon: Optional[bool] = Query(None, description="Filter documents expiring in 30 days"),
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    List all documents for an employee.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # 2. RBAC & Ownership Check
    is_manager = deps.has_permission(db, current_user, "1")
    is_owner = isinstance(current_user, Employee) and current_user.id == employee.id
    
    if not (isinstance(current_user, Organization) or is_manager or is_owner):
        raise HTTPException(
            status_code=403, 
            detail="You do not have permission to view documents for this employee"
        )
        
    # Query Docs
    query = db.query(EmployeeDocument).filter(
        EmployeeDocument.employee_id == employee.id
    )
    
    if document_type:
        query = query.filter(EmployeeDocument.document_type == document_type)
        
    if is_verified is not None:
        query = query.filter(EmployeeDocument.is_verified == is_verified)
        
    if expiring_soon:
        today = date.today()
        thirty_days = today + timedelta(days=30)
        query = query.filter(
            EmployeeDocument.expiry_date != None,
            EmployeeDocument.expiry_date >= today,
            EmployeeDocument.expiry_date <= thirty_days
        )
        
    docs = query.order_by(EmployeeDocument.created_at.desc()).all()
    
    return EmployeeDocumentListResponse(
        success=True,
        message="Documents retrieved successfully",
        data=docs
    )

@router.get("/{employee_id}/documents/{document_uuid}", response_model=EmployeeDocumentResponse)
def get_employee_document_detail(
    employee_id: str,
    document_uuid: str,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Get document details.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # Get Document
    doc = None
    try:
        doc_uuid_obj = uuid.UUID(document_uuid)
        doc = db.query(EmployeeDocument).filter(
            EmployeeDocument.uuid == doc_uuid_obj,
            EmployeeDocument.employee_id == employee.id
        ).first()
    except ValueError:
        pass
        
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
        
    return EmployeeDocumentResponse(
        success=True,
        message="Document details retrieved successfully",
        data=doc
    )

@router.get("/{employee_id}/documents/{document_uuid}/download")
def download_employee_document(
    employee_id: str,
    document_uuid: str,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Download a document.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # Get Document
    doc = None
    try:
        doc_uuid_obj = uuid.UUID(document_uuid)
        doc = db.query(EmployeeDocument).filter(
            EmployeeDocument.uuid == doc_uuid_obj,
            EmployeeDocument.employee_id == employee.id
        ).first()
    except ValueError:
        pass
        
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
        
    # 2. RBAC & Ownership Check
    is_manager = deps.has_permission(db, current_user, "1")
    is_owner = isinstance(current_user, Employee) and current_user.id == employee.id
    
    if not (isinstance(current_user, Organization) or is_manager or is_owner):
        raise HTTPException(
            status_code=403, 
            detail="You do not have permission to download this document"
        )
        
    # Construct File Path
    if not doc.file_url:
        raise HTTPException(status_code=404, detail="File URL missing")
        
    safe_filename = os.path.basename(doc.file_url)
    file_path = os.path.join("uploads", str(current_org.uuid), "employee", str(employee.uuid), "documents", safe_filename)
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File file not found on server")
        
    return FileResponse(
        path=file_path,
        filename=doc.file_name, # Serve with original name
        media_type=doc.mime_type
    )

@router.put("/{employee_id}/documents/{document_uuid}", response_model=EmployeeDocumentResponse)
def update_employee_document(
    employee_id: str,
    document_uuid: str,
    document_in: EmployeeDocumentUpdate,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Update document metadata.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # Get Document
    doc = None
    try:
        doc_uuid_obj = uuid.UUID(document_uuid)
        doc = db.query(EmployeeDocument).filter(
            EmployeeDocument.uuid == doc_uuid_obj,
            EmployeeDocument.employee_id == employee.id
        ).first()
    except ValueError:
        pass
        
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
        
    # 2. RBAC & Ownership Check
    is_manager = deps.has_permission(db, current_user, "3")
    is_owner = isinstance(current_user, Employee) and current_user.id == employee.id
    
    if not (isinstance(current_user, Organization) or is_manager or is_owner):
        raise HTTPException(
            status_code=403, 
            detail="You do not have permission to update this document"
        )
    
    # Verification Lock for Owner
    if is_owner and not (isinstance(current_user, Organization) or is_manager):
        if doc.is_verified:
            raise HTTPException(
                status_code=403, 
                detail="Verified documents cannot be updated by the owner. Please contact HR."
            )
        
    # Update
    update_data = document_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(doc, field, value)
        
    try:
        db.commit()
        db.refresh(doc)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return EmployeeDocumentResponse(
        success=True,
        message="Document updated successfully",
        data=doc
    )

@router.delete("/{employee_id}/documents/{document_uuid}", response_model=EmployeeDocumentResponse)
def delete_employee_document(
    employee_id: str,
    document_uuid: str,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Delete a document.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    # Get Document
    doc = None
    try:
        doc_uuid_obj = uuid.UUID(document_uuid)
        doc = db.query(EmployeeDocument).filter(
            EmployeeDocument.uuid == doc_uuid_obj,
            EmployeeDocument.employee_id == employee.id
        ).first()
    except ValueError:
        pass
        
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
        
    # 2. RBAC & Ownership Check
    is_manager = deps.has_permission(db, current_user, "4")
    is_owner = isinstance(current_user, Employee) and current_user.id == employee.id
    
    if not (isinstance(current_user, Organization) or is_manager or is_owner):
        raise HTTPException(
            status_code=403, 
            detail="You do not have permission to delete this document"
        )
    
    # Verification Lock for Owner
    if is_owner and not (isinstance(current_user, Organization) or is_manager):
        if doc.is_verified:
            raise HTTPException(
                status_code=403, 
                detail="Verified documents cannot be deleted by the owner. Please contact HR."
            )

    # Delete File from Disk
    if doc.file_url:
        safe_filename = os.path.basename(doc.file_url)
        file_path = os.path.join("uploads", str(current_org.uuid), "employee", str(employee.uuid), "documents", safe_filename)
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except OSError:
                pass
                
    # Delete Record
    db.delete(doc)
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return EmployeeDocumentResponse(
        success=True,
        message="Document deleted successfully",
        data=None
    )
@router.patch("/{employee_id}/documents/{document_uuid}/verify", response_model=EmployeeDocumentResponse)
def verify_employee_document(
    employee_id: str,
    document_uuid: str,
    verification_in: DocumentVerification,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Verify a document.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # 2. RBAC & Owner Check
    # Only Organizations or Managers with Perm Code 3 (Update/Manage) can verify
    is_manager = deps.has_permission(db, current_user, "3")
    
    if not (isinstance(current_user, Organization) or is_manager):
        raise HTTPException(
            status_code=403, 
            detail="You do not have permission to verify documents"
        )
    
    # Self-verification restriction
    if isinstance(current_user, Employee) and current_user.id == employee.id:
        raise HTTPException(
            status_code=403, 
            detail="You cannot verify your own documents"
        )

    # Get Document
    doc = None
    try:
        doc_uuid_obj = uuid.UUID(document_uuid)
        doc = db.query(EmployeeDocument).filter(
            EmployeeDocument.uuid == doc_uuid_obj,
            EmployeeDocument.employee_id == employee.id
        ).first()
    except ValueError:
        pass
        
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
        
    # Verify
    doc.is_verified = verification_in.is_verified
    doc.verification_notes = verification_in.verification_notes
    if verification_in.is_verified:
        doc.verified_at = datetime.utcnow()
    else:
        doc.verified_at = None
        
    try:
        db.commit()
        db.refresh(doc)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return EmployeeDocumentResponse(
        success=True,
        message="Document verified successfully",
        data=doc
    )

@router.get("/{employee_id}/active-delegations", response_model=ApprovalDelegationListResponse)
def get_employee_active_delegations(
    employee_id: str,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    date: Optional[date] = Query(None, description="Check delegations active on this date"),
    delegation_date: Optional[date] = Query(None, description="Check delegations active on this date (alternative param name)"),
    delegation_type: Optional[DelegationType] = Query(None, description="Filter by delegation type")
):
    """
    Get active delegations for an employee where they are either the delegator 
    or the delegate recipient.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:
        try:
            employee = db.query(Employee).filter(
                Employee.id == int(employee_id), Employee.organization_id == current_org.id
            ).first()
        except ValueError:
            pass
            
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # 2. Query Active Delegations
    query = db.query(ApprovalDelegation).options(
        joinedload(ApprovalDelegation.delegator),
        joinedload(ApprovalDelegation.delegate_to)
    ).filter(
        ApprovalDelegation.organization_id == current_org.id,
        ApprovalDelegation.is_active == True,
        or_(
            ApprovalDelegation.delegator_id == employee.id,
            ApprovalDelegation.delegate_to_id == employee.id
        )
    )

    # Resolve which date to use (prioritize 'date', then 'delegation_date')
    target_date = date or delegation_date

    if target_date:
        query = query.filter(
            ApprovalDelegation.from_date <= target_date,
            ApprovalDelegation.to_date >= target_date
        )

    if delegation_type:
        query = query.filter(ApprovalDelegation.delegation_type == delegation_type)

    delegations = query.all()

    # 3. Map to schema
    from app.api.v1.endpoints.approval_delegations import map_delegation_response

    data = [map_delegation_response(d) for d in delegations]

    return ApprovalDelegationListResponse(
        success=True,
        message="Active delegations retrieved successfully",
        data=data,
        pagination={
            "total_records": len(data),
            "current_page": 1,
            "total_pages": 1,
            "page_size": len(data)
        }
    )

@router.post("/{employee_id}/profile-picture", response_model=EmployeeResponse)
def upload_employee_profile_picture(
    employee_id: str,
    file: UploadFile = File(...),
    request: Request = None,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Upload or update employee profile picture.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # 2. RBAC & Ownership Check
    is_manager = deps.has_permission(db, current_user, "3")
    is_owner = isinstance(current_user, Employee) and current_user.id == employee.id
    
    if not (isinstance(current_user, Organization) or is_manager or is_owner):
        raise HTTPException(
            status_code=403, 
            detail="You do not have permission to update profile picture for this employee"
        )
        
    # 3. Save File
    upload_dir = os.path.join("uploads", str(current_org.uuid), "employee", str(employee.uuid), "profile")
    os.makedirs(upload_dir, exist_ok=True)
    
    import mimetypes
    file_ext = os.path.splitext(file.filename)[1]
    if not file_ext and file.content_type:
        file_ext = mimetypes.guess_extension(file.content_type) or ""
        
    safe_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = os.path.join(upload_dir, safe_filename)
    
    # Remove old picture if exists
    if employee.photograph_url and "/static/" in employee.photograph_url:
        parts = employee.photograph_url.split("/static/")
        if len(parts) == 2:
            old_file_path = os.path.join("uploads", parts[1])
            if os.path.exists(old_file_path):
                try:
                    os.remove(old_file_path)
                except OSError:
                    pass
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    # URL generation
    base_url = str(request.base_url).rstrip("/") if request else ""
    file_url = f"{base_url}/static/{current_org.uuid}/employee/{employee.uuid}/profile/{safe_filename}"
    
    # 4. Update DB Record
    employee.photograph_url = file_url
    
    try:
        db.commit()
        db.refresh(employee)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return EmployeeResponse(
        success=True,
        message="Profile picture updated successfully",
        data=employee
    )

@router.delete("/{employee_id}/profile-picture", response_model=EmployeeResponse)
def delete_employee_profile_picture(
    employee_id: str,
    db: Session = Depends(deps.get_db),
    current_org: Organization = Depends(deps.get_current_org),
    current_user: Union[Organization, Employee] = Depends(deps.get_current_user)
):
    """
    Delete employee profile picture.
    """
    # 1. Resolve Employee
    employee = None
    try:
        uuid_obj = uuid.UUID(employee_id)
        employee = db.query(Employee).filter(
            Employee.uuid == uuid_obj, Employee.organization_id == current_org.id
        ).first()
    except ValueError:
        pass
    
    if not employee:   
         try:
            int_id = int(employee_id)
            employee = db.query(Employee).filter(Employee.id == int_id, Employee.organization_id == current_org.id).first()
         except ValueError:
             pass
             
    if not employee:
        employee = db.query(Employee).filter(Employee.employee_code == employee_id, Employee.organization_id == current_org.id).first()
        
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # 2. RBAC & Ownership Check
    is_manager = deps.has_permission(db, current_user, "3")
    is_owner = isinstance(current_user, Employee) and current_user.id == employee.id
    
    if not (isinstance(current_user, Organization) or is_manager or is_owner):
        raise HTTPException(
            status_code=403, 
            detail="You do not have permission to delete profile picture for this employee"
        )
        
    # 3. Delete File from Disk
    if employee.photograph_url and "/static/" in employee.photograph_url:
        parts = employee.photograph_url.split("/static/")
        if len(parts) == 2:
            file_path = os.path.join("uploads", parts[1])
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except OSError:
                    pass
                    
    # 4. Update DB Record
    employee.photograph_url = None
    
    try:
        db.commit()
        db.refresh(employee)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return EmployeeResponse(
        success=True,
        message="Profile picture deleted successfully",
        data=employee
    )

